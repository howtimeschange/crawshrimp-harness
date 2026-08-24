"""Buyer-show AI workflow packaging and generation helpers."""
from __future__ import annotations

import re
import shutil
import time
import zipfile
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Mapping, Optional
from urllib.parse import parse_qs, urlparse

from core import ai_image_service, data_sink


BUYER_SHOW_TASK_ID = "buyer_show_ai_generate"
DEFAULT_IMAGE_SIZE = "2160x2880"
DEFAULT_OUTPUT_FORMAT = "png"
DEFAULT_QUALITY = "high"
DEFAULT_GENERATION_CONCURRENCY = 5
MAX_GENERATION_CONCURRENCY = 10
DEFAULT_GENERATION_ATTEMPTS = 3
MAX_GENERATION_ATTEMPTS = 3
DEFAULT_RESULT_DOWNLOAD_CONCURRENCY = 10
MAX_RESULT_DOWNLOAD_CONCURRENCY = 20
DEFAULT_RESULT_DOWNLOAD_ATTEMPTS_PER_URL = 3
RESUME_EXECUTE_MODES = {"resume", "recover", "resume-recover", "resume_recover"}
IMAGE_OUTPUT_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp"}
BUYER_SHOW_MODEL_OPTIONS = {
    "gpt-image-2k": {
        "model_id": "gpt-image-2k",
        "model": "gpt-image-2",
        "model_key_tier": "2k",
        "provider_family": "gpt-image",
    },
    "gpt-image-4k": {
        "model_id": "gpt-image-4k",
        "model": "gpt-image-2",
        "model_key_tier": "4k",
        "provider_family": "gpt-image",
    },
    "gemini-3.1-flash-image-preview": {
        "model_id": "gemini-3.1-flash-image-preview",
        "model": "gemini-3.1-flash-image-preview",
        "model_key_tier": "ai.1xm.gemini_3_1_flash_image_preview_key",
        "provider_family": "nano-banana",
        "resolution": "1K",
    },
    "gemini-3-pro-image-preview": {
        "model_id": "gemini-3-pro-image-preview",
        "model": "gemini-3-pro-image-preview",
        "model_key_tier": "ai.1xm.gemini_3_pro_image_preview_key",
        "provider_family": "nano-banana",
        "resolution": "2K",
    },
}
DEFAULT_BUYER_SHOW_MODEL_ID = "gpt-image-4k"
BUYER_SHOW_SIZE_OPTIONS = [
    {"ratio": "1:1", "size_2k": "2048x2048", "size_4k": "2880x2880", "width": 1, "height": 1},
    {"ratio": "3:4", "size_2k": "1536x2048", "size_4k": "2448x3264", "width": 3, "height": 4},
    {"ratio": "4:3", "size_2k": "2048x1536", "size_4k": "3264x2448", "width": 4, "height": 3},
    {"ratio": "4:5", "size_2k": "1536x1920", "size_4k": "2560x3200", "width": 4, "height": 5},
    {"ratio": "3:2", "size_2k": "1536x1024", "size_4k": "3504x2336", "width": 3, "height": 2},
    {"ratio": "2:3", "size_2k": "1024x1536", "size_4k": "2336x3504", "width": 2, "height": 3},
    {"ratio": "16:9", "size_2k": "2048x1152", "size_4k": "3840x2160", "width": 16, "height": 9},
    {"ratio": "9:16", "size_2k": "1152x2048", "size_4k": "2160x3840", "width": 9, "height": 16},
]

SUMMARY_COLUMNS = [
    "表格行号",
    "订单号",
    "款色号",
    "货号",
    "尺码",
    "唯一值",
    "AI素材库路径",
    "AI图包文件夹命名",
    "存放地址",
    "模拍文件",
    "模拍云盘路径",
    "模拍下载结果",
    "模拍本地文件",
    "平铺参考图",
    "平铺云盘路径",
    "平铺下载结果",
    "平铺本地文件",
    "下载素材文件",
    "AI任务ID",
    "1XM任务ID",
    "生图结果",
    "生图文件",
    "本地图包文件夹",
    "备注",
]

USAGE_COLUMNS = [
    "记录时间",
    "货号",
    "款色号",
    "唯一值",
    "源表行号",
    "模拍文件",
    "模拍云盘路径",
    "平铺云盘路径",
    "AI任务ID",
    "1XM任务ID",
    "输出文件",
]


def _compact(value: Any) -> str:
    return str(value or "").replace("\u3000", " ").strip()


def _safe_local_name(value: Any, fallback: str = "item") -> str:
    text = _compact(value)
    text = re.sub(r"[\\/:*?\"<>|]+", "_", text)
    text = re.sub(r"\s+", " ", text).strip(" ._")
    return text[:120].strip(" ._") or fallback


def _append_note(current: Any, addition: Any) -> str:
    left = _compact(current)
    right = _compact(addition)
    if not left:
        return right
    if not right:
        return left
    return f"{left}；{right}"


def _extract_style_code(style_color_code: Any) -> str:
    code = re.sub(r"\s+", "", str(style_color_code or "")).strip()
    if not code:
        return ""
    if "-" in code:
        return code.split("-", 1)[0].strip() or code
    if re.fullmatch(r"[A-Za-z0-9]+", code) and len(code) >= 17:
        return code[:-5]
    return code


def _ensure_unique_dir(path: Path) -> Path:
    if not path.exists():
        path.mkdir(parents=True, exist_ok=True)
        return path
    for index in range(2, 10000):
        candidate = path.with_name(f"{path.name}_{index}")
        if not candidate.exists():
            candidate.mkdir(parents=True, exist_ok=True)
            return candidate
    raise RuntimeError(f"Cannot allocate unique directory for {path}")


def _ensure_unique_path(path: Path) -> Path:
    if not path.exists():
        path.parent.mkdir(parents=True, exist_ok=True)
        return path
    for index in range(2, 10000):
        candidate = path.with_name(f"{path.stem}_{index}{path.suffix}")
        if not candidate.exists():
            candidate.parent.mkdir(parents=True, exist_ok=True)
            return candidate
    raise RuntimeError(f"Cannot allocate unique file name for {path.name}")


def _copy_file_to_unique_target(source: Path, target: Path) -> Path:
    destination = _ensure_unique_path(target)
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)
    return destination


def _image_suffix_from_file(source_path: Path, fallback: str = ".png") -> str:
    try:
        header = source_path.read_bytes()[:16]
    except FileNotFoundError:
        header = b""
    if header.startswith(b"\x89PNG\r\n\x1a\n"):
        return ".png"
    if header.startswith(b"\xff\xd8\xff"):
        return ".jpg"
    if header[:4] == b"RIFF" and header[8:12] == b"WEBP":
        return ".webp"
    suffix = source_path.suffix.lower()
    if suffix == ".jpeg":
        return ".jpg"
    if suffix in {".png", ".jpg", ".webp"}:
        return suffix
    return fallback


def _parse_image_dimensions_from_header(source_path: Path) -> Optional[tuple[int, int]]:
    try:
        header = source_path.read_bytes()[:512 * 1024]
    except OSError:
        return None
    if len(header) >= 24 and header.startswith(b"\x89PNG\r\n\x1a\n"):
        width = int.from_bytes(header[16:20], "big")
        height = int.from_bytes(header[20:24], "big")
        if width > 0 and height > 0:
            return width, height
    if len(header) >= 12 and header[:4] == b"RIFF" and header[8:12] == b"WEBP":
        try:
            from PIL import Image

            with Image.open(source_path) as image:
                return int(image.width), int(image.height)
        except Exception:
            return None
    if len(header) >= 4 and header.startswith(b"\xff\xd8"):
        index = 2
        while index + 9 < len(header):
            if header[index] != 0xFF:
                index += 1
                continue
            while index < len(header) and header[index] == 0xFF:
                index += 1
            if index >= len(header):
                break
            marker = header[index]
            index += 1
            if marker in {0x01, *range(0xD0, 0xD9)}:
                continue
            if index + 2 > len(header):
                break
            length = int.from_bytes(header[index:index + 2], "big")
            if length < 2 or index + length > len(header):
                break
            if marker in {0xC0, 0xC1, 0xC2, 0xC3, 0xC5, 0xC6, 0xC7, 0xC9, 0xCA, 0xCB, 0xCD, 0xCE, 0xCF}:
                height = int.from_bytes(header[index + 3:index + 5], "big")
                width = int.from_bytes(header[index + 5:index + 7], "big")
                if width > 0 and height > 0:
                    return width, height
                break
            index += length
    return None


def _read_image_dimensions(source_path: Path) -> Optional[tuple[int, int]]:
    try:
        from PIL import Image

        with Image.open(source_path) as image:
            return int(image.width), int(image.height)
    except Exception:
        return _parse_image_dimensions_from_header(source_path)


def _resolve_buyer_show_model_params(run_params: Mapping[str, Any]) -> dict:
    model_id = _compact(
        run_params.get("model_id")
        or run_params.get("ai_model_id")
        or run_params.get("buyer_show_model_id")
    )
    if model_id in BUYER_SHOW_MODEL_OPTIONS:
        return dict(BUYER_SHOW_MODEL_OPTIONS[model_id])

    model = _compact(run_params.get("model") or run_params.get("model_key"))
    if model in BUYER_SHOW_MODEL_OPTIONS:
        return dict(BUYER_SHOW_MODEL_OPTIONS[model])
    if model in {
        "gemini-3.1-flash-image-preview",
        "gemini-3-pro-image-preview",
    }:
        return dict(BUYER_SHOW_MODEL_OPTIONS[model])

    tier = _compact(run_params.get("model_key_tier") or run_params.get("one_xm_key_tier")).lower()
    if model == "gpt-image-2" and tier in {"2k", "4k"}:
        return dict(BUYER_SHOW_MODEL_OPTIONS[f"gpt-image-{tier}"])
    return dict(BUYER_SHOW_MODEL_OPTIONS[DEFAULT_BUYER_SHOW_MODEL_ID])


def _nearest_buyer_show_size(width: int, height: int, tier: str = "4k") -> dict:
    source_ratio = max(1, int(width or 0)) / float(max(1, int(height or 0)))
    option = min(
        BUYER_SHOW_SIZE_OPTIONS,
        key=lambda item: abs(source_ratio - (float(item["width"]) / float(item["height"]))),
    )
    size_key = "size_2k" if _compact(tier).lower() == "2k" else "size_4k"
    return {
        "size": str(option[size_key]),
        "ratio": str(option["ratio"]),
        "tier": "2k" if size_key == "size_2k" else "4k",
    }


def _resolve_buyer_show_image_size(row: Mapping[str, Any], run_params: Mapping[str, Any]) -> dict:
    model_path = Path(_compact(row.get("模拍本地文件"))).expanduser()
    model_params = _resolve_buyer_show_model_params(run_params)
    tier = _compact(model_params.get("model_key_tier")).lower()
    if tier not in {"2k", "4k"}:
        tier = "4k"
    dimensions = _read_image_dimensions(model_path) if model_path.is_file() else None
    if dimensions:
        width, height = dimensions
        resolved = _nearest_buyer_show_size(width, height, tier)
        return {
            **resolved,
            "source_width": width,
            "source_height": height,
            "strategy": "source_ratio",
        }
    return {
        "size": DEFAULT_IMAGE_SIZE,
        "ratio": "3:4",
        "tier": tier,
        "source_width": 0,
        "source_height": 0,
        "strategy": "fallback_default",
    }


def _resolve_output_root(run_params: Mapping[str, Any]) -> Path:
    export_folder = _compact(run_params.get("export_folder") or run_params.get("output_dir"))
    if export_folder:
        return Path(export_folder).expanduser()
    return Path.home() / "Downloads" / "AI 买家秀全量测试"


def _is_resume_mode(run_params: Mapping[str, Any]) -> bool:
    value = _compact(run_params.get("execute_mode")).lower()
    return value in RESUME_EXECUTE_MODES or value.replace("_", "-") in RESUME_EXECUTE_MODES


def _resolve_package_paths(
    output_root: Path,
    package_base: str,
    run_params: Mapping[str, Any],
    *,
    resume_enabled: bool,
) -> tuple[Path, Path]:
    if not resume_enabled:
        return _ensure_unique_dir(output_root / package_base), output_root

    resume_dir = _compact(
        run_params.get("resume_package_dir")
        or run_params.get("recover_package_dir")
        or run_params.get("existing_package_dir")
        or run_params.get("package_dir")
    )
    if resume_dir:
        package_root = Path(resume_dir).expanduser()
    elif output_root.name == package_base:
        package_root = output_root
    else:
        package_root = output_root / package_base

    if not package_root.exists():
        raise FileNotFoundError(f"续跑原图包目录不存在：{package_root}")
    if not package_root.is_dir():
        raise NotADirectoryError(f"续跑原图包路径不是目录：{package_root}")
    return package_root, package_root.parent


def _artifact_path(path: Path, *, resume_enabled: bool) -> Path:
    if not resume_enabled:
        return _ensure_unique_path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def _write_rows_xlsx(path: Path, rows: list[dict], columns: list[str], sheet_name: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Sheet1"
    header_fill = PatternFill(fill_type="solid", fgColor="F5F7FA")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    ws.freeze_panes = "A2"
    for row in rows or []:
        ws.append([str((row or {}).get(column, "")) for column in columns])
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for row_idx in range(1, ws.max_row + 1):
            max_len = max(max_len, len(str(ws.cell(row=row_idx, column=col_idx).value or "")))
        ws.column_dimensions[letter].width = min(max_len + 4, 60)
    wb.save(str(path))
    return str(path)


def _target_slot_from_row(row: Mapping[str, Any], run_params: Mapping[str, Any]) -> str:
    explicit = _compact(row.get("替换部位") or row.get("部位") or run_params.get("garment_slot")).lower()
    if explicit in {"套装", "两件套", "连体衣", "整套", "全身", "outfit", "set", "full"}:
        return "outfit"
    if explicit in {"上装", "上衣", "top", "upper"}:
        return "upper"
    if explicit in {"下装", "裤", "裙", "bottom", "lower"}:
        return "lower"
    if explicit in {"鞋", "鞋类", "shoes"}:
        return "shoes"
    haystack = " ".join([
        _compact(row.get("AI素材库路径")),
        _compact(row.get("品类")),
        _compact(row.get("类目")),
        _compact(row.get("AI图包文件夹命名")),
    ]).lower()
    if any(token in haystack for token in ["套装", "两件套", "连体衣", "背带", "jumpsuit", "romper", "set"]):
        return "outfit"
    if any(token in haystack for token in ["鞋", "靴", "sneaker", "shoe"]):
        return "shoes"
    if any(token in haystack for token in ["下装", "裤", "裙", "短裤", "长裤", "半身裙"]):
        return "lower"
    return "upper"


def build_buyer_show_prompt(row: Mapping[str, Any], run_params: Mapping[str, Any]) -> str:
    custom_prompt = _compact(row.get("自定义Prompt") or row.get("自定义提示词") or run_params.get("custom_prompt"))
    prompt_extra = _compact(run_params.get("prompt_extra"))
    if custom_prompt:
        return "\n".join(part for part in [custom_prompt, prompt_extra] if part)

    slot = _target_slot_from_row(row, run_params)
    path_text = _compact(row.get("AI素材库路径"))
    special = any(token in path_text for token in ["羽绒服", "两件套", "连体衣", "套装", "背带", "马甲"])
    if slot == "outfit":
        base = (
            "基于第一张人物模拍图生成真实电商买家秀。以第二张产品平铺图为唯一商品参考，"
            "把参考图中的完整套装商品穿到人物身上：上衣、下装、连体结构、颜色分区、领口、袖口、裤型、"
            "图案、刺绣或贴标都必须同时匹配参考图。不要保留原模拍图里原有服装的颜色和款式，"
            "不要只替换上衣或只替换裤子。保持人物脸部、发型、体型、姿态、手脚、背景、相机角度、"
            "光影和鞋子等非商品元素不变。避免把整个人重画，避免拼接感、错位、额外肢体、Logo 变形和多余文字。"
            "输出自然生活化、真实清晰的买家秀照片。"
        )
        return "\n".join(part for part in [base, prompt_extra] if part)

    slot_text = {
        "upper": "上装",
        "lower": "下装",
        "shoes": "鞋类",
    }.get(slot, "目标服饰部位")
    if special:
        slot_text = f"特殊款式的{slot_text}"

    base = (
        "基于第一张人物模拍图生成真实电商买家秀。以第二张产品平铺图为唯一商品参考，"
        f"只替换人物身上的{slot_text}，保持人物脸部、发型、体型、姿态、背景、相机角度、"
        "光影和其他未替换服饰不变。商品颜色、版型、材质、图案、开合结构和穿着层次要贴合参考图，"
        "避免把整个人重画，避免拼接感、错位、额外肢体、Logo 变形和多余文字。输出自然生活化、真实清晰的买家秀照片。"
    )
    return "\n".join(part for part in [base, prompt_extra] if part)


def _created_run_task_id(summary: Mapping[str, Any]) -> str:
    task_id = _compact(summary.get("task_id"))
    if task_id:
        return task_id
    for run in summary.get("runs") or []:
        if isinstance(run, Mapping):
            task_id = _compact(run.get("task_id"))
            if task_id:
                return task_id
    return ""


def _result_image_urls(summary: Mapping[str, Any]) -> list[str]:
    urls: list[str] = []
    for raw in summary.get("image_urls") or []:
        value = _compact(raw)
        if value:
            urls.append(value)
    for run in summary.get("runs") or []:
        if not isinstance(run, Mapping):
            continue
        for raw in run.get("image_urls") or []:
            value = _compact(raw)
            if value and value not in urls:
                urls.append(value)
    return urls


TRANSIENT_ERROR_PATTERN = re.compile(
    r"unexpected_eof|eof occurred|connection (?:reset|aborted)|"
    r"timed out|timeout|temporary failure|ssl|http\s*(?:408|425|429|500|502|503|504)\b",
    re.IGNORECASE,
)


def _looks_transient_error(value: Any) -> bool:
    return bool(TRANSIENT_ERROR_PATTERN.search(_compact(value)))


def _brief_error(value: Any, limit: int = 180) -> str:
    text = _compact(value)
    return text[:limit] + ("..." if len(text) > limit else "")


def _candidate_result_urls(url: str) -> list[str]:
    source_url = _compact(url)
    direct_urls: list[str] = []
    try:
        parsed = urlparse(source_url)
        query = parse_qs(parsed.query)
    except Exception:
        query = {}
    for key in ("url", "image_url", "src"):
        for raw in query.get(key) or []:
            value = _compact(raw)
            if value.startswith(("http://", "https://")) and value not in direct_urls:
                direct_urls.append(value)
    urls: list[str] = []
    if source_url:
        urls.append(source_url)
    for direct_url in direct_urls:
        if direct_url not in urls:
            urls.append(direct_url)
    return urls


def _run_buyer_show_ai_job_with_retry(
    job_uid: str,
    *,
    settings: Optional[Mapping[str, Any]],
    log: Callable[[str], None],
    label: str,
    attempts: int = 3,
) -> dict:
    last_result: dict = {}
    safe_attempts = max(1, min(MAX_GENERATION_ATTEMPTS, int(attempts or DEFAULT_GENERATION_ATTEMPTS)))
    for attempt in range(1, safe_attempts + 1):
        try:
            result = ai_image_service.run_job_with_one_xm(job_uid, settings=settings)
        except ai_image_service.MissingModelKeyError:
            raise
        except Exception as exc:
            result = {"ok": False, "summary": {"error": str(exc)}}
        last_result = result if isinstance(result, dict) else {}
        summary = last_result.get("summary") if isinstance(last_result.get("summary"), Mapping) else {}
        if last_result.get("ok"):
            return last_result
        error = summary.get("error") or "AI 生图失败"
        if attempt >= safe_attempts:
            return last_result
        log(f"[buyer-show] AI 生图失败，自动重跑 {attempt + 1}/{safe_attempts}: {label} / {_brief_error(error)}")
        time.sleep(min(2 * attempt, 8))
    return last_result


def _materialize_generated_image_with_retry(
    job_uid: str,
    url: str,
    *,
    log: Callable[[str], None],
    label: str,
    attempts_per_url: int = DEFAULT_RESULT_DOWNLOAD_ATTEMPTS_PER_URL,
) -> dict:
    errors: list[str] = []
    for url_index, candidate_url in enumerate(_candidate_result_urls(url)):
        for attempt in range(1, max(1, attempts_per_url) + 1):
            try:
                return ai_image_service.materialize_remote_image(
                    job_uid,
                    candidate_url,
                    allow_unlisted=url_index > 0,
                )
            except Exception as exc:
                errors.append(_brief_error(exc))
                if attempt >= attempts_per_url or not _looks_transient_error(exc):
                    break
                log(f"[buyer-show] AI 结果落图临时失败，重试 {attempt + 1}/{attempts_per_url}: {label} / {_brief_error(exc)}")
                time.sleep(min(1.5 * attempt, 8))
    detail = errors[-1] if errors else "未知错误"
    raise RuntimeError(f"生成图下载失败：{detail}")


def _normalize_generation_concurrency(run_params: Mapping[str, Any]) -> int:
    raw = _compact(
        run_params.get("ai_generation_concurrency")
        or run_params.get("generation_concurrency")
        or run_params.get("concurrency")
    )
    if not raw:
        return DEFAULT_GENERATION_CONCURRENCY
    try:
        value = int(float(raw))
    except Exception:
        return DEFAULT_GENERATION_CONCURRENCY
    return max(1, min(MAX_GENERATION_CONCURRENCY, value))


def _normalize_generation_attempts(run_params: Mapping[str, Any]) -> int:
    raw = _compact(
        run_params.get("ai_generation_retry_attempts")
        or run_params.get("generation_retry_attempts")
        or run_params.get("ai_generation_attempts")
    )
    if not raw:
        return DEFAULT_GENERATION_ATTEMPTS
    try:
        value = int(float(raw))
    except Exception:
        return DEFAULT_GENERATION_ATTEMPTS
    return max(1, min(MAX_GENERATION_ATTEMPTS, value))


def _normalize_result_download_concurrency(run_params: Mapping[str, Any], fallback: int) -> int:
    raw = _compact(
        run_params.get("ai_result_download_concurrency")
        or run_params.get("result_download_concurrency")
        or run_params.get("download_concurrency")
    )
    if not raw:
        return max(1, min(MAX_RESULT_DOWNLOAD_CONCURRENCY, int(fallback or DEFAULT_RESULT_DOWNLOAD_CONCURRENCY)))
    try:
        value = int(float(raw))
    except Exception:
        return max(1, min(MAX_RESULT_DOWNLOAD_CONCURRENCY, int(fallback or DEFAULT_RESULT_DOWNLOAD_CONCURRENCY)))
    return max(1, min(MAX_RESULT_DOWNLOAD_CONCURRENCY, value))


def _create_buyer_show_job(row: Mapping[str, Any], run_params: Mapping[str, Any], prompt: str, output_dir: Path) -> dict:
    model_path = _compact(row.get("模拍本地文件"))
    reference_path = _compact(row.get("平铺本地文件"))
    style_color_code = _compact(row.get("款色号"))
    style_code = _extract_style_code(style_color_code)
    model_params = _resolve_buyer_show_model_params(run_params)
    size_info = _resolve_buyer_show_image_size(row, run_params)
    title = _safe_local_name(
        f"AI买家秀 {style_color_code or style_code} {row.get('唯一值') or ''}",
        "AI买家秀",
    )
    params = {
        "prompt": prompt,
        "size": _compact(size_info.get("size")) or DEFAULT_IMAGE_SIZE,
        "quality": _compact(run_params.get("quality")) or DEFAULT_QUALITY,
        "output_format": _compact(run_params.get("output_format")) or DEFAULT_OUTPUT_FORMAT,
        "n": 1,
        "model_id": _compact(model_params.get("model_id")),
        "model": _compact(model_params.get("model")),
        "model_key_tier": _compact(model_params.get("model_key_tier")) or _compact(size_info.get("tier")) or "4k",
        "model_provider_family": _compact(model_params.get("provider_family")),
        "main_image_path": model_path,
        "reference_image_paths": [reference_path],
        "workflow": BUYER_SHOW_TASK_ID,
        "surface": "semir-cloud-drive",
        "style_code": style_code,
        "style_color_code": style_color_code,
        "ratio": _compact(size_info.get("ratio")),
        "source_image_width": int(size_info.get("source_width") or 0),
        "source_image_height": int(size_info.get("source_height") or 0),
        "source_image_ratio": _compact(size_info.get("ratio")),
        "image_size_strategy": _compact(size_info.get("strategy")),
        "unique_value": _compact(row.get("唯一值")),
        "model_cloud_path": _compact(row.get("模拍云盘路径")),
        "reference_cloud_path": _compact(row.get("平铺云盘路径")),
    }
    if _compact(model_params.get("resolution")):
        params["resolution"] = _compact(model_params.get("resolution"))
    job = data_sink.create_ai_image_job({
        "title": title,
        "prompt": prompt,
        "model_key": _compact(model_params.get("model")) or "gpt-image-2",
        "status": "draft",
        "output_dir": str(output_dir),
        "params": params,
        "summary": {
            "workflow": BUYER_SHOW_TASK_ID,
            "surface": "semir-cloud-drive",
            "style_code": style_code,
            "style_color_code": style_color_code,
            "model_id": _compact(model_params.get("model_id")),
            "model_key": _compact(model_params.get("model")),
            "model_key_tier": _compact(model_params.get("model_key_tier")),
            "unique_value": _compact(row.get("唯一值")),
            "model_cloud_path": _compact(row.get("模拍云盘路径")),
            "reference_cloud_path": _compact(row.get("平铺云盘路径")),
        },
    })
    job_uid = _compact(job.get("job_uid"))
    data_sink.create_ai_image_asset({
        "job_uid": job_uid,
        "kind": "main",
        "source_type": "local",
        "path": model_path,
        "sort_order": 0,
        "meta": {"role": "buyer_show_model", "style_code": style_code},
    })
    data_sink.create_ai_image_asset({
        "job_uid": job_uid,
        "kind": "reference",
        "source_type": "local",
        "path": reference_path,
        "sort_order": 1,
        "meta": {"role": "flat_product_reference", "style_code": style_code},
    })
    return job


def _is_successful_download(row: Mapping[str, Any], result_key: str, path_key: str) -> bool:
    return _compact(row.get(result_key)) == "已下载" and Path(_compact(row.get(path_key))).expanduser().is_file()


def _package_output_stem(row: Mapping[str, Any], ordinal: int) -> str:
    unique_value = _safe_local_name(row.get("唯一值") or row.get("AI图包文件夹命名"), "未命名")
    style_color_code = _safe_local_name(row.get("款色号"), "款色号")
    return f"{unique_value}_{ordinal:03d}_{style_color_code}_AI买家秀"


def _package_filename(row: Mapping[str, Any], ordinal: int, source_path: Path) -> str:
    stem = _package_output_stem(row, ordinal)
    suffix = _image_suffix_from_file(source_path)
    return f"{stem}{suffix}"


def _material_package_filename(row: Mapping[str, Any], ordinal: int, role: str, source_path: Path) -> str:
    unique_value = _safe_local_name(row.get("唯一值") or row.get("AI图包文件夹命名"), "未命名")
    style_color_code = _safe_local_name(row.get("款色号"), "款色号")
    suffix = _image_suffix_from_file(source_path, fallback=".jpg")
    role_name = _safe_local_name(role, "素材")
    return f"{unique_value}_{ordinal:03d}_{style_color_code}_{role_name}{suffix}"


def _copy_downloaded_materials_for_review(
    row: dict,
    package_root: Path,
    folder_ordinals: dict[str, int],
) -> list[str]:
    unique_folder = _safe_local_name(row.get("唯一值") or row.get("AI图包文件夹命名"), "未命名")
    target_dir = package_root / unique_folder
    copied_files: list[str] = []
    for result_key, path_key, role in [
        ("模拍下载结果", "模拍本地文件", "模拍原图"),
        ("平铺下载结果", "平铺本地文件", "平铺参考图"),
    ]:
        if not _is_successful_download(row, result_key, path_key):
            continue
        source = Path(_compact(row.get(path_key))).expanduser()
        folder_ordinals[unique_folder] = int(folder_ordinals.get(unique_folder) or 0) + 1
        target = target_dir / _material_package_filename(row, folder_ordinals[unique_folder], role, source)
        copied = _copy_file_to_unique_target(source, target)
        copied_files.append(str(copied))
    if copied_files:
        row["本地图包文件夹"] = str(target_dir)
        row["下载素材文件"] = "\n".join(copied_files)
    return copied_files


def _make_usage_export_row(row: Mapping[str, Any], output_file: str, created_at: str) -> dict:
    return {
        "记录时间": created_at,
        "货号": _extract_style_code(row.get("款色号")),
        "款色号": _compact(row.get("款色号")),
        "唯一值": _compact(row.get("唯一值")),
        "源表行号": _compact(row.get("表格行号")),
        "模拍文件": _compact(row.get("模拍文件")),
        "模拍云盘路径": _compact(row.get("模拍云盘路径")),
        "平铺云盘路径": _compact(row.get("平铺云盘路径")),
        "AI任务ID": _compact(row.get("AI任务ID")),
        "1XM任务ID": _compact(row.get("1XM任务ID")),
        "输出文件": output_file,
    }


def _buyer_show_row_label(row: Mapping[str, Any]) -> str:
    style_color_code = _compact(row.get("款色号"))
    style_code = _extract_style_code(style_color_code)
    return f"{style_color_code or style_code} / {row.get('模拍文件') or ''}"


def _record_buyer_show_usage(row: dict, output_files: list[str], prompt: str) -> None:
    style_color_code = _compact(row.get("款色号"))
    style_code = _extract_style_code(style_color_code)
    model_cloud_path = _compact(row.get("模拍云盘路径"))
    target_dir = Path(_compact(row.get("本地图包文件夹"))).expanduser()
    created_at = datetime.now().isoformat(timespec="seconds")
    row["__usage_record_time"] = created_at
    data_sink.create_buyer_show_material_usage({
        "style_code": style_code,
        "style_color_code": style_color_code,
        "model_cloud_path": model_cloud_path,
        "model_filename": _compact(row.get("模拍文件")),
        "model_local_path": _compact(row.get("模拍本地文件")),
        "reference_cloud_path": _compact(row.get("平铺云盘路径")),
        "reference_local_path": _compact(row.get("平铺本地文件")),
        "output_file": "\n".join(output_files),
        "source_row_no": int(row.get("表格行号") or 0),
        "unique_value": _compact(row.get("唯一值")),
        "ai_job_uid": _compact(row.get("AI任务ID")),
        "ai_task_id": _compact(row.get("1XM任务ID")),
        "meta": {
            "workflow": BUYER_SHOW_TASK_ID,
            "package_root": str(target_dir.parent),
            "prompt": prompt,
        },
    })


def _find_existing_buyer_show_output(row: Mapping[str, Any]) -> Optional[Path]:
    target_dir = Path(_compact(row.get("本地图包文件夹"))).expanduser()
    if not target_dir.is_dir():
        return None
    ordinal_start = max(1, int(row.get("__package_ordinal_start") or 1))
    stem = _package_output_stem(row, ordinal_start)
    candidates: list[Path] = []
    for path in target_dir.iterdir():
        if not path.is_file():
            continue
        if path.suffix.lower() not in IMAGE_OUTPUT_SUFFIXES:
            continue
        if not path.name.startswith(stem):
            continue
        try:
            if path.stat().st_size <= 0:
                continue
        except OSError:
            continue
        candidates.append(path)
    if not candidates:
        return None
    return sorted(candidates, key=lambda item: (0 if item.stem == stem else 1, item.name))[0]


def _mark_existing_buyer_show_output(
    row: dict,
    existing_path: Path,
    *,
    run_params: Mapping[str, Any],
    log: Callable[[str], None],
) -> None:
    usage = data_sink.find_buyer_show_material_usage(_compact(row.get("款色号")), _compact(row.get("模拍云盘路径")))
    if usage:
        row["AI任务ID"] = _compact(row.get("AI任务ID")) or _compact(usage.get("ai_job_uid"))
        row["1XM任务ID"] = _compact(row.get("1XM任务ID")) or _compact(usage.get("ai_task_id"))
    row["生图结果"] = "已生成"
    row["生图文件"] = str(existing_path)
    row["__生成图明细"] = [{
        "filename": existing_path.name,
        "local_path": str(existing_path),
        "source_url": "",
        "resume": True,
    }]
    row["备注"] = _append_note(row.get("备注"), "续跑复用原图包已有成品")
    prompt = _compact(row.get("__generation_prompt")) or build_buyer_show_prompt(row, run_params)
    _record_buyer_show_usage(row, [str(existing_path)], prompt)
    log(f"[buyer-show] 续跑复用已有成品: {_buyer_show_row_label(row)}")


def _prepare_buyer_show_resume_row(
    row: dict,
    *,
    run_params: Mapping[str, Any],
    log: Callable[[str], None],
) -> bool:
    existing = _find_existing_buyer_show_output(row)
    if existing:
        _mark_existing_buyer_show_output(row, existing, run_params=run_params, log=log)
        return True

    job = data_sink.find_completed_buyer_show_ai_image_job(
        style_color_code=_compact(row.get("款色号")),
        main_image_path=_compact(row.get("模拍本地文件")),
        reference_image_path=_compact(row.get("平铺本地文件")),
    )
    if not job:
        return False

    summary = job.get("summary") if isinstance(job.get("summary"), Mapping) else {}
    urls = _result_image_urls(summary)
    if not urls:
        log(f"[buyer-show] 续跑找到已完成 AI 任务但缺少图片链接，将重新生成: {_buyer_show_row_label(row)}")
        return False

    row["AI任务ID"] = _compact(job.get("job_uid"))
    row["1XM任务ID"] = _created_run_task_id(summary)
    row["__generation_urls"] = urls
    row["__generation_prompt"] = _compact(job.get("prompt")) or _compact(summary.get("prompt")) or build_buyer_show_prompt(row, run_params)
    row["生图结果"] = "待落图"
    row["备注"] = _append_note(row.get("备注"), "续跑复用已完成 AI 生图链接")
    log(f"[buyer-show] 续跑复用已完成 AI 任务链接: {_buyer_show_row_label(row)}")
    return True


def _prepare_buyer_show_generation_row(
    row: dict,
    *,
    run_params: Mapping[str, Any],
    settings: Optional[Mapping[str, Any]],
    log: Callable[[str], None],
) -> dict:
    label = _buyer_show_row_label(row)
    target_dir = Path(_compact(row.get("本地图包文件夹"))).expanduser()
    target_dir.mkdir(parents=True, exist_ok=True)

    try:
        prompt = build_buyer_show_prompt(row, run_params)
        job = _create_buyer_show_job(row, run_params, prompt, target_dir)
        job_uid = _compact(job.get("job_uid"))
        row["AI任务ID"] = job_uid
        result = _run_buyer_show_ai_job_with_retry(
            job_uid,
            settings=settings,
            log=log,
            label=label,
            attempts=_normalize_generation_attempts(run_params),
        )
        summary = result.get("summary") if isinstance(result.get("summary"), Mapping) else {}
        row["1XM任务ID"] = _created_run_task_id(summary)
        if not result.get("ok"):
            row["生图结果"] = "生成失败"
            row["备注"] = _append_note(row.get("备注"), summary.get("error") or "AI 生图失败")
            log(f"[buyer-show] 生成失败: {label} / {_brief_error(summary.get('error') or 'AI 生图失败')}")
            return row

        urls = _result_image_urls(summary)
        if not urls:
            row["生图结果"] = "生成失败"
            row["备注"] = _append_note(row.get("备注"), "AI 任务完成但未返回图片地址")
            log(f"[buyer-show] 生成失败: {label} / AI 任务完成但未返回图片地址")
            return row

        row["生图结果"] = "待落图"
        row["__generation_urls"] = urls
        row["__generation_prompt"] = prompt
        return row
    except ai_image_service.MissingModelKeyError as exc:
        row["生图结果"] = "配置缺失"
        row["备注"] = _append_note(row.get("备注"), str(exc))
        log(f"[buyer-show] 生成失败: {label} / {_brief_error(exc)}")
        return row
    except Exception as exc:
        row["生图结果"] = "生成失败"
        row["备注"] = _append_note(row.get("备注"), str(exc))
        log(f"[buyer-show] 生成失败: {label} / {_brief_error(exc)}")
        return row


def _materialize_buyer_show_generation_row(
    row: dict,
    *,
    run_params: Mapping[str, Any],
    log: Callable[[str], None],
) -> dict:
    label = _buyer_show_row_label(row)
    target_dir = Path(_compact(row.get("本地图包文件夹"))).expanduser()
    target_dir.mkdir(parents=True, exist_ok=True)
    ordinal_start = max(1, int(row.get("__package_ordinal_start") or 1))
    job_uid = _compact(row.get("AI任务ID"))
    urls = [url for url in row.get("__generation_urls") or [] if _compact(url)]
    prompt = _compact(row.get("__generation_prompt")) or build_buyer_show_prompt(row, run_params)

    if _compact(row.get("生图结果")) != "待落图":
        return row
    if not job_uid or not urls:
        row["生图结果"] = "生成失败"
        row["备注"] = _append_note(row.get("备注"), "AI 任务完成但缺少可落图链接")
        log(f"[buyer-show] 生成失败: {label} / AI 任务完成但缺少可落图链接")
        return row

    try:
        copied_files: list[str] = []
        generated_details: list[dict] = []
        for url_index, url in enumerate(urls):
            materialized = _materialize_generated_image_with_retry(job_uid, url, log=log, label=label)
            source = Path(_compact(materialized.get("path"))).expanduser()
            if not source.is_file():
                raise FileNotFoundError(f"生成图物化失败：{source}")
            target = target_dir / _package_filename(row, ordinal_start + url_index, source)
            if _is_resume_mode(run_params) and target.exists() and target.is_file():
                try:
                    if target.stat().st_size <= 0:
                        target.unlink()
                except OSError:
                    pass
            copied = _copy_file_to_unique_target(source, target)
            copied_files.append(str(copied))
            generated_details.append({"filename": copied.name, "local_path": str(copied), "source_url": url})

        row["生图结果"] = "已生成"
        row["生图文件"] = "\n".join(copied_files)
        row["__生成图明细"] = generated_details
        _record_buyer_show_usage(row, copied_files, prompt)
        log(f"[buyer-show] 已生成: {label}")
        return row
    except Exception as exc:
        row["生图结果"] = "落图失败"
        row["备注"] = _append_note(row.get("备注"), str(exc))
        log(f"[buyer-show] 落图失败: {label} / {_brief_error(exc)}")
        return row


def _generate_buyer_show_row(
    row: dict,
    *,
    run_params: Mapping[str, Any],
    settings: Optional[Mapping[str, Any]],
    log: Callable[[str], None],
) -> dict:
    _prepare_buyer_show_generation_row(row, run_params=run_params, settings=settings, log=log)
    return _materialize_buyer_show_generation_row(row, run_params=run_params, log=log)


def _run_buyer_show_generation_pipeline(
    generation_rows: list[dict],
    *,
    run_params: Mapping[str, Any],
    settings: Optional[Mapping[str, Any]],
    generation_concurrency: int,
    log: Callable[[str], None],
    materialize_only_rows: Optional[list[dict]] = None,
) -> None:
    prepared_rows = list(materialize_only_rows or [])
    if not generation_rows and not prepared_rows:
        return

    worker_count = min(max(1, generation_concurrency), len(generation_rows)) if generation_rows else 1
    download_total = max(1, len(generation_rows) + len(prepared_rows))
    download_worker_count = min(
        _normalize_result_download_concurrency(run_params, DEFAULT_RESULT_DOWNLOAD_CONCURRENCY),
        download_total,
    )
    if generation_rows:
        log(f"[buyer-show] AI 生图并发窗口：{worker_count}；待生成 {len(generation_rows)} 行")
    if prepared_rows:
        log(f"[buyer-show] 续跑待落图 {len(prepared_rows)} 行；直接进入下载队列")
    log(f"[buyer-show] AI 结果落图并发窗口：{download_worker_count}；生成链接就入下载队列")

    generation_completed = 0
    download_submitted = 0
    download_completed = 0

    with (
        ThreadPoolExecutor(max_workers=worker_count, thread_name_prefix="buyer-show-ai") as generation_executor,
        ThreadPoolExecutor(max_workers=download_worker_count, thread_name_prefix="buyer-show-download") as download_executor,
    ):
        generation_futures = {
            generation_executor.submit(
                _prepare_buyer_show_generation_row,
                row,
                run_params=run_params,
                settings=settings,
                log=log,
            ): row
            for row in generation_rows
        }
        download_futures: dict[Any, dict] = {}
        for row in prepared_rows:
            download_futures[
                download_executor.submit(
                    _materialize_buyer_show_generation_row,
                    row,
                    run_params=run_params,
                    log=log,
                )
            ] = row
            download_submitted += 1

        while generation_futures or download_futures:
            done, _pending = wait(
                {*generation_futures.keys(), *download_futures.keys()},
                return_when=FIRST_COMPLETED,
            )
            for future in done:
                if future in generation_futures:
                    row = generation_futures.pop(future)
                    try:
                        future.result()
                    except Exception as exc:
                        row["生图结果"] = "生成失败"
                        row["备注"] = _append_note(row.get("备注"), str(exc))
                        log(f"[buyer-show] 生成失败: {row.get('款色号') or row.get('货号') or ''} / {row.get('模拍文件') or ''} / {_brief_error(exc)}")
                    generation_completed += 1
                    if _compact(row.get("生图结果")) == "待落图":
                        download_futures[
                            download_executor.submit(
                                _materialize_buyer_show_generation_row,
                                row,
                                run_params=run_params,
                                log=log,
                            )
                        ] = row
                        download_submitted += 1
                    if generation_completed % 5 == 0 or generation_completed == len(generation_rows):
                        log(f"[buyer-show] AI 生图链接收集进度 {generation_completed}/{len(generation_rows)}；已入落图队列 {download_submitted}")
                    continue

                row = download_futures.pop(future)
                try:
                    future.result()
                except Exception as exc:
                    row["生图结果"] = "落图失败"
                    row["备注"] = _append_note(row.get("备注"), str(exc))
                    log(f"[buyer-show] 落图失败: {row.get('款色号') or row.get('货号') or ''} / {row.get('模拍文件') or ''} / {_brief_error(exc)}")
                download_completed += 1
                if download_completed % 5 == 0 or (
                    not generation_futures and download_completed == download_submitted
                ):
                    log(f"[buyer-show] AI 结果落图进度 {download_completed}/{download_submitted}")


def finalize_buyer_show_outputs(
    *,
    data_rows: list,
    runtime_files: list,
    exported_files: list,
    run_params: Mapping[str, Any],
    runtime_artifact_dir: str,
    settings: Optional[Mapping[str, Any]] = None,
    log: Callable[[str], None] = print,
) -> list[str]:
    data_sink.init_db()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_root = _resolve_output_root(run_params)
    output_root.mkdir(parents=True, exist_ok=True)
    package_base = _safe_local_name(
        run_params.get("package_name") or f"AI买家秀_{timestamp}",
        f"AI买家秀_{timestamp}",
    )
    resume_enabled = _is_resume_mode(run_params)
    package_root, zip_output_root = _resolve_package_paths(
        output_root,
        package_base,
        run_params,
        resume_enabled=resume_enabled,
    )

    rows = [dict(row) if isinstance(row, Mapping) else {"备注": str(row)} for row in (data_rows or [])]
    usage_export_rows: list[dict] = []
    folder_ordinals: dict[str, int] = {}
    material_folder_ordinals: dict[str, int] = {}
    seen_usage_keys: set[tuple[str, str]] = set()
    execute_mode = _compact(run_params.get("execute_mode") or "generate").lower()
    generate_enabled = execute_mode not in {"download_only", "download-only", "plan", "false", "0"}
    enforce_usage = _compact(run_params.get("usage_record_mode") or "enforce").lower() != "ignore"
    usage_blocks_enabled = enforce_usage and not resume_enabled
    max_generate_jobs_raw = _compact(run_params.get("max_generate_jobs"))
    max_generate_jobs = int(max_generate_jobs_raw) if max_generate_jobs_raw.isdigit() else 0
    generation_concurrency = _normalize_generation_concurrency(run_params)
    scheduled_count = 0
    generation_rows: list[dict] = []
    materialize_only_rows: list[dict] = []

    log(f"[buyer-show] 开始后处理 {len(rows)} 行，输出目录：{package_root}")
    if resume_enabled:
        log("[buyer-show] 当前为原包续跑模式：复用已有成品或已完成 AI 链接，缺口才重新生图")

    for index, row in enumerate(rows):
        style_color_code = _compact(row.get("款色号"))
        style_code = _extract_style_code(style_color_code)
        row["货号"] = style_code
        model_cloud_path = _compact(row.get("模拍云盘路径"))
        usage_key = (style_color_code, model_cloud_path)
        if not generate_enabled:
            copied_materials = _copy_downloaded_materials_for_review(row, package_root, material_folder_ordinals)
            if copied_materials:
                row["备注"] = _append_note(row.get("备注"), f"已打包下载素材 {len(copied_materials)} 个")

        if not _is_successful_download(row, "模拍下载结果", "模拍本地文件"):
            row["生图结果"] = "素材不完整"
            row["备注"] = _append_note(row.get("备注"), "模拍图未成功下载")
            continue
        if not _is_successful_download(row, "平铺下载结果", "平铺本地文件"):
            row["生图结果"] = "素材不完整"
            row["备注"] = _append_note(row.get("备注"), "平铺参考图未成功下载")
            continue
        if not style_color_code or not model_cloud_path:
            row["生图结果"] = "参数缺失"
            row["备注"] = _append_note(row.get("备注"), "缺少款色号或模拍云盘路径")
            continue

        if resume_enabled:
            unique_folder = _safe_local_name(row.get("唯一值") or row.get("AI图包文件夹命名"), "未命名")
            target_dir = package_root / unique_folder
            target_dir.mkdir(parents=True, exist_ok=True)
            row["本地图包文件夹"] = str(target_dir)
            folder_ordinals[unique_folder] = int(folder_ordinals.get(unique_folder) or 0) + 1
            row["__package_ordinal_start"] = folder_ordinals[unique_folder]
            if _prepare_buyer_show_resume_row(row, run_params=run_params, log=log):
                if _compact(row.get("生图结果")) == "待落图":
                    materialize_only_rows.append(row)
                continue

        if usage_blocks_enabled and usage_key in seen_usage_keys:
            row["生图结果"] = "已跳过"
            row["备注"] = _append_note(row.get("备注"), "本批次同款色号已使用过这张模拍图")
            continue
        if usage_blocks_enabled:
            existing = data_sink.find_buyer_show_material_usage(style_color_code, model_cloud_path)
            if existing:
                row["生图结果"] = "已跳过"
                row["备注"] = _append_note(row.get("备注"), "使用记录已存在，同款色号不重复调用这张模拍图")
                row["AI任务ID"] = _compact(existing.get("ai_job_uid"))
                row["1XM任务ID"] = _compact(existing.get("ai_task_id"))
                row["生图文件"] = _compact(existing.get("output_file"))
                continue
        if max_generate_jobs > 0 and scheduled_count >= max_generate_jobs:
            row["生图结果"] = "已跳过"
            row["备注"] = _append_note(row.get("备注"), f"超过本次最多真实生成 {max_generate_jobs} 条")
            continue
        if not generate_enabled:
            row["生图结果"] = "待生成"
            row["备注"] = _append_note(row.get("备注"), "当前为下载/打包预演模式，未提交 AI 生图")
            continue

        if not resume_enabled:
            unique_folder = _safe_local_name(row.get("唯一值") or row.get("AI图包文件夹命名"), "未命名")
            target_dir = package_root / unique_folder
            target_dir.mkdir(parents=True, exist_ok=True)
            row["本地图包文件夹"] = str(target_dir)
            folder_ordinals[unique_folder] = int(folder_ordinals.get(unique_folder) or 0) + 1
            row["__package_ordinal_start"] = folder_ordinals[unique_folder]
        scheduled_count += 1
        generation_rows.append(row)
        if usage_blocks_enabled:
            seen_usage_keys.add(usage_key)

    _run_buyer_show_generation_pipeline(
        generation_rows,
        run_params=run_params,
        settings=settings,
        generation_concurrency=generation_concurrency,
        log=log,
        materialize_only_rows=materialize_only_rows,
    )

    usage_export_rows = [
        _make_usage_export_row(row, _compact(row.get("生图文件")), _compact(row.get("__usage_record_time")))
        for row in rows
        if _compact(row.get("生图结果")) == "已生成" and _compact(row.get("__usage_record_time"))
    ]

    summary_path = _artifact_path(package_root / f"{package_root.name}_执行结果.xlsx", resume_enabled=resume_enabled)
    usage_path = _artifact_path(package_root / f"{package_root.name}_使用记录.xlsx", resume_enabled=resume_enabled)
    _write_rows_xlsx(summary_path, rows, SUMMARY_COLUMNS, "执行结果")
    _write_rows_xlsx(usage_path, usage_export_rows or [], USAGE_COLUMNS, "使用记录")

    zip_path = _artifact_path(zip_output_root / f"{package_root.name}.zip", resume_enabled=resume_enabled)
    started = time.monotonic()
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for file_path in sorted(package_root.rglob("*")):
            if file_path.is_file():
                archive.write(file_path, arcname=str(file_path.relative_to(package_root.parent)))
    log(f"[buyer-show] 本地图包已打包：{zip_path} ({time.monotonic() - started:.1f}s)")

    final_refs = [str(package_root), str(zip_path), str(summary_path), str(usage_path)]
    for file_path in exported_files or []:
        value = _compact(file_path)
        if value and Path(value).expanduser().is_file():
            final_refs.append(value)
    return final_refs
