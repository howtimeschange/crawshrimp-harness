"""
Data persistence layer
- Task run state stored in SQLite
- Data records exported to Excel / JSON
- Filename templates support {date}, {datetime}, {timestamp}, {adapter_id}, {task_id}
"""
import json
import logging
import os
import re
import sqlite3
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, List, Mapping, Optional

from core import runtime_paths
from core.models import TaskRun, TaskStatus

logger = logging.getLogger(__name__)

MAX_EXPORT_FILENAME_LENGTH = 140


class AiVideoRequestUidConflictError(Exception):
    """Raised when a globally unique AI-video requestUid belongs to another Job."""


class AiVideoRetryConflictError(Exception):
    """Raised when a retry loses the Job status/current-Run compare-and-swap."""


class _SafeTemplateVars(dict):
    def __missing__(self, key):
        return ""


def _data_root() -> Path:
    return runtime_paths.child_dir("data")


def _db_path() -> Path:
    return runtime_paths.data_root() / "crawshrimp.db"


def _harden_db_file_permissions() -> None:
    """Best-effort POSIX hardening for the local SQLite file that stores machine tokens."""
    if os.name != "posix":
        return
    db_path = _db_path()
    try:
        db_path.parent.mkdir(parents=True, exist_ok=True)
        os.chmod(db_path.parent, 0o700)
        if db_path.exists():
            os.chmod(db_path, 0o600)
    except OSError as exc:
        logger.warning("Unable to harden Crawshrimp SQLite permissions: %s", exc)


def _get_conn() -> sqlite3.Connection:
    _harden_db_file_permissions()
    conn = sqlite3.connect(str(_db_path()))
    conn.row_factory = sqlite3.Row
    _harden_db_file_permissions()
    return conn


def init_db():
    """Create tables if not exists"""
    with _get_conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS task_runs (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                adapter_id   TEXT NOT NULL,
                task_id      TEXT NOT NULL,
                status       TEXT NOT NULL DEFAULT 'idle',
                started_at   TEXT,
                finished_at  TEXT,
                records_count INTEGER DEFAULT 0,
                error        TEXT,
                output_files TEXT DEFAULT '[]'
            )
        """)
        _ensure_column(conn, "task_runs", "last_seen_at", "TEXT")
        _ensure_column(conn, "task_runs", "phase", "TEXT DEFAULT ''")
        _ensure_column(conn, "task_runs", "current_row", "INTEGER DEFAULT 0")
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_task_runs_adapter_task
            ON task_runs (adapter_id, task_id)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS task_instances (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                instance_uid  TEXT NOT NULL UNIQUE,
                adapter_id    TEXT NOT NULL,
                task_id       TEXT NOT NULL,
                title         TEXT NOT NULL,
                status        TEXT NOT NULL DEFAULT 'draft',
                current_step  TEXT NOT NULL DEFAULT 'config',
                params_json   TEXT NOT NULL DEFAULT '{}',
                summary_json  TEXT NOT NULL DEFAULT '{}',
                last_run_id   INTEGER,
                archived      INTEGER NOT NULL DEFAULT 0,
                created_at    TEXT NOT NULL,
                updated_at    TEXT NOT NULL,
                completed_at  TEXT
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_task_instances_adapter_task_status
            ON task_instances (adapter_id, task_id, status, updated_at)
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_task_instances_updated
            ON task_instances (updated_at)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS task_instance_runs (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                instance_uid  TEXT NOT NULL,
                run_id        INTEGER NOT NULL,
                purpose       TEXT NOT NULL DEFAULT 'main',
                created_at    TEXT NOT NULL,
                UNIQUE(instance_uid, run_id, purpose)
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_task_instance_runs_instance
            ON task_instance_runs (instance_uid, created_at)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS task_instance_artifacts (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                instance_uid  TEXT NOT NULL,
                kind          TEXT NOT NULL,
                label         TEXT NOT NULL,
                path          TEXT NOT NULL,
                meta_json     TEXT NOT NULL DEFAULT '{}',
                created_at    TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_task_instance_artifacts_instance
            ON task_instance_artifacts (instance_uid, created_at)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS task_instance_events (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                instance_uid  TEXT NOT NULL,
                event_type    TEXT NOT NULL,
                message       TEXT NOT NULL,
                meta_json     TEXT NOT NULL DEFAULT '{}',
                created_at    TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_task_instance_events_instance
            ON task_instance_events (instance_uid, created_at)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS task_schedules (
                id                 INTEGER PRIMARY KEY AUTOINCREMENT,
                schedule_uid       TEXT NOT NULL UNIQUE,
                adapter_id         TEXT NOT NULL,
                task_id            TEXT NOT NULL,
                title              TEXT NOT NULL,
                enabled            INTEGER NOT NULL DEFAULT 1,
                frequency          TEXT NOT NULL,
                time_of_day        TEXT NOT NULL,
                weekday            INTEGER,
                params_json        TEXT NOT NULL DEFAULT '{}',
                notify_channel     TEXT NOT NULL DEFAULT 'dingtalk',
                notify_template    TEXT NOT NULL DEFAULT '',
                last_run_id        INTEGER,
                last_instance_uid  TEXT,
                last_status        TEXT,
                last_error         TEXT,
                last_triggered_at  TEXT,
                archived           INTEGER NOT NULL DEFAULT 0,
                created_at         TEXT NOT NULL,
                updated_at         TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_task_schedules_adapter_task_enabled
            ON task_schedules (adapter_id, task_id, enabled, archived, updated_at)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS cloud_machine_credentials (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                machine_id TEXT NOT NULL DEFAULT '',
                machine_token TEXT NOT NULL DEFAULT '',
                machine_name TEXT NOT NULL DEFAULT '',
                capabilities_json TEXT NOT NULL DEFAULT '[]',
                updated_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS cloud_job_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_uid TEXT NOT NULL,
                event_type TEXT NOT NULL,
                message TEXT NOT NULL DEFAULT '',
                payload_json TEXT NOT NULL DEFAULT '{}',
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS cloud_job_completion_results (
                job_uid TEXT PRIMARY KEY,
                lease_id TEXT NOT NULL DEFAULT '',
                result_json TEXT NOT NULL DEFAULT '{}',
                last_error TEXT NOT NULL DEFAULT '',
                attempt_count INTEGER NOT NULL DEFAULT 0,
                next_attempt_at TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        """)
        _ensure_column(conn, "cloud_job_completion_results", "lease_id", "TEXT NOT NULL DEFAULT ''")
        _ensure_column(conn, "cloud_job_completion_results", "last_error", "TEXT NOT NULL DEFAULT ''")
        _ensure_column(conn, "cloud_job_completion_results", "attempt_count", "INTEGER NOT NULL DEFAULT 0")
        _ensure_column(conn, "cloud_job_completion_results", "next_attempt_at", "TEXT NOT NULL DEFAULT ''")
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_cloud_job_completion_results_due
            ON cloud_job_completion_results (next_attempt_at, created_at)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ai_image_jobs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_uid TEXT NOT NULL UNIQUE,
                title TEXT NOT NULL DEFAULT '',
                prompt TEXT NOT NULL DEFAULT '',
                model_key TEXT NOT NULL DEFAULT 'gpt-image-2',
                status TEXT NOT NULL DEFAULT 'draft',
                output_dir TEXT NOT NULL DEFAULT '',
                params_json TEXT NOT NULL DEFAULT '{}',
                summary_json TEXT NOT NULL DEFAULT '{}',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        """)
        _ensure_column(conn, "ai_image_jobs", "output_dir", "TEXT NOT NULL DEFAULT ''")
        _ensure_column(conn, "ai_image_jobs", "pinned_at", "TEXT NOT NULL DEFAULT ''")
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_ai_image_jobs_updated
            ON ai_image_jobs (updated_at)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ai_image_assets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_uid TEXT NOT NULL UNIQUE,
                job_uid TEXT NOT NULL DEFAULT '',
                kind TEXT NOT NULL DEFAULT 'reference',
                source_type TEXT NOT NULL DEFAULT 'local',
                path TEXT NOT NULL DEFAULT '',
                url TEXT NOT NULL DEFAULT '',
                mime_type TEXT NOT NULL DEFAULT '',
                sort_order INTEGER NOT NULL DEFAULT 0,
                meta_json TEXT NOT NULL DEFAULT '{}',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_ai_image_assets_job_order
            ON ai_image_assets (job_uid, sort_order, id)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ai_image_canvases (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                canvas_uid TEXT NOT NULL UNIQUE,
                job_uid TEXT NOT NULL DEFAULT '',
                title TEXT NOT NULL DEFAULT '',
                canvas_json TEXT NOT NULL DEFAULT '{}',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_ai_image_canvases_job
            ON ai_image_canvases (job_uid, updated_at)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ai_video_jobs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_uid TEXT NOT NULL UNIQUE,
                request_uid TEXT NOT NULL UNIQUE,
                title TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'draft',
                provider TEXT NOT NULL DEFAULT 'seedance',
                model TEXT NOT NULL DEFAULT '',
                prompt TEXT NOT NULL DEFAULT '',
                parameters_json TEXT NOT NULL DEFAULT '{}',
                current_run_uid TEXT NOT NULL DEFAULT '',
                output_dir TEXT NOT NULL DEFAULT '',
                deleted_at TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_ai_video_jobs_updated
            ON ai_video_jobs (updated_at DESC)
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_ai_video_jobs_status
            ON ai_video_jobs (status, updated_at DESC)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ai_video_assets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_uid TEXT NOT NULL UNIQUE,
                job_uid TEXT NOT NULL DEFAULT '',
                kind TEXT NOT NULL DEFAULT 'image',
                role TEXT NOT NULL DEFAULT 'reference_image',
                source_type TEXT NOT NULL DEFAULT 'local_file',
                original_name TEXT NOT NULL DEFAULT '',
                local_path TEXT NOT NULL DEFAULT '',
                remote_url TEXT NOT NULL DEFAULT '',
                mime_type TEXT NOT NULL DEFAULT '',
                width INTEGER NOT NULL DEFAULT 0,
                height INTEGER NOT NULL DEFAULT 0,
                size_bytes INTEGER NOT NULL DEFAULT 0,
                sha256 TEXT NOT NULL DEFAULT '',
                sort_order INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_ai_video_assets_job_order
            ON ai_video_assets (job_uid, sort_order, id)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ai_video_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_uid TEXT NOT NULL UNIQUE,
                request_uid TEXT NOT NULL UNIQUE,
                job_uid TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'queued',
                provider TEXT NOT NULL DEFAULT '',
                model TEXT NOT NULL DEFAULT '',
                input_snapshot_json TEXT NOT NULL DEFAULT '{}',
                provider_task_id TEXT NOT NULL DEFAULT '',
                provider_status TEXT NOT NULL DEFAULT '',
                archive_status TEXT NOT NULL DEFAULT 'none',
                output_json TEXT NOT NULL DEFAULT '{}',
                error_json TEXT NOT NULL DEFAULT '{}',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                submitted_at TEXT NOT NULL DEFAULT '',
                completed_at TEXT NOT NULL DEFAULT ''
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_ai_video_runs_job
            ON ai_video_runs (job_uid, created_at DESC)
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_ai_video_runs_active
            ON ai_video_runs (status, updated_at DESC)
        """)
        conn.commit()
    _harden_db_file_permissions()


def _ensure_column(conn: sqlite3.Connection, table_name: str, column_name: str, definition: str) -> None:
    columns = {
        str(row["name"])
        for row in conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    }
    if column_name not in columns:
        conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {definition}")


def _now_iso() -> str:
    return datetime.now().isoformat()


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _json_dumps(value: Any) -> str:
    return json.dumps(value or {}, ensure_ascii=False)


def _row_to_dict(row: Optional[sqlite3.Row]) -> Optional[dict]:
    return dict(row) if row else None


def _json_loads_object(value: Any) -> dict:
    try:
        parsed = json.loads(value) if isinstance(value, str) else value
    except Exception:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def save_cloud_machine_credentials(machine_id: str, machine_token: str, machine_name: str, capabilities: list[str]) -> dict:
    now = _now_iso()
    with _get_conn() as conn:
        conn.execute(
            """
            INSERT INTO cloud_machine_credentials (
                id, machine_id, machine_token, machine_name, capabilities_json, updated_at
            )
            VALUES (1, ?, ?, ?, ?, ?)
            ON CONFLICT(id) DO UPDATE SET
                machine_id = excluded.machine_id,
                machine_token = excluded.machine_token,
                machine_name = excluded.machine_name,
                capabilities_json = excluded.capabilities_json,
                updated_at = excluded.updated_at
            """,
            (machine_id, machine_token, machine_name, json.dumps(capabilities or [], ensure_ascii=False), now),
        )
        conn.commit()
    return get_cloud_machine_credentials() or {}


def get_cloud_machine_credentials() -> dict | None:
    with _get_conn() as conn:
        row = conn.execute("SELECT * FROM cloud_machine_credentials WHERE id = 1").fetchone()
    if not row:
        return None
    data = dict(row)
    try:
        data["capabilities"] = json.loads(data.pop("capabilities_json") or "[]")
    except Exception:
        data["capabilities"] = []
    return data


def clear_cloud_machine_credentials() -> None:
    with _get_conn() as conn:
        conn.execute("DELETE FROM cloud_machine_credentials WHERE id = 1")
        conn.commit()


def save_pending_cloud_job_completion(
    job_uid: str,
    lease_id: str,
    result: Any,
    last_error: str = "",
) -> dict:
    job_uid = str(job_uid or "").strip()
    if not job_uid:
        return {}
    lease_id = str(lease_id or "").strip()
    now = _utc_now_iso()
    result_json = _json_dumps(result if isinstance(result, Mapping) else {"value": result})
    with _get_conn() as conn:
        conn.execute(
            """
            INSERT INTO cloud_job_completion_results (
                job_uid, lease_id, result_json, last_error, attempt_count,
                next_attempt_at, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, 0, '', ?, ?)
            ON CONFLICT(job_uid) DO UPDATE SET
                lease_id = excluded.lease_id,
                result_json = excluded.result_json,
                last_error = excluded.last_error,
                attempt_count = 0,
                next_attempt_at = '',
                updated_at = excluded.updated_at
            """,
            (job_uid, lease_id, result_json, str(last_error or ""), now, now),
        )
        conn.commit()
        row = conn.execute(
            "SELECT * FROM cloud_job_completion_results WHERE job_uid = ?",
            (job_uid,),
        ).fetchone()
    return _cloud_job_completion_from_row(row) or {}


def get_pending_cloud_job_completion(job_uid: str) -> dict | None:
    job_uid = str(job_uid or "").strip()
    if not job_uid:
        return None
    with _get_conn() as conn:
        row = conn.execute("SELECT * FROM cloud_job_completion_results WHERE job_uid = ?", (job_uid,)).fetchone()
    if not row:
        return None
    return _json_loads_object(dict(row).get("result_json"))


def list_pending_cloud_job_completions(limit: int = 20) -> list[dict]:
    now = _utc_now_iso()
    with _get_conn() as conn:
        rows = conn.execute(
            """
            SELECT * FROM cloud_job_completion_results
            WHERE next_attempt_at = '' OR next_attempt_at <= ?
            ORDER BY created_at ASC, job_uid ASC
            LIMIT ?
            """,
            (now, max(1, int(limit or 20))),
        ).fetchall()
    return [entry for row in rows if (entry := _cloud_job_completion_from_row(row)) is not None]


def mark_pending_cloud_job_completion_attempt(job_uid: str, error: str, next_attempt_at: str) -> None:
    job_uid = str(job_uid or "").strip()
    if not job_uid:
        return
    with _get_conn() as conn:
        conn.execute(
            """
            UPDATE cloud_job_completion_results
            SET last_error = ?,
                attempt_count = attempt_count + 1,
                next_attempt_at = ?,
                updated_at = ?
            WHERE job_uid = ?
            """,
            (str(error or ""), str(next_attempt_at or ""), _utc_now_iso(), job_uid),
        )
        conn.commit()


def clear_pending_cloud_job_completion(job_uid: str) -> None:
    job_uid = str(job_uid or "").strip()
    if not job_uid:
        return
    with _get_conn() as conn:
        conn.execute("DELETE FROM cloud_job_completion_results WHERE job_uid = ?", (job_uid,))
        conn.commit()


def _cloud_job_completion_from_row(row: Optional[sqlite3.Row]) -> dict | None:
    if not row:
        return None
    data = dict(row)
    data["result"] = _json_loads_object(data.pop("result_json", "{}"))
    data["attempt_count"] = int(data.get("attempt_count") or 0)
    return data


def record_cloud_job_event(job_uid: str, event_type: str, message: str = "", payload: Optional[Mapping[str, Any]] = None) -> dict:
    now = _now_iso()
    with _get_conn() as conn:
        cursor = conn.execute(
            """
            INSERT INTO cloud_job_events (job_uid, event_type, message, payload_json, created_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (job_uid, event_type, message, _json_dumps(dict(payload or {})), now),
        )
        conn.commit()
        event_id = cursor.lastrowid
        row = conn.execute("SELECT * FROM cloud_job_events WHERE id = ?", (event_id,)).fetchone()
    data = dict(row)
    data["payload"] = _json_loads_object(data.pop("payload_json"))
    return data


def list_cloud_job_events(job_uid: str, limit: int = 100) -> list[dict]:
    with _get_conn() as conn:
        rows = conn.execute(
            """
            SELECT * FROM cloud_job_events
            WHERE job_uid = ?
            ORDER BY id DESC
            LIMIT ?
            """,
            (job_uid, max(1, int(limit or 100))),
        ).fetchall()
    result = []
    for row in reversed(rows):
        data = dict(row)
        data["payload"] = _json_loads_object(data.pop("payload_json"))
        result.append(data)
    return result


def _ai_image_job_from_row(row: Optional[sqlite3.Row]) -> Optional[dict]:
    data = _row_to_dict(row)
    if not data:
        return None
    data["params"] = _json_loads_object(data.pop("params_json", "{}"))
    data["summary"] = _json_loads_object(data.pop("summary_json", "{}"))
    return data


def _ai_image_asset_from_row(row: Optional[sqlite3.Row]) -> Optional[dict]:
    data = _row_to_dict(row)
    if not data:
        return None
    data["meta"] = _json_loads_object(data.pop("meta_json", "{}"))
    return data


def _ai_image_canvas_from_row(row: Optional[sqlite3.Row]) -> Optional[dict]:
    data = _row_to_dict(row)
    if not data:
        return None
    data["canvas"] = _json_loads_object(data.pop("canvas_json", "{}"))
    return data


def create_ai_image_job(payload: Optional[Mapping[str, Any]] = None) -> dict:
    source = dict(payload or {})
    now = _now_iso()
    job_uid = str(source.get("job_uid") or "").strip() or uuid.uuid4().hex
    with _get_conn() as conn:
        conn.execute(
            """
            INSERT INTO ai_image_jobs (
                job_uid, title, prompt, model_key, status, output_dir, params_json,
                summary_json, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                job_uid,
                str(source.get("title") or "").strip() or "未命名生图任务",
                str(source.get("prompt") or "").strip(),
                str(source.get("model_key") or "gpt-image-2").strip() or "gpt-image-2",
                str(source.get("status") or "draft").strip() or "draft",
                str(source.get("output_dir") or "").strip(),
                _json_dumps(source.get("params") if isinstance(source.get("params"), Mapping) else {}),
                _json_dumps(source.get("summary") if isinstance(source.get("summary"), Mapping) else {}),
                now,
                now,
            ),
        )
        conn.commit()
    return get_ai_image_job(job_uid) or {}


def get_ai_image_job(job_uid: str) -> Optional[dict]:
    with _get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM ai_image_jobs WHERE job_uid=? LIMIT 1",
            (str(job_uid or "").strip(),),
        ).fetchone()
    return _ai_image_job_from_row(row)


def list_ai_image_jobs(limit: int = 100) -> list[dict]:
    try:
        safe_limit = max(1, min(int(limit), 500))
    except Exception:
        safe_limit = 100
    with _get_conn() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM ai_image_jobs
            ORDER BY
                CASE WHEN pinned_at <> '' THEN 0 ELSE 1 END,
                pinned_at DESC,
                updated_at DESC,
                id DESC
            LIMIT ?
            """,
            (safe_limit,),
        ).fetchall()
    return [item for row in rows if (item := _ai_image_job_from_row(row))]


def list_active_ai_image_jobs() -> list[dict]:
    with _get_conn() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM ai_image_jobs
            WHERE lower(status) IN ('queued', 'running')
               OR summary_json LIKE '%"status": "queued"%'
               OR summary_json LIKE '%"status": "running"%'
               OR summary_json LIKE '%"status":"queued"%'
               OR summary_json LIKE '%"status":"running"%'
            ORDER BY updated_at DESC, id DESC
            """
        ).fetchall()
    return [item for row in rows if (item := _ai_image_job_from_row(row))]


def update_ai_image_job(job_uid: str, payload: Optional[Mapping[str, Any]] = None) -> dict:
    source = dict(payload or {})
    allowed = {"title", "prompt", "model_key", "status", "output_dir"}
    updates: dict[str, Any] = {}
    for key in allowed:
        if key in source:
            updates[key] = str(source.get(key) or "").strip()
    if "params" in source:
        updates["params_json"] = _json_dumps(source.get("params") if isinstance(source.get("params"), Mapping) else {})
    if "summary" in source:
        updates["summary_json"] = _json_dumps(source.get("summary") if isinstance(source.get("summary"), Mapping) else {})
    updates["updated_at"] = _now_iso()
    uid = str(job_uid or "").strip()
    assignments = ", ".join(f"{key}=?" for key in updates)
    with _get_conn() as conn:
        conn.execute(f"UPDATE ai_image_jobs SET {assignments} WHERE job_uid=?", [*updates.values(), uid])
        conn.commit()
    return get_ai_image_job(uid) or {}


def set_ai_image_job_pinned(job_uid: str, pinned: bool) -> dict:
    uid = str(job_uid or "").strip()
    if not uid:
        return {}
    pinned_at = _utc_now_iso() if bool(pinned) else ""
    with _get_conn() as conn:
        conn.execute(
            "UPDATE ai_image_jobs SET pinned_at=? WHERE job_uid=?",
            (pinned_at, uid),
        )
        conn.commit()
    return get_ai_image_job(uid) or {}


def delete_ai_image_job(job_uid: str) -> bool:
    uid = str(job_uid or "").strip()
    if not uid:
        return False
    with _get_conn() as conn:
        found = conn.execute(
            "SELECT 1 FROM ai_image_jobs WHERE job_uid=? LIMIT 1",
            (uid,),
        ).fetchone()
        if not found:
            return False
        conn.execute("DELETE FROM ai_image_assets WHERE job_uid=?", (uid,))
        conn.execute("DELETE FROM ai_image_canvases WHERE job_uid=?", (uid,))
        conn.execute("DELETE FROM ai_image_jobs WHERE job_uid=?", (uid,))
        conn.commit()
    return True


def create_ai_image_asset(payload: Optional[Mapping[str, Any]] = None) -> dict:
    source = dict(payload or {})
    now = _now_iso()
    asset_uid = str(source.get("asset_uid") or "").strip() or uuid.uuid4().hex
    with _get_conn() as conn:
        conn.execute(
            """
            INSERT INTO ai_image_assets (
                asset_uid, job_uid, kind, source_type, path, url, mime_type,
                sort_order, meta_json, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                asset_uid,
                str(source.get("job_uid") or "").strip(),
                str(source.get("kind") or "reference").strip() or "reference",
                str(source.get("source_type") or "local").strip() or "local",
                str(source.get("path") or "").strip(),
                str(source.get("url") or "").strip(),
                str(source.get("mime_type") or source.get("mime") or "").strip(),
                int(source.get("sort_order") or 0),
                _json_dumps(source.get("meta") if isinstance(source.get("meta"), Mapping) else {}),
                now,
                now,
            ),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM ai_image_assets WHERE asset_uid=?", (asset_uid,)).fetchone()
    return _ai_image_asset_from_row(row) or {}


def list_ai_image_assets(job_uid: str = "", limit: int = 200) -> list[dict]:
    clauses = []
    params: list[Any] = []
    uid = str(job_uid or "").strip()
    if uid:
        clauses.append("job_uid=?")
        params.append(uid)
    where = "WHERE " + " AND ".join(clauses) if clauses else ""
    try:
        safe_limit = max(1, min(int(limit), 1000))
    except Exception:
        safe_limit = 200
    with _get_conn() as conn:
        rows = conn.execute(
            f"""
            SELECT *
            FROM ai_image_assets
            {where}
            ORDER BY sort_order ASC, id ASC
            LIMIT ?
            """,
            [*params, safe_limit],
        ).fetchall()
    return [item for row in rows if (item := _ai_image_asset_from_row(row))]


def create_ai_image_canvas(payload: Optional[Mapping[str, Any]] = None) -> dict:
    source = dict(payload or {})
    now = _now_iso()
    canvas_uid = str(source.get("canvas_uid") or "").strip() or uuid.uuid4().hex
    with _get_conn() as conn:
        conn.execute(
            """
            INSERT INTO ai_image_canvases (
                canvas_uid, job_uid, title, canvas_json, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                canvas_uid,
                str(source.get("job_uid") or "").strip(),
                str(source.get("title") or "").strip() or "未命名画布",
                _json_dumps(source.get("canvas") if isinstance(source.get("canvas"), Mapping) else {}),
                now,
                now,
            ),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM ai_image_canvases WHERE canvas_uid=?", (canvas_uid,)).fetchone()
    return _ai_image_canvas_from_row(row) or {}


def get_ai_image_canvas(canvas_uid: str) -> Optional[dict]:
    with _get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM ai_image_canvases WHERE canvas_uid=? LIMIT 1",
            (str(canvas_uid or "").strip(),),
        ).fetchone()
    return _ai_image_canvas_from_row(row)


def list_ai_image_canvases(job_uid: str = "", limit: int = 100) -> list[dict]:
    uid = str(job_uid or "").strip()
    where = "WHERE job_uid=?" if uid else ""
    params: list[Any] = [uid] if uid else []
    try:
        safe_limit = max(1, min(int(limit), 500))
    except Exception:
        safe_limit = 100
    with _get_conn() as conn:
        rows = conn.execute(
            f"""
            SELECT *
            FROM ai_image_canvases
            {where}
            ORDER BY updated_at DESC, id DESC
            LIMIT ?
            """,
            [*params, safe_limit],
        ).fetchall()
    return [item for row in rows if (item := _ai_image_canvas_from_row(row))]


def _ai_video_job_from_row(row: Optional[sqlite3.Row]) -> Optional[dict]:
    data = _row_to_dict(row)
    if not data:
        return None
    data["id"] = str(data.get("job_uid") or "")
    data["requestUid"] = str(data.get("request_uid") or "")
    data["parameters"] = _json_loads_object(data.pop("parameters_json", "{}"))
    data["currentRunId"] = str(data.get("current_run_uid") or "") or None
    data["outputDir"] = str(data.get("output_dir") or "")
    data["createdAt"] = str(data.get("created_at") or "")
    data["updatedAt"] = str(data.get("updated_at") or "")
    deleted_at = str(data.get("deleted_at") or "").strip()
    data["deletedAt"] = deleted_at or None
    return data


def _ai_video_asset_from_row(row: Optional[sqlite3.Row]) -> Optional[dict]:
    data = _row_to_dict(row)
    if not data:
        return None
    return {
        "id": str(data.get("asset_uid") or ""),
        "jobId": str(data.get("job_uid") or ""),
        "kind": str(data.get("kind") or "image"),
        "role": str(data.get("role") or "reference_image"),
        "sourceType": str(data.get("source_type") or "local_file"),
        "originalName": str(data.get("original_name") or ""),
        "localPath": str(data.get("local_path") or "") or None,
        "remoteUrl": str(data.get("remote_url") or "") or None,
        "mimeType": str(data.get("mime_type") or ""),
        "width": int(data.get("width") or 0),
        "height": int(data.get("height") or 0),
        "sizeBytes": int(data.get("size_bytes") or 0),
        "sha256": str(data.get("sha256") or ""),
        "sortOrder": int(data.get("sort_order") or 0),
        "createdAt": str(data.get("created_at") or ""),
    }


def _ai_video_run_from_row(row: Optional[sqlite3.Row]) -> Optional[dict]:
    data = _row_to_dict(row)
    if not data:
        return None
    output = _json_loads_object(data.pop("output_json", "{}"))
    error = _json_loads_object(data.pop("error_json", "{}"))
    return {
        "id": str(data.get("run_uid") or ""),
        "requestUid": str(data.get("request_uid") or ""),
        "jobId": str(data.get("job_uid") or ""),
        "status": str(data.get("status") or ""),
        "provider": str(data.get("provider") or ""),
        "model": str(data.get("model") or ""),
        "inputSnapshot": _json_loads_object(data.pop("input_snapshot_json", "{}")),
        "providerTaskId": str(data.get("provider_task_id") or "") or None,
        "providerStatus": str(data.get("provider_status") or "") or None,
        "archiveStatus": str(data.get("archive_status") or "none"),
        "output": output or None,
        "error": error or None,
        "createdAt": str(data.get("created_at") or ""),
        "updatedAt": str(data.get("updated_at") or ""),
        "submittedAt": str(data.get("submitted_at") or "") or None,
        "completedAt": str(data.get("completed_at") or "") or None,
    }


def get_ai_video_job_by_request_uid(request_uid: str) -> Optional[dict]:
    uid = str(request_uid or "").strip()
    if not uid:
        return None
    with _get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM ai_video_jobs WHERE request_uid=? LIMIT 1",
            (uid,),
        ).fetchone()
    return _ai_video_job_from_row(row)


def get_ai_video_run_by_request_uid(request_uid: str) -> Optional[dict]:
    uid = str(request_uid or "").strip()
    if not uid:
        return None
    with _get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM ai_video_runs WHERE request_uid=? LIMIT 1",
            (uid,),
        ).fetchone()
    return _ai_video_run_from_row(row)


def create_ai_video_job_with_run(
    job_payload: Optional[Mapping[str, Any]] = None,
    assets: Optional[list[Mapping[str, Any]]] = None,
    run_payload: Optional[Mapping[str, Any]] = None,
) -> dict:
    """Atomically create one Job, its Assets, and the first Run."""
    job_source = dict(job_payload or {})
    run_source = dict(run_payload or {})
    now = _now_iso()
    job_uid = str(job_source.get("id") or job_source.get("job_uid") or "").strip() or f"avj_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:4]}"
    run_uid = str(run_source.get("id") or run_source.get("run_uid") or "").strip() or f"avr_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:4]}"
    job_request_uid = str(job_source.get("requestUid") or job_source.get("request_uid") or "").strip()
    run_request_uid = str(run_source.get("requestUid") or run_source.get("request_uid") or job_request_uid).strip()
    if not job_request_uid or not run_request_uid:
        raise ValueError("requestUid is required")

    parameters = job_source.get("parameters") if isinstance(job_source.get("parameters"), Mapping) else {}
    input_snapshot = run_source.get("inputSnapshot") if isinstance(run_source.get("inputSnapshot"), Mapping) else {}
    with _get_conn() as conn:
        conn.execute("BEGIN IMMEDIATE")
        existing_job_row = conn.execute(
            "SELECT * FROM ai_video_jobs WHERE request_uid=? LIMIT 1",
            (job_request_uid,),
        ).fetchone()
        if existing_job_row:
            existing_job = _ai_video_job_from_row(existing_job_row) or {}
            existing_run_row = None
            current_run_uid = str(existing_job.get("currentRunId") or "")
            if current_run_uid:
                existing_run_row = conn.execute(
                    "SELECT * FROM ai_video_runs WHERE run_uid=? LIMIT 1",
                    (current_run_uid,),
                ).fetchone()
            if not existing_run_row:
                existing_run_row = conn.execute(
                    "SELECT * FROM ai_video_runs WHERE request_uid=? LIMIT 1",
                    (run_request_uid,),
                ).fetchone()
            existing_run = _ai_video_run_from_row(existing_run_row) or {}
            conn.commit()
            return {
                "job": get_ai_video_job(existing_job.get("id") or "") or existing_job,
                "run": existing_run,
                "reused": True,
            }

        existing_run_row = conn.execute(
            "SELECT * FROM ai_video_runs WHERE request_uid=? LIMIT 1",
            (run_request_uid,),
        ).fetchone()
        if existing_run_row:
            raise AiVideoRequestUidConflictError(
                "AI video create requestUid is already owned by a Run without a matching Job"
            )

        conn.execute(
            """
            INSERT INTO ai_video_jobs (
                job_uid, request_uid, title, status, provider, model, prompt,
                parameters_json, current_run_uid, output_dir, deleted_at, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '', ?, ?)
            """,
            (
                job_uid,
                job_request_uid,
                str(job_source.get("title") or "").strip() or "未命名视频任务",
                str(job_source.get("status") or "queued").strip() or "queued",
                str(job_source.get("provider") or "seedance").strip() or "seedance",
                str(job_source.get("model") or "").strip(),
                str(job_source.get("prompt") or "").strip(),
                _json_dumps(parameters),
                run_uid,
                str(job_source.get("outputDir") or job_source.get("output_dir") or "").strip(),
                now,
                now,
            ),
        )
        for index, asset in enumerate(assets or []):
            source = dict(asset or {})
            asset_uid = str(source.get("id") or source.get("asset_uid") or "").strip() or f"ava_{uuid.uuid4().hex[:10]}"
            conn.execute(
                """
                INSERT INTO ai_video_assets (
                    asset_uid, job_uid, kind, role, source_type, original_name, local_path,
                    remote_url, mime_type, width, height, size_bytes, sha256, sort_order, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    asset_uid,
                    job_uid,
                    str(source.get("kind") or "image"),
                    str(source.get("role") or "reference_image"),
                    str(source.get("sourceType") or source.get("source_type") or "local_file"),
                    str(source.get("originalName") or source.get("original_name") or ""),
                    str(source.get("localPath") or source.get("local_path") or ""),
                    str(source.get("remoteUrl") or source.get("remote_url") or ""),
                    str(source.get("mimeType") or source.get("mime_type") or ""),
                    int(source.get("width") or 0),
                    int(source.get("height") or 0),
                    int(source.get("sizeBytes") or source.get("size_bytes") or 0),
                    str(source.get("sha256") or ""),
                    int(source.get("sortOrder") if source.get("sortOrder") is not None else source.get("sort_order") or index),
                    now,
                ),
            )
        conn.execute(
            """
            INSERT INTO ai_video_runs (
                run_uid, request_uid, job_uid, status, provider, model, input_snapshot_json,
                provider_task_id, provider_status, archive_status, output_json, error_json,
                created_at, updated_at, submitted_at, completed_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                run_uid,
                run_request_uid,
                job_uid,
                str(run_source.get("status") or "queued").strip() or "queued",
                str(run_source.get("provider") or job_source.get("provider") or "").strip(),
                str(run_source.get("model") or job_source.get("model") or "").strip(),
                _json_dumps(input_snapshot),
                str(run_source.get("providerTaskId") or run_source.get("provider_task_id") or ""),
                str(run_source.get("providerStatus") or run_source.get("provider_status") or ""),
                str(run_source.get("archiveStatus") or run_source.get("archive_status") or "none"),
                _json_dumps(run_source.get("output") if isinstance(run_source.get("output"), Mapping) else {}),
                _json_dumps(run_source.get("error") if isinstance(run_source.get("error"), Mapping) else {}),
                now,
                now,
                str(run_source.get("submittedAt") or run_source.get("submitted_at") or ""),
                str(run_source.get("completedAt") or run_source.get("completed_at") or ""),
            ),
        )
        conn.commit()
    return {
        "job": get_ai_video_job(job_uid) or {},
        "run": get_ai_video_run(run_uid) or {},
        "reused": False,
    }


def get_ai_video_job(job_uid: str) -> Optional[dict]:
    uid = str(job_uid or "").strip()
    if not uid:
        return None
    with _get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM ai_video_jobs WHERE job_uid=? LIMIT 1",
            (uid,),
        ).fetchone()
    job = _ai_video_job_from_row(row)
    if not job:
        return None
    job["assets"] = list_ai_video_assets(uid)
    job["runs"] = list_ai_video_runs(uid)
    return job


def list_ai_video_jobs(
    *,
    status: str = "",
    provider: str = "",
    limit: int = 50,
    include_deleted: bool = False,
) -> list[dict]:
    try:
        safe_limit = max(1, min(int(limit), 200))
    except Exception:
        safe_limit = 50
    clauses: list[str] = []
    params: list[Any] = []
    if not include_deleted:
        clauses.append("(deleted_at IS NULL OR deleted_at = '')")
    status_value = str(status or "").strip()
    if status_value and status_value != "all":
        clauses.append("status=?")
        params.append(status_value)
    provider_value = str(provider or "").strip()
    if provider_value:
        clauses.append("provider=?")
        params.append(provider_value)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    with _get_conn() as conn:
        rows = conn.execute(
            f"""
            SELECT *
            FROM ai_video_jobs
            {where}
            ORDER BY updated_at DESC, id DESC
            LIMIT ?
            """,
            [*params, safe_limit],
        ).fetchall()
    jobs = []
    for row in rows:
        job = _ai_video_job_from_row(row)
        if not job:
            continue
        job["assets"] = list_ai_video_assets(job["id"])
        current_run = get_ai_video_run(job.get("currentRunId") or "")
        job["currentRun"] = current_run
        jobs.append(job)
    return jobs


def list_active_ai_video_runs() -> list[dict]:
    with _get_conn() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM ai_video_runs
            WHERE lower(status) IN ('queued', 'running', 'downloading')
            ORDER BY updated_at ASC, id ASC
            """
        ).fetchall()
    return [item for row in rows if (item := _ai_video_run_from_row(row))]


def update_ai_video_job(job_uid: str, payload: Optional[Mapping[str, Any]] = None) -> dict:
    source = dict(payload or {})
    uid = str(job_uid or "").strip()
    if not uid:
        return {}
    updates: dict[str, Any] = {}
    mapping = {
        "title": "title",
        "status": "status",
        "provider": "provider",
        "model": "model",
        "prompt": "prompt",
        "outputDir": "output_dir",
        "output_dir": "output_dir",
        "currentRunId": "current_run_uid",
        "current_run_uid": "current_run_uid",
        "deletedAt": "deleted_at",
        "deleted_at": "deleted_at",
    }
    for src_key, column in mapping.items():
        if src_key in source:
            value = source.get(src_key)
            updates[column] = "" if value is None else str(value).strip()
    if "parameters" in source:
        updates["parameters_json"] = _json_dumps(source.get("parameters") if isinstance(source.get("parameters"), Mapping) else {})
    updates["updated_at"] = _now_iso()
    assignments = ", ".join(f"{key}=?" for key in updates)
    with _get_conn() as conn:
        conn.execute(f"UPDATE ai_video_jobs SET {assignments} WHERE job_uid=?", [*updates.values(), uid])
        conn.commit()
    return get_ai_video_job(uid) or {}


def replace_ai_video_assets(job_uid: str, assets: Optional[list[Mapping[str, Any]]] = None) -> list[dict]:
    uid = str(job_uid or "").strip()
    if not uid:
        return []
    now = _now_iso()
    with _get_conn() as conn:
        conn.execute("DELETE FROM ai_video_assets WHERE job_uid=?", (uid,))
        for index, asset in enumerate(assets or []):
            source = dict(asset or {})
            asset_uid = str(source.get("id") or source.get("asset_uid") or "").strip() or f"ava_{uuid.uuid4().hex[:10]}"
            conn.execute(
                """
                INSERT INTO ai_video_assets (
                    asset_uid, job_uid, kind, role, source_type, original_name, local_path,
                    remote_url, mime_type, width, height, size_bytes, sha256, sort_order, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    asset_uid,
                    uid,
                    str(source.get("kind") or "image"),
                    str(source.get("role") or "reference_image"),
                    str(source.get("sourceType") or source.get("source_type") or "local_file"),
                    str(source.get("originalName") or source.get("original_name") or ""),
                    str(source.get("localPath") or source.get("local_path") or ""),
                    str(source.get("remoteUrl") or source.get("remote_url") or ""),
                    str(source.get("mimeType") or source.get("mime_type") or ""),
                    int(source.get("width") or 0),
                    int(source.get("height") or 0),
                    int(source.get("sizeBytes") or source.get("size_bytes") or 0),
                    str(source.get("sha256") or ""),
                    int(source.get("sortOrder") if source.get("sortOrder") is not None else source.get("sort_order") or index),
                    now,
                ),
            )
        conn.execute("UPDATE ai_video_jobs SET updated_at=? WHERE job_uid=?", (now, uid))
        conn.commit()
    return list_ai_video_assets(uid)


def list_ai_video_assets(job_uid: str = "") -> list[dict]:
    uid = str(job_uid or "").strip()
    with _get_conn() as conn:
        if uid:
            rows = conn.execute(
                """
                SELECT * FROM ai_video_assets
                WHERE job_uid=?
                ORDER BY sort_order ASC, id ASC
                """,
                (uid,),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM ai_video_assets ORDER BY sort_order ASC, id ASC LIMIT 500"
            ).fetchall()
    return [item for row in rows if (item := _ai_video_asset_from_row(row))]


def create_ai_video_run(payload: Optional[Mapping[str, Any]] = None) -> dict:
    source = dict(payload or {})
    now = _now_iso()
    run_uid = str(source.get("id") or source.get("run_uid") or "").strip() or f"avr_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:4]}"
    request_uid = str(source.get("requestUid") or source.get("request_uid") or "").strip()
    if not request_uid:
        raise ValueError("requestUid is required")
    job_uid = str(source.get("jobId") or source.get("job_uid") or "").strip()
    with _get_conn() as conn:
        conn.execute("BEGIN IMMEDIATE")
        existing_row = conn.execute(
            "SELECT * FROM ai_video_runs WHERE request_uid=? LIMIT 1",
            (request_uid,),
        ).fetchone()
        if existing_row:
            existing = _ai_video_run_from_row(existing_row) or {}
            if str(existing.get("jobId") or "").strip() != job_uid:
                raise AiVideoRequestUidConflictError("AI video requestUid belongs to another Job")
            conn.commit()
            existing["_reused"] = True
            return existing
        if "expectedCurrentRunId" in source or "expectedJobStatus" in source:
            expected_current_run_uid = str(source.get("expectedCurrentRunId") or "").strip()
            expected_job_status = str(source.get("expectedJobStatus") or "").strip()
            job_row = conn.execute(
                "SELECT status, current_run_uid, deleted_at FROM ai_video_jobs WHERE job_uid=? LIMIT 1",
                (job_uid,),
            ).fetchone()
            if (
                not job_row
                or str(job_row["deleted_at"] or "").strip()
                or str(job_row["status"] or "").strip() != expected_job_status
                or str(job_row["current_run_uid"] or "").strip() != expected_current_run_uid
            ):
                raise AiVideoRetryConflictError("AI video Job changed before retry Run creation")
        conn.execute(
            """
            INSERT INTO ai_video_runs (
                run_uid, request_uid, job_uid, status, provider, model, input_snapshot_json,
                provider_task_id, provider_status, archive_status, output_json, error_json,
                created_at, updated_at, submitted_at, completed_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                run_uid,
                request_uid,
                job_uid,
                str(source.get("status") or "queued").strip() or "queued",
                str(source.get("provider") or "").strip(),
                str(source.get("model") or "").strip(),
                _json_dumps(source.get("inputSnapshot") if isinstance(source.get("inputSnapshot"), Mapping) else {}),
                str(source.get("providerTaskId") or source.get("provider_task_id") or ""),
                str(source.get("providerStatus") or source.get("provider_status") or ""),
                str(source.get("archiveStatus") or source.get("archive_status") or "none"),
                _json_dumps(source.get("output") if isinstance(source.get("output"), Mapping) else {}),
                _json_dumps(source.get("error") if isinstance(source.get("error"), Mapping) else {}),
                now,
                now,
                str(source.get("submittedAt") or source.get("submitted_at") or ""),
                str(source.get("completedAt") or source.get("completed_at") or ""),
            ),
        )
        if job_uid:
            conn.execute(
                "UPDATE ai_video_jobs SET current_run_uid=?, status=?, updated_at=? WHERE job_uid=?",
                (run_uid, str(source.get("status") or "queued"), now, job_uid),
            )
        conn.commit()
    created = get_ai_video_run(run_uid) or {}
    created["_reused"] = False
    return created


def get_ai_video_run(run_uid: str) -> Optional[dict]:
    uid = str(run_uid or "").strip()
    if not uid:
        return None
    with _get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM ai_video_runs WHERE run_uid=? LIMIT 1",
            (uid,),
        ).fetchone()
    return _ai_video_run_from_row(row)


def list_ai_video_runs(job_uid: str = "", limit: int = 50) -> list[dict]:
    uid = str(job_uid or "").strip()
    try:
        safe_limit = max(1, min(int(limit), 200))
    except Exception:
        safe_limit = 50
    with _get_conn() as conn:
        if uid:
            rows = conn.execute(
                """
                SELECT * FROM ai_video_runs
                WHERE job_uid=?
                ORDER BY created_at DESC, id DESC
                LIMIT ?
                """,
                (uid, safe_limit),
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT * FROM ai_video_runs
                ORDER BY created_at DESC, id DESC
                LIMIT ?
                """,
                (safe_limit,),
            ).fetchall()
    return [item for row in rows if (item := _ai_video_run_from_row(row))]


def update_ai_video_run(run_uid: str, payload: Optional[Mapping[str, Any]] = None) -> dict:
    """Update mutable execution fields only; never rewrite input snapshot."""
    source = dict(payload or {})
    uid = str(run_uid or "").strip()
    if not uid:
        return {}
    updates: dict[str, Any] = {}
    mapping = {
        "status": "status",
        "providerTaskId": "provider_task_id",
        "provider_task_id": "provider_task_id",
        "providerStatus": "provider_status",
        "provider_status": "provider_status",
        "archiveStatus": "archive_status",
        "archive_status": "archive_status",
        "submittedAt": "submitted_at",
        "submitted_at": "submitted_at",
        "completedAt": "completed_at",
        "completed_at": "completed_at",
    }
    for src_key, column in mapping.items():
        if src_key in source:
            value = source.get(src_key)
            updates[column] = "" if value is None else str(value).strip()
    if "output" in source:
        updates["output_json"] = _json_dumps(source.get("output") if isinstance(source.get("output"), Mapping) else {})
    if "error" in source:
        updates["error_json"] = _json_dumps(source.get("error") if isinstance(source.get("error"), Mapping) else {})
    updates["updated_at"] = _now_iso()
    assignments = ", ".join(f"{key}=?" for key in updates)
    with _get_conn() as conn:
        conn.execute(f"UPDATE ai_video_runs SET {assignments} WHERE run_uid=?", [*updates.values(), uid])
        if "status" in source:
            job_uid = conn.execute(
                "SELECT job_uid FROM ai_video_runs WHERE run_uid=? LIMIT 1",
                (uid,),
            ).fetchone()
            if job_uid and str(job_uid["job_uid"] or "").strip():
                jid = str(job_uid["job_uid"])
                current = conn.execute(
                    "SELECT current_run_uid FROM ai_video_jobs WHERE job_uid=? LIMIT 1",
                    (jid,),
                ).fetchone()
                if current and str(current["current_run_uid"] or "") == uid:
                    conn.execute(
                        "UPDATE ai_video_jobs SET status=?, updated_at=? WHERE job_uid=?",
                        (str(source.get("status") or "").strip(), updates["updated_at"], jid),
                    )
        conn.commit()
    return get_ai_video_run(uid) or {}


def cancel_ai_video_run_if_unsubmitted(run_uid: str) -> Optional[dict]:
    """Atomically cancel the current queued Run only before provider submission."""
    uid = str(run_uid or "").strip()
    if not uid:
        return None
    now = _now_iso()
    with _get_conn() as conn:
        conn.execute("BEGIN IMMEDIATE")
        run_row = conn.execute(
            "SELECT * FROM ai_video_runs WHERE run_uid=? LIMIT 1",
            (uid,),
        ).fetchone()
        if (
            not run_row
            or str(run_row["status"] or "").strip() != "queued"
            or str(run_row["provider_task_id"] or "").strip()
            or str(run_row["submitted_at"] or "").strip()
            or str(run_row["provider_status"] or "").strip()
        ):
            conn.commit()
            return None
        job_uid = str(run_row["job_uid"] or "").strip()
        job_row = conn.execute(
            "SELECT status, current_run_uid, deleted_at FROM ai_video_jobs WHERE job_uid=? LIMIT 1",
            (job_uid,),
        ).fetchone()
        if (
            not job_row
            or str(job_row["status"] or "").strip() != "queued"
            or str(job_row["current_run_uid"] or "").strip() != uid
            or str(job_row["deleted_at"] or "").strip()
        ):
            conn.commit()
            return None
        conn.execute(
            "UPDATE ai_video_runs SET status='cancelled', error_json='{}', updated_at=? WHERE run_uid=?",
            (now, uid),
        )
        conn.execute(
            "UPDATE ai_video_jobs SET status='cancelled', updated_at=? WHERE job_uid=? AND current_run_uid=?",
            (now, job_uid, uid),
        )
        conn.commit()
    return get_ai_video_run(uid)


def soft_delete_ai_video_job(job_uid: str) -> bool:
    uid = str(job_uid or "").strip()
    if not uid:
        return False
    job = get_ai_video_job(uid)
    if not job:
        return False
    update_ai_video_job(uid, {"deletedAt": _now_iso()})
    return True


def create_task_instance(adapter_id: str, task_id: str, title: str, params: Optional[Mapping[str, Any]] = None) -> dict:
    """Insert a draft task instance and return the inserted row."""
    now = _now_iso()
    instance_uid = uuid.uuid4().hex
    with _get_conn() as conn:
        conn.execute("""
            INSERT INTO task_instances (
                instance_uid, adapter_id, task_id, title, status, current_step,
                params_json, summary_json, archived, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, 'draft', 'config', ?, '{}', 0, ?, ?)
        """, (
            instance_uid,
            str(adapter_id or "").strip(),
            str(task_id or "").strip(),
            str(title or "").strip() or "未命名任务",
            _json_dumps(dict(params or {})),
            now,
            now,
        ))
        conn.commit()
    return get_task_instance(instance_uid) or {}


def get_task_instance(instance_uid: str) -> Optional[dict]:
    """Return one task instance row by uid."""
    with _get_conn() as conn:
        row = conn.execute("""
            SELECT *
            FROM task_instances
            WHERE instance_uid=?
            LIMIT 1
        """, (str(instance_uid or "").strip(),)).fetchone()
        return _row_to_dict(row)


def get_task_instance_detail(instance_uid: str) -> dict:
    """Return one task instance plus parsed params, parsed summary, runs, artifacts, and events."""
    uid = str(instance_uid or "").strip()
    instance = get_task_instance(uid)
    if not instance:
        return {}

    with _get_conn() as conn:
        runs = conn.execute("""
            SELECT tir.*, tr.adapter_id, tr.task_id, tr.status, tr.started_at, tr.finished_at,
                   tr.records_count, tr.error, tr.output_files
            FROM task_instance_runs tir
            LEFT JOIN task_runs tr ON tr.id = tir.run_id
            WHERE tir.instance_uid=?
            ORDER BY tir.id DESC
        """, (uid,)).fetchall()
        artifacts = conn.execute("""
            SELECT *
            FROM task_instance_artifacts
            WHERE instance_uid=?
            ORDER BY id ASC
        """, (uid,)).fetchall()
        events = conn.execute("""
            SELECT *
            FROM task_instance_events
            WHERE instance_uid=?
            ORDER BY id ASC
        """, (uid,)).fetchall()

    detail = dict(instance)
    detail["params"] = _json_loads_object(detail.get("params_json"))
    detail["summary"] = _json_loads_object(detail.get("summary_json"))
    detail["runs"] = [dict(row) for row in runs]
    detail["artifacts"] = [dict(row) for row in artifacts]
    detail["events"] = [dict(row) for row in events]
    for artifact in detail["artifacts"]:
        artifact["meta"] = _json_loads_object(artifact.get("meta_json"))
    for event in detail["events"]:
        event["meta"] = _json_loads_object(event.get("meta_json"))
    return detail


def list_task_instances(
    status_group: str = "",
    adapter_id: str = "",
    task_id: str = "",
    keyword: str = "",
    limit: int = 100,
) -> list[dict]:
    """Return task instances filtered by status group and metadata."""
    status_groups = {
        "current": ("draft", "queued", "running", "generating", "creating", "waiting_approval"),
        "pending": ("waiting_approval", "failed", "create_failed", "partial_failed"),
        "history": ("completed", "stopped", "archived"),
    }
    clauses = []
    params: list[Any] = []

    group = str(status_group or "").strip()
    statuses = status_groups.get(group)
    if statuses:
        placeholders = ",".join("?" for _ in statuses)
        if group == "history":
            clauses.append(f"(status IN ({placeholders}) OR archived=1)")
        else:
            clauses.append(f"status IN ({placeholders})")
            clauses.append("archived=0")
        params.extend(statuses)
    elif group:
        clauses.append("status=?")
        clauses.append("archived=0")
        params.append(group)
    else:
        clauses.append("archived=0")

    adapter = str(adapter_id or "").strip()
    if adapter:
        clauses.append("adapter_id=?")
        params.append(adapter)

    task = str(task_id or "").strip()
    if task:
        clauses.append("task_id=?")
        params.append(task)

    term = str(keyword or "").strip()
    if term:
        clauses.append("(title LIKE ? OR instance_uid LIKE ? OR task_id LIKE ?)")
        like = f"%{term}%"
        params.extend([like, like, like])

    try:
        safe_limit = max(1, min(int(limit), 500))
    except Exception:
        safe_limit = 100

    where = " AND ".join(clauses) if clauses else "1=1"
    with _get_conn() as conn:
        rows = conn.execute(f"""
            SELECT *
            FROM task_instances
            WHERE {where}
            ORDER BY updated_at DESC, id DESC
            LIMIT ?
        """, [*params, safe_limit]).fetchall()
        return [dict(row) for row in rows]


def find_task_instance_by_approval_batch_id(batch_id: str) -> Optional[dict]:
    """Return the first task instance whose summary stores the approval batch id."""
    target = str(batch_id or "").strip()
    if not target:
        return None
    with _get_conn() as conn:
        rows = conn.execute("""
            SELECT *
            FROM task_instances
            WHERE summary_json LIKE ?
            ORDER BY updated_at DESC, id DESC
            LIMIT 20
        """, (f"%{target}%",)).fetchall()
    for row in rows:
        item = dict(row)
        summary = _json_loads_object(item.get("summary_json"))
        if str(summary.get("approval_batch_id") or "").strip() == target:
            return item
    return None


def update_task_instance(instance_uid: str, **fields) -> dict:
    """Update allowed task instance fields and return the updated row."""
    allowed = {
        "title",
        "status",
        "current_step",
        "params_json",
        "summary_json",
        "last_run_id",
        "archived",
        "completed_at",
    }
    updates = {}
    for key, value in fields.items():
        if key == "params":
            updates["params_json"] = _json_dumps(value)
        elif key == "summary":
            updates["summary_json"] = _json_dumps(value)
        elif key == "archived":
            archived = 1 if value else 0
            updates["archived"] = archived
            if archived:
                updates.setdefault("status", "archived")
                updates.setdefault("completed_at", _now_iso())
        elif key in allowed:
            updates[key] = value

    status = str(updates.get("status") or "").strip()
    if status in {"completed", "stopped", "archived"} and not updates.get("completed_at"):
        updates["completed_at"] = _now_iso()

    updates["updated_at"] = _now_iso()
    assignments = ", ".join(f"{key}=?" for key in updates.keys())
    values = list(updates.values())
    uid = str(instance_uid or "").strip()
    with _get_conn() as conn:
        conn.execute(
            f"UPDATE task_instances SET {assignments} WHERE instance_uid=?",
            [*values, uid],
        )
        conn.commit()
    return get_task_instance(uid) or {}


def link_task_instance_run(instance_uid: str, run_id: int, purpose: str = "main") -> None:
    """Associate a task_runs row with a task instance."""
    uid = str(instance_uid or "").strip()
    rid = int(run_id)
    now = _now_iso()
    with _get_conn() as conn:
        conn.execute("""
            INSERT OR IGNORE INTO task_instance_runs (instance_uid, run_id, purpose, created_at)
            VALUES (?, ?, ?, ?)
        """, (uid, rid, str(purpose or "main").strip() or "main", now))
        conn.execute("""
            UPDATE task_instances
            SET last_run_id=?, updated_at=?
            WHERE instance_uid=?
        """, (rid, now, uid))
        conn.commit()


def add_task_instance_artifact(
    instance_uid: str,
    kind: str,
    label: str,
    path: str,
    meta: Optional[Mapping[str, Any]] = None,
) -> dict:
    """Insert or update an artifact row and return it."""
    uid = str(instance_uid or "").strip()
    path_text = str(path or "").strip()
    now = _now_iso()
    with _get_conn() as conn:
        existing = None
        if path_text:
            existing = conn.execute("""
                SELECT id
                FROM task_instance_artifacts
                WHERE instance_uid=? AND path=?
                ORDER BY id ASC
                LIMIT 1
            """, (uid, path_text)).fetchone()
        if existing:
            artifact_id = int(existing["id"])
            conn.execute("""
                UPDATE task_instance_artifacts
                SET kind=?, label=?, meta_json=?
                WHERE id=?
            """, (
                str(kind or "").strip() or "file",
                str(label or "").strip() or "输出文件",
                _json_dumps(dict(meta or {})),
                artifact_id,
            ))
            conn.execute("UPDATE task_instances SET updated_at=? WHERE instance_uid=?", (now, uid))
            conn.commit()
            row = conn.execute("SELECT * FROM task_instance_artifacts WHERE id=?", (artifact_id,)).fetchone()
            return dict(row) if row else {}
        cur = conn.execute("""
            INSERT INTO task_instance_artifacts (instance_uid, kind, label, path, meta_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            uid,
            str(kind or "").strip() or "file",
            str(label or "").strip() or "输出文件",
            path_text,
            _json_dumps(dict(meta or {})),
            now,
        ))
        conn.execute("UPDATE task_instances SET updated_at=? WHERE instance_uid=?", (now, uid))
        conn.commit()
        row = conn.execute("SELECT * FROM task_instance_artifacts WHERE id=?", (cur.lastrowid,)).fetchone()
        return dict(row) if row else {}


def add_task_instance_event(
    instance_uid: str,
    event_type: str,
    message: str,
    meta: Optional[Mapping[str, Any]] = None,
) -> dict:
    """Insert an event row and return it."""
    uid = str(instance_uid or "").strip()
    with _get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO task_instance_events (instance_uid, event_type, message, meta_json, created_at)
            VALUES (?, ?, ?, ?, ?)
        """, (
            uid,
            str(event_type or "").strip() or "event",
            str(message or "").strip(),
            _json_dumps(dict(meta or {})),
            _now_iso(),
        ))
        conn.execute("UPDATE task_instances SET updated_at=? WHERE instance_uid=?", (_now_iso(), uid))
        conn.commit()
        row = conn.execute("SELECT * FROM task_instance_events WHERE id=?", (cur.lastrowid,)).fetchone()
        return dict(row) if row else {}


def create_task_schedule(
    adapter_id: str,
    task_id: str,
    title: str,
    frequency: str,
    time_of_day: str,
    weekday: Optional[int] = None,
    params: Optional[Mapping[str, Any]] = None,
    notify_channel: str = "dingtalk",
    notify_template: str = "",
    enabled: bool = True,
) -> dict:
    """Insert a persisted schedule definition and return it."""
    now = _now_iso()
    schedule_uid = uuid.uuid4().hex
    with _get_conn() as conn:
        conn.execute("""
            INSERT INTO task_schedules (
                schedule_uid, adapter_id, task_id, title, enabled, frequency,
                time_of_day, weekday, params_json, notify_channel, notify_template,
                archived, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
        """, (
            schedule_uid,
            str(adapter_id or "").strip(),
            str(task_id or "").strip(),
            str(title or "").strip() or "未命名定时任务",
            1 if enabled else 0,
            str(frequency or "").strip(),
            str(time_of_day or "").strip(),
            weekday,
            _json_dumps(dict(params or {})),
            str(notify_channel or "").strip() or "dingtalk",
            str(notify_template or "").strip(),
            now,
            now,
        ))
        conn.commit()
    return get_task_schedule(schedule_uid) or {}


def get_task_schedule(schedule_uid: str) -> Optional[dict]:
    """Return one schedule definition by uid."""
    with _get_conn() as conn:
        row = conn.execute("""
            SELECT *
            FROM task_schedules
            WHERE schedule_uid=?
            LIMIT 1
        """, (str(schedule_uid or "").strip(),)).fetchone()
        return _row_to_dict(row)


def get_task_schedule_detail(schedule_uid: str) -> dict:
    """Return one schedule plus parsed params."""
    schedule = get_task_schedule(schedule_uid)
    if not schedule:
        return {}
    detail = dict(schedule)
    detail["params"] = _json_loads_object(detail.get("params_json"))
    return detail


def list_task_schedules(
    adapter_id: str = "",
    task_id: str = "",
    enabled: Optional[bool] = None,
    keyword: str = "",
    include_archived: bool = False,
    limit: int = 100,
) -> list[dict]:
    """Return persisted schedule definitions filtered by metadata."""
    clauses = []
    params: list[Any] = []

    if not include_archived:
        clauses.append("archived=0")

    adapter = str(adapter_id or "").strip()
    if adapter:
        clauses.append("adapter_id=?")
        params.append(adapter)

    task = str(task_id or "").strip()
    if task:
        clauses.append("task_id=?")
        params.append(task)

    if enabled is not None:
        clauses.append("enabled=?")
        params.append(1 if enabled else 0)

    term = str(keyword or "").strip()
    if term:
        clauses.append("(title LIKE ? OR schedule_uid LIKE ? OR task_id LIKE ?)")
        like = f"%{term}%"
        params.extend([like, like, like])

    try:
        safe_limit = max(1, min(int(limit), 500))
    except Exception:
        safe_limit = 100

    where = " AND ".join(clauses) if clauses else "1=1"
    with _get_conn() as conn:
        try:
            rows = conn.execute(f"""
                SELECT *
                FROM task_schedules
                WHERE {where}
                ORDER BY updated_at DESC, id DESC
                LIMIT ?
            """, [*params, safe_limit]).fetchall()
        except sqlite3.OperationalError as exc:
            if "no such table: task_schedules" in str(exc):
                return []
            raise
        return [dict(row) for row in rows]


def update_task_schedule(schedule_uid: str, **fields) -> dict:
    """Update allowed schedule fields and return the updated row."""
    allowed = {
        "title",
        "enabled",
        "frequency",
        "time_of_day",
        "weekday",
        "params_json",
        "notify_channel",
        "notify_template",
        "last_run_id",
        "last_instance_uid",
        "last_status",
        "last_error",
        "last_triggered_at",
        "archived",
    }
    updates: dict[str, Any] = {}
    for key, value in fields.items():
        if key == "params":
            updates["params_json"] = _json_dumps(value)
        elif key in {"enabled", "archived"}:
            updates[key] = 1 if value else 0
        elif key in allowed:
            updates[key] = value

    updates["updated_at"] = _now_iso()
    assignments = ", ".join(f"{key}=?" for key in updates.keys())
    values = list(updates.values())
    uid = str(schedule_uid or "").strip()
    with _get_conn() as conn:
        conn.execute(
            f"UPDATE task_schedules SET {assignments} WHERE schedule_uid=?",
            [*values, uid],
        )
        conn.commit()
    return get_task_schedule(uid) or {}


def archive_task_schedule(schedule_uid: str) -> dict:
    """Archive a persisted schedule and disable it."""
    return update_task_schedule(schedule_uid, archived=True, enabled=False)


def record_task_schedule_run(
    schedule_uid: str,
    run_id: Optional[int] = None,
    instance_uid: str = "",
    status: str = "",
    error: str = "",
) -> dict:
    """Store the latest execution result on a schedule definition."""
    fields: dict[str, Any] = {
        "last_triggered_at": _now_iso(),
        "last_instance_uid": str(instance_uid or "").strip(),
        "last_status": str(status or "").strip(),
        "last_error": str(error or "").strip(),
    }
    if run_id is not None:
        fields["last_run_id"] = int(run_id)
    return update_task_schedule(schedule_uid, **fields)


def begin_run(adapter_id: str, task_id: str) -> int:
    """Record a task run start, return run_id"""
    now = datetime.now().isoformat()
    with _get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO task_runs (adapter_id, task_id, status, started_at, last_seen_at, phase, current_row)
            VALUES (?, ?, 'running', ?, ?, '', 0)
        """, (adapter_id, task_id, now, now))
        conn.commit()
        return cur.lastrowid


def heartbeat_run(run_id: int, phase: str = "", current_row: int = 0, records_count: Optional[int] = None):
    fields = [
        "last_seen_at=?",
        "phase=?",
        "current_row=?",
    ]
    now = datetime.now().isoformat()
    values: list[Any] = [
        now,
        str(phase or "").strip(),
        max(0, int(current_row or 0)),
    ]
    if records_count is not None:
        fields.append("records_count=?")
        values.append(max(0, int(records_count or 0)))
    values.append(int(run_id))
    with _get_conn() as conn:
        conn.execute(f"""
            UPDATE task_runs
            SET {", ".join(fields)}
            WHERE id=?
        """, values)
        conn.commit()


def finish_run(run_id: int, records_count: int, output_files: List[str]):
    now = datetime.now().isoformat()
    with _get_conn() as conn:
        conn.execute("""
            UPDATE task_runs
            SET status='done', finished_at=?, last_seen_at=?, records_count=?, output_files=?, error=NULL
            WHERE id=?
        """, (now, now, records_count, json.dumps(output_files), run_id))
        conn.commit()


def fail_run(run_id: int, error: str, records_count: int = 0, output_files: Optional[List[str]] = None):
    now = datetime.now().isoformat()
    with _get_conn() as conn:
        conn.execute("""
            UPDATE task_runs
            SET status='error', finished_at=?, last_seen_at=?, records_count=?, output_files=?, error=?
            WHERE id=?
        """, (
            now,
            now,
            max(0, int(records_count or 0)),
            json.dumps(output_files or []),
            error,
            run_id,
        ))
        conn.commit()


def stop_run(run_id: int, records_count: int, output_files: List[str], error: str = ""):
    now = datetime.now().isoformat()
    with _get_conn() as conn:
        conn.execute("""
            UPDATE task_runs
            SET status='stopped', finished_at=?, last_seen_at=?, records_count=?, output_files=?, error=?
            WHERE id=?
        """, (now, now, records_count, json.dumps(output_files), error, run_id))
        conn.commit()


def stop_orphaned_active_runs(error: str = "任务运行时后端已重启，已自动标记为停止") -> int:
    """Mark runs that cannot have an in-memory worker after backend startup as stopped."""
    with _get_conn() as conn:
        cur = conn.execute("""
            UPDATE task_runs
            SET status='stopped', finished_at=?, last_seen_at=?, error=COALESCE(NULLIF(error, ''), ?)
            WHERE status IN ('running', 'pausing', 'paused', 'stopping')
        """, (datetime.now().isoformat(), datetime.now().isoformat(), error))
        conn.commit()
        return int(cur.rowcount or 0)


def list_active_runs(statuses: Optional[Iterable[str]] = None) -> List[dict]:
    """Return runs that are still considered active by the desktop runtime."""
    active_statuses = [
        str(item or "").strip()
        for item in (statuses or ("running", "pausing", "paused", "stopping"))
        if str(item or "").strip()
    ]
    if not active_statuses:
        return []
    placeholders = ",".join("?" for _ in active_statuses)
    with _get_conn() as conn:
        rows = conn.execute(f"""
            SELECT *
            FROM task_runs
            WHERE status IN ({placeholders})
        """, active_statuses).fetchall()
        return [dict(row) for row in rows]


def get_latest_run(adapter_id: str, task_id: str) -> Optional[dict]:
    with _get_conn() as conn:
        row = conn.execute("""
            SELECT * FROM task_runs
            WHERE adapter_id=? AND task_id=?
            ORDER BY id DESC LIMIT 1
        """, (adapter_id, task_id)).fetchone()
        return dict(row) if row else None


def get_latest_task_instance_run(instance_uid: str) -> Optional[dict]:
    """Return the latest task_runs row linked to one task instance."""
    uid = str(instance_uid or "").strip()
    if not uid:
        return None
    with _get_conn() as conn:
        row = conn.execute("""
            SELECT tr.*, tir.purpose, tir.created_at AS linked_at
            FROM task_instance_runs tir
            LEFT JOIN task_runs tr ON tr.id = tir.run_id
            WHERE tir.instance_uid=?
            ORDER BY tir.id DESC
            LIMIT 1
        """, (uid,)).fetchone()
        return dict(row) if row else None


def list_runs(adapter_id: str, task_id: str, limit: int = 20) -> List[dict]:
    with _get_conn() as conn:
        rows = conn.execute("""
            SELECT * FROM task_runs
            WHERE adapter_id=? AND task_id=?
            ORDER BY id DESC LIMIT ?
        """, (adapter_id, task_id, limit)).fetchall()
        return [dict(r) for r in rows]


def _normalize_output_file_path(path: str) -> str:
    try:
        return str(Path(str(path)).expanduser().resolve(strict=False))
    except Exception:
        return str(path or '').strip()


def remove_output_files(paths: List[str]) -> dict:
    target_paths = {
        _normalize_output_file_path(path)
        for path in (paths or [])
        if str(path or '').strip()
    }
    target_paths.discard('')
    if not target_paths:
        return {"updated_runs": 0, "removed_refs": 0}

    updated_runs = 0
    removed_refs = 0

    with _get_conn() as conn:
        rows = conn.execute("""
            SELECT id, output_files
            FROM task_runs
            WHERE output_files IS NOT NULL AND output_files != '[]'
        """).fetchall()

        for row in rows:
            raw_files = row["output_files"]
            try:
                files = json.loads(raw_files) if isinstance(raw_files, str) else raw_files
            except Exception:
                continue
            if not isinstance(files, list):
                continue

            kept_files = []
            changed = False
            for item in files:
                file_path = str(item or '').strip()
                if not file_path:
                    changed = True
                    continue
                if _normalize_output_file_path(file_path) in target_paths:
                    removed_refs += 1
                    changed = True
                    continue
                kept_files.append(file_path)

            if changed:
                conn.execute(
                    "UPDATE task_runs SET output_files=? WHERE id=?",
                    (json.dumps(kept_files), row["id"]),
                )
                updated_runs += 1

        conn.commit()

    return {"updated_runs": updated_runs, "removed_refs": removed_refs}


def is_output_file_path(path: str) -> bool:
    target_path = _normalize_output_file_path(path)
    if not target_path:
        return False

    with _get_conn() as conn:
        rows = conn.execute("""
            SELECT output_files
            FROM task_runs
            WHERE output_files IS NOT NULL AND output_files != '[]'
        """).fetchall()

        for row in rows:
            raw_files = row["output_files"]
            try:
                files = json.loads(raw_files) if isinstance(raw_files, str) else raw_files
            except Exception:
                continue
            if not isinstance(files, list):
                continue
            for item in files:
                file_path = str(item or '').strip()
                if file_path and _normalize_output_file_path(file_path) == target_path:
                    return True

    return False


# ─── Export functions ───

def _sanitize_filename(text: Any, fallback: str = "output") -> str:
    value = str(text or "").strip()
    value = re.sub(r"\s+", " ", value)
    value = re.sub(r"[\x00-\x1f]+", "", value)
    value = re.sub(r'[\\/:*?"<>|]+', "_", value)
    value = value.strip(" .")
    return value or fallback


def _shorten_filename(filename: str, max_length: int = MAX_EXPORT_FILENAME_LENGTH) -> str:
    value = _sanitize_filename(filename)
    if len(value) <= max_length:
        return value

    suffix = Path(value).suffix
    stem = value[:-len(suffix)] if suffix else value
    budget = max_length - len(suffix)
    if budget <= 0:
        return value[:max_length].strip(" ._") or "output"

    tail_budget = min(64, max(24, budget // 2))
    head_budget = max(12, budget - tail_budget - 1)
    shortened = f"{stem[:head_budget].rstrip(' ._')}_{stem[-tail_budget:].lstrip(' ._')}{suffix}"
    return shortened[:max_length].strip(" ._") or f"output{suffix}"


def _render_filename(template: str, adapter_id: str, task_id: str,
                     filename_vars: Optional[Mapping[str, Any]] = None) -> str:
    now = datetime.now()
    vars_map = {
        "date": now.strftime("%Y%m%d"),
        "datetime": now.strftime("%Y%m%d_%H%M%S"),
        "timestamp": now.strftime("%Y%m%d-%H%M%S"),
        "adapter_id": adapter_id,
        "task_id": task_id,
    }
    if filename_vars:
        for key, value in filename_vars.items():
            vars_map[str(key)] = _sanitize_filename(value, "")
    return _shorten_filename(template.format_map(_SafeTemplateVars(vars_map)))


def _ensure_unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    idx = 2
    while True:
        candidate = path.with_name(f"{stem}_{idx}{suffix}")
        if not candidate.exists():
            return candidate
        idx += 1


def _normalize_sheet_title(name: Any, fallback: str) -> str:
    value = _sanitize_filename(name, fallback).strip()
    if not value:
        value = fallback
    return value[:31] or fallback[:31] or "Sheet1"


def _clean_excel_row(row: Mapping[str, Any]) -> dict[str, Any]:
    cleaned: dict[str, Any] = {}
    for key, value in (row or {}).items():
        key_text = str(key or "")
        if key_text.startswith("__"):
            continue
        cleaned[key_text] = value
    return cleaned


def _normalize_column_groups(column_groups: Optional[Iterable[Mapping[str, Any]]]) -> list[dict[str, Any]]:
    normalized_groups = []
    for group in column_groups or []:
        if hasattr(group, "model_dump"):
            group = group.model_dump()
        label = str((group or {}).get("label") or "").strip()
        columns = [
            str(col or "").strip()
            for col in ((group or {}).get("columns") or [])
            if str(col or "").strip()
        ]
        if not label or not columns:
            continue
        normalized_groups.append({"label": label, "columns": columns})
    return normalized_groups


def _resolve_headers(data: List[dict], column_order: Optional[List[str]] = None) -> list[str]:
    headers: list[str] = []
    seen_headers = set()
    for key in column_order or []:
        key_text = str(key or "").strip()
        if not key_text or key_text.startswith("__") or key_text in seen_headers:
            continue
        seen_headers.add(key_text)
        headers.append(key_text)
    for row in data:
        for key in row.keys():
            key_text = str(key or "").strip()
            if not key_text or key_text.startswith("__") or key_text in seen_headers:
                continue
            seen_headers.add(key_text)
            headers.append(key_text)
    return headers


def _write_excel_sheet(ws, data: List[dict], column_order: Optional[List[str]] = None,
                       column_groups: Optional[Iterable[Mapping[str, Any]]] = None) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    headers = _resolve_headers(data, column_order)
    if not headers:
        headers = ["提示"]
        data = [{"提示": "无数据"}]

    header_fill = PatternFill(fill_type="solid", fgColor="F5F7FA")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    def _leaf_header_label(column_key: str) -> str:
        value = str(column_key or "").strip()
        if "/" in value:
            return value.split("/")[-1].strip() or value
        return value

    normalized_groups = _normalize_column_groups(column_groups)
    group_lookup = {}
    for group in normalized_groups:
        for column in group["columns"]:
            group_lookup[column] = group["label"]

    has_grouped_header = bool(normalized_groups)
    if has_grouped_header:
        top_row = []
        leaf_row = []
        for header in headers:
            group_label = group_lookup.get(header, "")
            if group_label:
                top_row.append(group_label)
                leaf_row.append(_leaf_header_label(header))
            else:
                top_row.append(_leaf_header_label(header))
                leaf_row.append("")

        ws.append(top_row)
        ws.append(leaf_row)

        col_index = 1
        while col_index <= len(headers):
            header = headers[col_index - 1]
            group_label = group_lookup.get(header, "")
            if not group_label:
                ws.merge_cells(start_row=1, start_column=col_index, end_row=2, end_column=col_index)
                col_index += 1
                continue

            end_col = col_index
            while end_col < len(headers) and group_lookup.get(headers[end_col], "") == group_label:
                end_col += 1
            if end_col > col_index:
                ws.merge_cells(start_row=1, start_column=col_index, end_row=1, end_column=end_col)
            col_index = end_col + 1

        for row_no in (1, 2):
            for cell in ws[row_no]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
        ws.freeze_panes = "A3"
    else:
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        ws.freeze_panes = "A2"

    for row in data:
        ws.append([str(row.get(h, "")) for h in headers])

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(value or "")))
        ws.column_dimensions[letter].width = min(max_len + 4, 60)


def prepare_artifact_dir(adapter_id: str, task_id: str, run_id: int, kind: str = "artifacts") -> str:
    """Create and return a per-run artifact directory."""
    out_dir = artifact_dir_path(adapter_id, task_id, run_id, kind)
    out_dir.mkdir(parents=True, exist_ok=True)
    return str(out_dir)


def artifact_dir_path(adapter_id: str, task_id: str, run_id: int, kind: str = "artifacts") -> Path:
    """Return the per-run artifact directory path without creating it."""
    safe_kind = _sanitize_filename(kind, "artifacts")
    return _data_root() / adapter_id / task_id / safe_kind / str(run_id)


def export_excel(
    data: List[dict],
    adapter_id: str,
    task_id: str,
    filename_template: str = "{task_id}_{date}.xlsx",
    filename_vars: Optional[Mapping[str, Any]] = None,
    column_order: Optional[List[str]] = None,
    column_groups: Optional[List[Mapping[str, Any]]] = None,
    sheet_key: Optional[str] = None,
    sheet_configs: Optional[List[Mapping[str, Any]]] = None,
) -> str:
    """
    Export data to Excel file.
    Returns absolute path of written file.
    """
    from openpyxl import Workbook

    if not data:
        raise ValueError("No data to export")

    filename = _render_filename(filename_template, adapter_id, task_id, filename_vars)
    out_dir = _data_root() / adapter_id / task_id
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = _ensure_unique_path(out_dir / filename)

    wb = Workbook()
    cleaned_data = [_clean_excel_row(row) for row in data]
    normalized_sheet_key = str(sheet_key or "").strip()

    if not normalized_sheet_key:
        ws = wb.active
        ws.title = _normalize_sheet_title(task_id, task_id[:31] or "Sheet1")
        _write_excel_sheet(ws, cleaned_data, column_order, column_groups)
        total_rows = len(cleaned_data)
    else:
        sheet_rows: dict[str, list[dict[str, Any]]] = {}
        sheet_order: list[str] = []
        for raw_row in data:
            sheet_name = str((raw_row or {}).get(normalized_sheet_key) or "").strip() or "Sheet1"
            if sheet_name not in sheet_rows:
                sheet_rows[sheet_name] = []
                sheet_order.append(sheet_name)
            sheet_rows[sheet_name].append(_clean_excel_row(raw_row))

        config_by_name: dict[str, Mapping[str, Any]] = {}
        ordered_names: list[str] = []
        for config in sheet_configs or []:
            if hasattr(config, "model_dump"):
                config = config.model_dump()
            name = str((config or {}).get("name") or "").strip()
            if not name:
                continue
            config_by_name[name] = config
            ordered_names.append(name)

        final_sheet_names = [name for name in ordered_names if name in sheet_rows]
        final_sheet_names.extend(name for name in sheet_order if name not in final_sheet_names)
        if not final_sheet_names:
            final_sheet_names = ["Sheet1"]
            sheet_rows["Sheet1"] = cleaned_data

        ws = wb.active
        first_name = final_sheet_names[0]
        ws.title = _normalize_sheet_title(first_name, "Sheet1")
        first_config = config_by_name.get(first_name) or {}
        _write_excel_sheet(
            ws,
            sheet_rows.get(first_name) or [],
            list((first_config or {}).get("columns") or column_order or []),
            (first_config or {}).get("column_groups") or column_groups,
        )

        for name in final_sheet_names[1:]:
            config = config_by_name.get(name) or {}
            sheet = wb.create_sheet(title=_normalize_sheet_title(name, "Sheet"))
            _write_excel_sheet(
                sheet,
                sheet_rows.get(name) or [],
                list((config or {}).get("columns") or column_order or []),
                (config or {}).get("column_groups") or column_groups,
            )

        total_rows = sum(len(rows) for rows in sheet_rows.values())

    wb.save(str(out_path))
    logger.info(f"Excel exported: {out_path} ({total_rows} rows)")
    return str(out_path)


def export_json(data: List[dict], adapter_id: str, task_id: str,
                filename_template: str = "{task_id}_{date}.json",
                filename_vars: Optional[Mapping[str, Any]] = None) -> str:
    """
    Export data to JSON file.
    Returns absolute path of written file.
    """
    filename = _render_filename(filename_template, adapter_id, task_id, filename_vars)
    out_dir = _data_root() / adapter_id / task_id
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = _ensure_unique_path(out_dir / filename)

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    logger.info(f"JSON exported: {out_path} ({len(data)} records)")
    return str(out_path)


def export_desktop_excel(data: List[dict], adapter_id: str, task_id: str,
                         filename_template: str = "{task_id}_{date}.xlsx",
                         filename_vars: Optional[Mapping[str, Any]] = None) -> str:
    """Export to user Desktop (mirrors temu-assistant behavior)"""
    from openpyxl import Workbook
    desktop = Path.home() / "Desktop"
    filename = _render_filename(filename_template, adapter_id, task_id, filename_vars)
    out_path = _ensure_unique_path(desktop / filename)
    wb = Workbook()
    ws = wb.active
    ws.title = task_id[:31]
    if data:
        headers = list(data[0].keys())
        ws.append(headers)
        for row in data:
            ws.append([str(row.get(h, "")) for h in headers])
    wb.save(str(out_path))
    logger.info(f"Excel exported to Desktop: {out_path}")
    return str(out_path)
