#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
import tempfile
import time
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from core import shenhui_shoe_packaging as shoe  # noqa: E402
from core import shenhui_shoe_rules as shoe_rules  # noqa: E402


BASE_COLUMNS = [
    "输入款号",
    "颜色",
    "原文件名",
    "云盘路径",
    "规则槽位",
    "输出文件名",
    "处理动作",
    "下载结果",
    "本地文件",
    "压缩结果",
    "规则告警",
    "品类来源",
    "备注",
]


def text(value: Any) -> str:
    return str(value or "").strip()


def summarize_logs(logs: list[Any]) -> dict[str, Any]:
    lines = [text(item) for item in logs]
    correction_lines = [line for line in lines if "鞋品确定性校验：" in line]
    return {
        "correction_count": len(correction_lines),
        "pose_model_attempts": sum(
            "鞋品姿势识别模型" in line and "尝试：" in line
            for line in lines
        ),
        "ocr_model_attempts": sum("鞋盒标签 OCR 模型尝试" in line for line in lines),
        "soft_timeout_count": sum("60 秒软超时" in line for line in lines),
        "timeout_probe_count": sum("单批耐心复测" in line or "单次耐心复测" in line for line in lines),
        "fallback_count": sum(
            "快速 fallback" in line or "优先切换独立 fallback" in line
            for line in lines
        ),
        "strategy_lines": [line for line in lines if "鞋品姿势识别策略：" in line],
    }


def rows_from_xlsx(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        rows.append({
            headers[index]: (values[index] if index < len(values) else "")
            for index in range(len(headers))
        })
    wb.close()
    return rows


def prepared_rows(source_root: Path, report_xlsx: Path, style: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen_paths: set[Path] = set()
    for row in rows_from_xlsx(report_xlsx):
        slot = text(row.get("规则槽位"))
        if text(row.get("输入款号")) != style:
            continue
        if slot != "原始素材" and not re.fullmatch(r"yk\d+", slot, flags=re.IGNORECASE):
            continue
        output_name = text(row.get("输出文件名"))
        if not output_name:
            continue
        reported_local_path = Path(text(row.get("本地文件"))).expanduser()
        local_path = (
            reported_local_path
            if reported_local_path.is_file()
            else source_root / style / output_name
        )
        if local_path in seen_paths:
            continue
        seen_paths.add(local_path)
        current = {key: ("" if value is None else value) for key, value in row.items()}
        current["本地文件"] = str(local_path)
        current["下载结果"] = "已下载"
        current["__shenhui_group_code"] = style
        current["__shoe_original_filename"] = text(row.get("原文件名")) or local_path.name
        color_folder = output_name.split("/", 1)[0]
        color_match = (
            re.search(r"(?<!\d)(\d{5})(?!\d)", text(row.get("颜色")))
            or re.search(r"[-_](\d{5})(?:\D|$)", color_folder)
            or re.search(r"(?<!\d)(\d{5})(?!\d)", color_folder)
        )
        current["__shoe_color_code"] = color_match.group(1) if color_match else ""
        rows.append(current)
    return rows


def write_report_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    extra_columns: list[str] = []
    for row in rows:
        for key in row:
            key_text = text(key)
            if key_text and key_text not in BASE_COLUMNS and not key_text.startswith("__"):
                extra_columns.append(key_text)
    columns = [*BASE_COLUMNS, *dict.fromkeys(extra_columns)]
    wb = Workbook()
    ws = wb.active
    ws.title = "结果"
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column, "") for column in columns])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def _copy_file_preserving_relative(source: Path, source_root: Path, target_root: Path) -> Path:
    relative = source.relative_to(source_root)
    target = target_root / relative
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, target)
    return target


def _is_contact_sheet_artifact(path: Path) -> bool:
    if path.suffix.lower() not in {".jpg", ".jpeg", ".png"}:
        return False
    return bool(
        re.fullmatch(
            r"\d{5}(?:-\d+|-all|-global-\d+|-overview)",
            path.stem,
            flags=re.IGNORECASE,
        )
    )


def collect_contact_sheet_artifacts(attempt_root: Path) -> list[Path]:
    analysis_root = attempt_root / "_shoe_analysis"
    if not analysis_root.is_dir():
        return []
    return sorted(
        path
        for path in analysis_root.rglob("*")
        if path.is_file() and _is_contact_sheet_artifact(path)
    )


def collect_analysis_artifacts(attempt_root: Path) -> list[Path]:
    analysis_root = attempt_root / "_shoe_analysis"
    if not analysis_root.is_dir():
        return []
    return sorted(
        path
        for path in analysis_root.rglob("*")
        if path.is_file() and path.suffix.lower() in {".jpg", ".jpeg", ".png", ".json"}
    )


def create_visual_review_sheet(
    style_root: Path,
    report_rows: list[dict[str, Any]],
    target: Path,
) -> Path | None:
    entries: list[tuple[Path, str]] = []
    seen: set[Path] = set()
    for row in report_rows:
        slot = text(row.get("规则槽位"))
        if slot == "原始素材" or text(row.get("下载结果")) != "已下载":
            continue
        output_path = row_output_path(style_root, row)
        if (
            not output_path.is_file()
            or output_path in seen
            or output_path.suffix.lower() not in {".jpg", ".jpeg", ".png"}
        ):
            continue
        seen.add(output_path)
        label = f"{slot} {output_path.relative_to(style_root)}"
        entries.append((output_path, label))
    if not entries:
        return None

    from PIL import Image, ImageDraw, ImageFont, ImageOps

    columns = 5
    tile_width = 220
    image_height = 170
    label_height = 36
    rows = (len(entries) + columns - 1) // columns
    sheet = Image.new(
        "RGB",
        (tile_width * columns, (image_height + label_height) * rows),
        "white",
    )
    draw = ImageDraw.Draw(sheet)
    font = ImageFont.load_default()
    for index, (source, label) in enumerate(entries):
        left = (index % columns) * tile_width
        top = (index // columns) * (image_height + label_height)
        with Image.open(source) as opened:
            image = shoe._image_rgb_on_white(ImageOps.exif_transpose(opened))
            image.thumbnail((tile_width - 12, image_height - 12), Image.Resampling.LANCZOS)
            x = left + (tile_width - image.width) // 2
            y = top + (image_height - image.height) // 2
            sheet.paste(image, (x, y))
        draw.rectangle(
            (left, top + image_height, left + tile_width, top + image_height + label_height),
            fill=(245, 245, 245),
        )
        draw.text(
            (left + 5, top + image_height + 6),
            label[:42],
            fill="black",
            font=font,
        )
    target.parent.mkdir(parents=True, exist_ok=True)
    sheet.save(target, format="JPEG", quality=82, optimize=True)
    return target


def copy_final_artifacts(
    *,
    output_root: Path,
    attempt_root: Path,
    style: str,
    style_root: Path,
    contact_sheets: list[Path],
    analysis_artifacts: list[Path] | None = None,
) -> dict[str, Any]:
    final_root = output_root / "final"
    final_style_root = final_root / style
    evidence_root = final_root / "evidence" / style
    if final_style_root.exists():
        shutil.rmtree(final_style_root)
    if evidence_root.exists():
        shutil.rmtree(evidence_root)
    final_root.mkdir(parents=True, exist_ok=True)
    shutil.copytree(style_root, final_style_root)

    copied: dict[str, Any] = {
        "final_style_root": str(final_style_root),
        "final_evidence_root": str(evidence_root),
    }
    evidence_root.mkdir(parents=True, exist_ok=True)
    for name in ("report.xlsx", "report_rows.json", "logs.txt", "visual-review.jpg"):
        source = attempt_root / name
        if source.is_file():
            target = evidence_root / name
            shutil.copy2(source, target)
            copied[name.replace("-", "_").replace(".", "_")] = str(target)
    if contact_sheets:
        contact_root = evidence_root / "contact-sheets"
        copied_sheets = [
            str(_copy_file_preserving_relative(path, attempt_root, contact_root))
            for path in contact_sheets
            if path.is_file() and attempt_root in path.parents
        ]
        copied["final_contact_sheets"] = copied_sheets
    analysis_artifacts = analysis_artifacts or []
    json_artifacts = [
        path
        for path in analysis_artifacts
        if path.is_file()
        and path.suffix.lower() == ".json"
        and attempt_root in path.parents
    ]
    if json_artifacts:
        analysis_root = evidence_root / "analysis-artifacts"
        copied_json = [
            str(_copy_file_preserving_relative(path, attempt_root, analysis_root))
            for path in json_artifacts
        ]
        copied["final_analysis_artifacts"] = copied_json
    shutil.copy2(attempt_root / "report.xlsx", final_root / f"{style}-report.xlsx")
    return copied


def _existing_attempt_path(
    output_root: Path,
    style: str,
    attempt: str,
) -> Path:
    return output_root / "_attempts" / style / f"attempt-{attempt}"


def _path_from_validation(
    value: Any,
    *,
    attempt_root: Path,
) -> Path:
    path = Path(text(value))
    if not path.is_absolute():
        path = attempt_root / path
    return path


def _write_validation(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _load_attempt_report_rows(attempt_root: Path, report_xlsx: Path) -> list[dict[str, Any]]:
    report_json = attempt_root / "report_rows.json"
    if report_json.is_file():
        payload = json.loads(report_json.read_text(encoding="utf-8"))
        if not isinstance(payload, list):
            raise ValueError("report_rows.json is not a list")
        return [row for row in payload if isinstance(row, dict)]
    return rows_from_xlsx(report_xlsx)


def finalize_existing_attempt(
    *,
    output_root: Path,
    attempt_root: Path,
    style: str,
    visual_review_note: str = "",
) -> tuple[dict[str, Any], int]:
    validation_path = attempt_root / "validation.json"
    if not validation_path.is_file():
        return (
            {
                "style": style,
                "attempt_root": str(attempt_root),
                "issues": ["existing attempt validation.json missing"],
            },
            2,
        )
    try:
        summary = json.loads(validation_path.read_text(encoding="utf-8"))
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        return (
            {
                "style": style,
                "attempt_root": str(attempt_root),
                "issues": [f"existing attempt validation.json cannot be read: {exc}"],
            },
            2,
        )
    if not isinstance(summary, dict):
        return (
            {
                "style": style,
                "attempt_root": str(attempt_root),
                "issues": ["existing attempt validation.json is not an object"],
            },
            2,
        )

    issues: list[str] = []
    if summary.get("issues") != []:
        issues.append("existing attempt validation issues are not empty")

    style_root_value = text(summary.get("style_root"))
    if not style_root_value:
        issues.append("existing attempt style_root missing from validation")
        style_root = attempt_root / style
    else:
        style_root = _path_from_validation(style_root_value, attempt_root=attempt_root)
        if not style_root.is_dir():
            issues.append(f"existing attempt style_root missing: {style_root}")

    report_xlsx_value = text(summary.get("report_xlsx")) or str(attempt_root / "report.xlsx")
    report_xlsx = _path_from_validation(report_xlsx_value, attempt_root=attempt_root)
    if not report_xlsx.is_file():
        issues.append(f"existing attempt report.xlsx missing: {report_xlsx}")
    if not (attempt_root / "report.xlsx").is_file():
        issues.append(f"existing attempt canonical report.xlsx missing: {attempt_root / 'report.xlsx'}")

    visual_review_value = text(summary.get("visual_review")) or str(attempt_root / "visual-review.jpg")
    visual_review = _path_from_validation(visual_review_value, attempt_root=attempt_root)
    if not visual_review.is_file():
        issues.append(f"existing attempt visual-review missing: {visual_review}")

    contact_sheets = collect_contact_sheet_artifacts(attempt_root)
    if not contact_sheets:
        issues.append("existing attempt contact_sheets missing or empty")

    analysis_artifacts = collect_analysis_artifacts(attempt_root)
    report_rows: list[dict[str, Any]] = []
    report_rows_loaded = False
    if report_xlsx.is_file():
        try:
            report_rows = _load_attempt_report_rows(attempt_root, report_xlsx)
            report_rows_loaded = True
        except Exception as exc:
            issues.append(f"existing attempt report rows cannot be read: {exc}")
    if style_root.is_dir() and report_rows_loaded:
        validation_issues, validation_warnings = validate_style(
            style=style,
            style_root=style_root,
            report_rows=report_rows,
            category=text(summary.get("category")),
        )
        issues.extend(validation_issues)
        if validation_warnings:
            summary["finalize_warnings"] = validation_warnings

    if issues:
        failed = dict(summary)
        failed["finalize_existing"] = True
        failed["finalize_issues"] = issues
        failed["issues"] = issues
        _write_validation(validation_path, failed)
        print(json.dumps(failed, ensure_ascii=False, indent=2))
        return failed, 2

    summary["finalize_existing"] = True
    summary["visual_review_status"] = "approved"
    summary["visual_review_note"] = visual_review_note
    summary.update(
        copy_final_artifacts(
            output_root=output_root,
            attempt_root=attempt_root,
            style=style,
            style_root=style_root,
            contact_sheets=contact_sheets,
            analysis_artifacts=analysis_artifacts,
        )
    )
    _write_validation(validation_path, summary)
    final_evidence_root = text(summary.get("final_evidence_root"))
    if final_evidence_root:
        shutil.copy2(validation_path, Path(final_evidence_root) / "validation.json")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return summary, 0


def source_exists_without_copy(report_rows: list[dict[str, Any]], color: str, source_name: str) -> bool:
    if "拷贝" not in source_name:
        return False
    source_key = shoe._copy_variant_key(source_name)
    for row in report_rows:
        if text(row.get("颜色")) != color or text(row.get("规则槽位")) != "原始素材":
            continue
        original_name = text(row.get("原文件名"))
        if "拷贝" not in original_name and shoe._copy_variant_key(original_name) == source_key:
            return True
    return False


def image_size_mode(path: Path) -> tuple[tuple[int, int], str]:
    from PIL import Image

    with Image.open(path) as image:
        return image.size, image.mode


def image_supports_transparency(path: Path) -> bool:
    from PIL import Image

    with Image.open(path) as image:
        return image.mode in {"RGBA", "LA"} or "transparency" in image.info


def image_has_transparent_pixels(path: Path) -> bool:
    from PIL import Image

    with Image.open(path) as image:
        if image.mode not in {"RGBA", "LA"} and "transparency" not in image.info:
            return False
        alpha = image.convert("RGBA").getchannel("A")
        minimum, _maximum = alpha.getextrema()
        return minimum < 255


def validate_transparent_padding(
    path: Path,
    source_size: tuple[int, int],
    *,
    canvas_size: int = shoe.SHOE_CHANNEL_CANVAS_SIZE,
) -> list[str]:
    """Require transparent canvas bands whenever thumbnail fitting leaves space."""

    from PIL import Image

    source_width, source_height = source_size
    if source_width <= 0 or source_height <= 0:
        return [f"cannot verify transparent padding for invalid source size: {source_size}"]
    scale = min(
        1.0,
        canvas_size / source_width,
        canvas_size / source_height,
    )
    fitted_width = max(1, round(source_width * scale))
    fitted_height = max(1, round(source_height * scale))
    needs_horizontal_padding = fitted_width < canvas_size
    needs_vertical_padding = fitted_height < canvas_size
    if not needs_horizontal_padding and not needs_vertical_padding:
        return []

    with Image.open(path) as opened:
        alpha = opened.convert("RGBA").getchannel("A")
        width, height = alpha.size
        if (width, height) != (canvas_size, canvas_size):
            return [
                f"cannot verify transparent padding on non-{canvas_size}x{canvas_size} image: "
                f"{path} size={(width, height)}"
            ]
        if needs_horizontal_padding:
            left_opaque = alpha.crop((0, 0, 1, height)).getextrema()[1]
            right_opaque = alpha.crop((width - 1, 0, width, height)).getextrema()[1]
            if left_opaque != 0 or right_opaque != 0:
                return [
                    f"missing transparent padding: {path} source={source_size} "
                    f"fitted={(fitted_width, fitted_height)}"
                ]
        if needs_vertical_padding:
            top_opaque = alpha.crop((0, 0, width, 1)).getextrema()[1]
            bottom_opaque = alpha.crop((0, height - 1, width, height)).getextrema()[1]
            if top_opaque != 0 or bottom_opaque != 0:
                return [
                    f"missing transparent padding: {path} source={source_size} "
                    f"fitted={(fitted_width, fitted_height)}"
                ]
    return []


def image_background_luma(path: Path) -> float:
    try:
        return shoe._binary_pose_feature(path).background_luma
    except Exception:
        return 0.0


def red_pixel_count(path: Path) -> int:
    from PIL import Image

    with Image.open(path) as image:
        rgb = image.convert("RGB")
        return sum(
            1
            for red, green, blue in rgb.getdata()
            if red > 220 and green < 65 and blue < 65
        )


def _red_components(path: Path) -> tuple[tuple[int, int], list[set[tuple[int, int]]]]:
    from PIL import Image

    with Image.open(path) as image:
        rgb = image.convert("RGB")
        width, height = rgb.size
        red_pixels = {
            (x, y)
            for y in range(height)
            for x in range(width)
            if (
                (pixel := rgb.getpixel((x, y)))[0] > 180
                and pixel[0] > pixel[1] * 1.8
                and pixel[0] > pixel[2] * 1.8
                and pixel[1] < 130
                and pixel[2] < 130
            )
        }
    components: list[set[tuple[int, int]]] = []
    while red_pixels:
        start = red_pixels.pop()
        component = {start}
        stack = [start]
        while stack:
            x, y = stack.pop()
            for dx in (-1, 0, 1):
                for dy in (-1, 0, 1):
                    if not dx and not dy:
                        continue
                    neighbor = (x + dx, y + dy)
                    if neighbor in red_pixels:
                        red_pixels.remove(neighbor)
                        component.add(neighbor)
                        stack.append(neighbor)
        if len(component) >= 20:
            components.append(component)
    return (width, height), components


def _red_rectangle_bounds(path: Path) -> list[tuple[int, int, int, int]]:
    (_width, _height), components = _red_components(path)
    rectangles: list[tuple[float, tuple[int, int, int, int]]] = []
    for component in components:
        xs = [point[0] for point in component]
        ys = [point[1] for point in component]
        bounds = (min(xs), min(ys), max(xs), max(ys))
        left, top, right, bottom = bounds
        box_width = right - left + 1
        box_height = bottom - top + 1
        if box_width < 12 or box_height < 8:
            continue
        edge_band = max(2, min(5, round(min(box_width, box_height) * 0.08)))
        edge_pixels = sum(
            x - left < edge_band
            or right - x < edge_band
            or y - top < edge_band
            or bottom - y < edge_band
            for x, y in component
        )
        perimeter_band = max(1, 2 * (box_width + box_height) * edge_band)
        edge_score = edge_pixels / perimeter_band
        if edge_score < 0.25:
            continue
        rectangles.append((edge_score, bounds))
    return [bounds for _score, bounds in sorted(rectangles, reverse=True)]


def red_box_contains_bbox(
    path: Path,
    bbox: tuple[float, float, float, float],
) -> bool:
    from PIL import Image

    with Image.open(path) as image:
        width, height = image.size
    text_bounds = (
        bbox[0] * width,
        bbox[1] * height,
        bbox[2] * width,
        bbox[3] * height,
    )
    tolerance = max(2.0, min(width, height) * 0.01)
    return any(
        left - tolerance <= text_bounds[0]
        and top - tolerance <= text_bounds[1]
        and right + tolerance >= text_bounds[2]
        and bottom + tolerance >= text_bounds[3]
        for left, top, right, bottom in _red_rectangle_bounds(path)
    )


def validate_tmq_style_code(path: Path, style_code: str) -> list[str]:
    from PIL import Image

    try:
        ocr = shoe.ocr_service.recognize_image_with_tesseract_js(
            path,
            lang="eng",
            whitelist="0123456789",
        )
    except Exception as exc:
        return [f"tmq OCR failed for {style_code}: {exc}"]
    words = [
        word
        for word in (ocr.get("words") or [])
        if isinstance(word, shoe.ocr_service.OcrWord)
    ]
    with Image.open(path) as opened:
        image = opened.convert("RGB")
    bbox = shoe.ocr_service.refine_style_code_bbox(
        image=image,
        label_bbox=None,
        style_code_bbox=None,
        style_code=style_code,
        ocr_words=words,
    )
    if bbox is not None and red_box_contains_bbox(path, bbox):
        return []

    # Full-image Tesseract occasionally returns the correct 12 digits but a
    # bounding box that also spans the adjacent balabala logo. Verify the text
    # inside the detected red rectangle directly before reporting a failure.
    rectangles = _red_rectangle_bounds(path)
    with tempfile.TemporaryDirectory(prefix="crawshrimp-tmq-red-box-") as tmpdir:
        for index, (left, top, right, bottom) in enumerate(rectangles):
            inset = max(1, round(min(right - left, bottom - top) * 0.02))
            cropped = image.crop((
                left + inset,
                top + inset,
                max(left + inset + 1, right - inset + 1),
                max(top + inset + 1, bottom - inset + 1),
            ))
            if cropped.height < 160:
                scale = 160 / max(1, cropped.height)
                cropped = cropped.resize(
                    (
                        max(1, round(cropped.width * scale)),
                        max(1, round(cropped.height * scale)),
                    ),
                    Image.Resampling.LANCZOS,
                )
            crop_path = Path(tmpdir) / f"red-box-{index}.png"
            cropped.save(crop_path, format="PNG")
            try:
                crop_ocr = shoe.ocr_service.recognize_image_with_tesseract_js(
                    crop_path,
                    lang="eng",
                    whitelist="0123456789",
                )
            except Exception:
                continue
            if style_code in re.sub(r"\D", "", text(crop_ocr.get("text"))):
                return []
    if bbox is None:
        return [f"tmq does not OCR the full style code {style_code}"]
    return [f"tmq red box does not contain style code {style_code}"]


def validate_color_folder_names(
    color_dirs: list[Path],
    report_rows: list[dict[str, Any]],
) -> list[str]:
    issues: list[str] = []
    verified_by_code: dict[str, str] = {}
    for row in report_rows:
        if text(row.get("鞋盒命名已验证")) != "是":
            continue
        verified_name = text(row.get("鞋盒款色名"))
        color = text(row.get("颜色"))
        match = re.search(r"(\d{5})", verified_name) or re.search(r"(\d{5})", color)
        if match and verified_name:
            verified_by_code[match.group(1)] = verified_name
    for color_dir in color_dirs:
        actual = re.sub(r"^\d+\.", "", color_dir.name)
        match = re.search(r"(\d{5})", actual)
        if not match:
            issues.append(f"{color_dir.name} missing verified 5-digit color code")
            continue
        expected = verified_by_code.get(match.group(1))
        if not expected:
            issues.append(
                f"{color_dir.name} has no verified shoe-box color name evidence"
            )
            continue
        expected_safe = shoe._safe_path_component(expected)
        if actual != expected_safe:
            issues.append(
                f"color folder {color_dir.name} does not match verified "
                f"shoe-box name {expected_safe}"
            )
    return issues


def validate_channel_file_names(
    style_root: Path,
    color_dirs: list[Path],
    report_rows: list[dict[str, Any]],
) -> list[str]:
    verified_by_code: dict[str, str] = {}
    for row in report_rows:
        if text(row.get("鞋盒命名已验证")) != "是":
            continue
        verified_name = text(row.get("鞋盒款色名"))
        color = text(row.get("颜色"))
        match = re.search(r"(\d{5})", verified_name) or re.search(r"(\d{5})", color)
        if match and verified_name:
            verified_by_code[match.group(1)] = shoe._safe_path_component(verified_name)

    expected_color_names: set[str] = set()
    for color_dir in color_dirs:
        match = re.search(r"(\d{5})", color_dir.name)
        if not match:
            continue
        verified_name = verified_by_code.get(match.group(1))
        if verified_name:
            expected_color_names.add(verified_name)

    issues: list[str] = []
    for prefix in ("jdt", "wpt30"):
        expected = {
            f"{prefix}.{color_name}.png"
            for color_name in expected_color_names
        }
        actual = {
            path.name
            for path in style_root.glob(f"{prefix}.*.png")
            if path.is_file()
        }
        issues.extend(
            f"missing verified channel file {name}"
            for name in sorted(expected - actual)
        )
        issues.extend(
            f"unexpected or misnamed channel file {name}"
            for name in sorted(actual - expected)
        )
    return issues


def validate_tms_tmz5_source_families(
    report_rows: list[dict[str, Any]],
) -> list[str]:
    tmz5_by_color: dict[str, set[str]] = {}
    for row in report_rows:
        if (
            text(row.get("规则槽位")) != "tmz5"
            or text(row.get("下载结果")) != "已下载"
        ):
            continue
        color = text(row.get("颜色"))
        family = shoe._copy_variant_key(text(row.get("原文件名")))
        if color and family:
            tmz5_by_color.setdefault(color, set()).add(family)

    issues: list[str] = []
    for row in report_rows:
        if (
            text(row.get("规则槽位")) != "tms"
            or text(row.get("下载结果")) != "已下载"
        ):
            continue
        color = text(row.get("颜色"))
        expected_families = tmz5_by_color.get(color)
        if not expected_families:
            continue
        actual_family = shoe._copy_variant_key(text(row.get("原文件名")))
        if actual_family not in expected_families:
            issues.append(
                f"{color} tms source does not match tmz5: "
                f"actual={actual_family or 'empty'} "
                f"expected={','.join(sorted(expected_families))}"
            )
    return issues


def validate_wpz5_box_distinct(color_dir: Path) -> list[str]:
    wpz5 = color_dir / "wpz (15).jpg"
    wpz6 = color_dir / "wpz (16).jpg"
    if not wpz5.is_file() or not wpz6.is_file():
        return []
    if wpz5.read_bytes() == wpz6.read_bytes():
        return [f"{color_dir.name} wpz (15).jpg and wpz (16).jpg are identical"]
    first = shoe._binary_pose_feature(wpz5)
    second = shoe._binary_pose_feature(wpz6)
    distance = shoe._binary_pose_distance(first, second)
    if distance <= shoe.SHOE_MAIN_SLOT_DUPLICATE_MAX_DISTANCE:
        return [
            f"{color_dir.name} wpz (15).jpg and wpz (16).jpg are near duplicates "
            f"distance={distance:.4f}"
        ]
    return []


def row_output_path(style_root: Path, row: dict[str, Any]) -> Path:
    output_name = text(row.get("输出文件名"))
    return style_root / output_name if output_name else Path()


def _json_object(value: Any) -> dict[str, Any] | None:
    if isinstance(value, dict):
        return value
    raw = text(value)
    if not raw:
        return None
    try:
        parsed = json.loads(raw)
    except (TypeError, ValueError, json.JSONDecodeError):
        return None
    return parsed if isinstance(parsed, dict) else None


def _candidate_fact(value: Any) -> shoe_rules.CandidateFacts | None:
    if not isinstance(value, dict):
        return None
    matched_slots = value.get("matched_slots")
    if not isinstance(matched_slots, list):
        matched_slots = []
    try:
        return shoe_rules.CandidateFacts(
            candidate_id=text(value.get("candidate_id")),
            filename=text(value.get("filename")),
            asset_type=text(value.get("asset_type")).lower(),
            shoe_count=text(value.get("shoe_count")).lower(),
            pose=text(value.get("pose")).lower(),
            background=text(value.get("background")).lower(),
            complete=bool(value.get("complete")),
            side=text(value.get("side")).lower(),
            outsole_visible=bool(value.get("outsole_visible")),
            feature_card=bool(value.get("feature_card")),
            confidence=float(value.get("confidence") or 0.0),
            matched_slots=tuple(text(item) for item in matched_slots if text(item)),
        )
    except (TypeError, ValueError):
        return None


def validate_semantic_rows(
    report_rows: list[dict[str, Any]],
    *,
    category: str,
    require_evidence: bool = False,
) -> list[str]:
    issues: list[str] = []
    semantic_slots = {
        *{f"tmz{index}" for index in range(1, 6)},
        *{f"wpz{index}" for index in range(1, 7)},
        *{f"yq{index}" for index in range(1, 4)},
        "o",
        "tms",
        "yx",
    }
    for row in report_rows:
        slot = text(row.get("规则槽位"))
        evidence = _json_object(row.get("语义属性"))
        consensus = _json_object(row.get("模型共识"))
        if evidence is None and consensus is None:
            if (
                require_evidence
                and slot in semantic_slots
                and text(row.get("下载结果")) == "已下载"
            ):
                issues.append(f"{slot} missing semantic model evidence")
            continue
        if consensus is None:
            issues.append(f"{slot} missing model consensus evidence")
            continue
        votes = int(consensus.get("votes") or 0)
        required_votes = max(1, int(consensus.get("required_votes") or 2))
        model_ids = {
            text(item)
            for item in (consensus.get("models") or [])
            if text(item)
        }
        if (
            text(consensus.get("status")) != "locked"
            or votes < required_votes
            or len(model_ids) < required_votes
        ):
            issues.append(
                f"{slot} semantic consensus insufficient {votes}/{required_votes}"
            )
            continue
        if evidence is None:
            issues.append(f"{slot} missing candidate semantic facts")
            continue
        evidence_slot = text(evidence.get("slot")) or slot
        expected_evidence_slot = shoe._semantic_vote_slot(slot, category)
        if evidence_slot != expected_evidence_slot:
            issues.append(
                f"{slot} semantic evidence slot mismatch: "
                f"{evidence_slot} != {expected_evidence_slot}"
            )
            continue
        source_family = shoe._copy_variant_key(text(row.get("原文件名")))
        selected_family = text(consensus.get("selected_family"))
        if not source_family or selected_family != source_family:
            issues.append(
                f"{slot} semantic consensus does not match actual source: "
                f"selected={selected_family or 'empty'} actual={source_family or 'empty'}"
            )
            continue
        facts_by_model: dict[str, list[shoe_rules.CandidateFacts]] = {}
        for item in evidence.get("models") or []:
            if not isinstance(item, dict):
                continue
            model_id = text(item.get("model_id"))
            fact = _candidate_fact(item.get("fact"))
            if not model_id or fact is None:
                continue
            if model_id not in model_ids:
                continue
            facts_by_model.setdefault(model_id, []).append(fact)

        valid_models: set[str] = set()
        for model_id in sorted(model_ids):
            facts = facts_by_model.get(model_id) or []
            source_facts = [
                fact
                for fact in facts
                if shoe._copy_variant_key(fact.filename) == source_family
            ]
            if not source_facts:
                issues.append(
                    f"{slot} model {model_id} facts do not match actual source: "
                    f"{facts[0].filename if facts else 'empty'} != "
                    f"{text(row.get('原文件名'))}"
                )
                continue
            reasons: list[str] = []
            for fact in source_facts:
                valid, reason = shoe_rules.candidate_is_valid_for_slot(
                    fact,
                    evidence_slot,
                    category,
                )
                if valid:
                    valid_models.add(model_id)
                    break
                if reason:
                    reasons.append(reason)
            if model_id not in valid_models:
                issues.append(
                    f"{slot} model {model_id} invalid semantics: "
                    f"{reasons[0] if reasons else 'no valid supporting fact'}"
                )
        if len(valid_models) < required_votes:
            issues.append(
                f"{slot} valid semantic facts insufficient "
                f"{len(valid_models)}/{required_votes}"
            )
    return issues


def validate_style(
    *,
    style: str,
    style_root: Path,
    report_rows: list[dict[str, Any]],
    category: str,
) -> tuple[list[str], list[str]]:
    issues: list[str] = []
    warnings: list[str] = []
    root_required = [f"tmz ({index}).jpg" for index in range(1, 6)]
    root_required.extend(["tmq.jpg", "tmt.png"])
    for name in root_required:
        if not (style_root / name).is_file():
            issues.append(f"missing root {name}")
    color_dirs = sorted(
        [
            path
            for path in style_root.iterdir()
            if path.is_dir() and re.match(r"^\d+\.", path.name)
        ],
        key=lambda path: path.name,
    ) if style_root.is_dir() else []
    if not color_dirs:
        issues.append("missing color directories")
    issues.extend(validate_color_folder_names(color_dirs, report_rows))
    issues.extend(validate_channel_file_names(style_root, color_dirs, report_rows))

    original_size_by_color_name: dict[tuple[str, str], tuple[int, int]] = {}
    for row in report_rows:
        if text(row.get("规则槽位")) != "原始素材":
            continue
        output_path = row_output_path(style_root, row)
        if output_path.is_file():
            try:
                size, _mode = image_size_mode(output_path)
                original_size_by_color_name[(text(row.get("颜色")), text(row.get("原文件名")))] = size
            except Exception:
                warnings.append(f"cannot inspect original material {output_path}")

    for row in report_rows:
        slot = text(row.get("规则槽位"))
        if slot not in {"jdt", "wpt30", "tmt", "tmq", "tms"}:
            continue
        output_path = row_output_path(style_root, row)
        if not output_path.is_file():
            continue
        try:
            size, mode = image_size_mode(output_path)
        except Exception as exc:
            issues.append(f"{slot} cannot be opened: {output_path} {exc}")
            continue
        if slot in {"jdt", "tmt"}:
            if size != (shoe.SHOE_CHANNEL_CANVAS_SIZE, shoe.SHOE_CHANNEL_CANVAS_SIZE):
                issues.append(f"{slot} is not 800x800: {output_path} size={size}")
            if mode != "RGBA":
                issues.append(f"{slot} is not transparent PNG/RGBA: {output_path} mode={mode}")
            source_size = original_size_by_color_name.get((
                text(row.get("颜色")),
                text(row.get("原文件名")),
            ))
            if source_size:
                issues.extend(validate_transparent_padding(output_path, source_size))
            else:
                issues.append(
                    f"{slot} cannot verify transparent padding without source dimensions: "
                    f"{output_path}"
                )
        elif slot == "wpt30":
            if output_path.stat().st_size >= shoe.SHOE_WPT_MAX_BYTES:
                issues.append(
                    f"wpt30 exceeds 600KB: {output_path} bytes={output_path.stat().st_size}"
                )
            if not image_supports_transparency(output_path):
                issues.append(f"wpt30 has no transparency channel: {output_path} mode={mode}")
            elif not image_has_transparent_pixels(output_path):
                issues.append(f"wpt30 has no transparent pixels: {output_path}")
            source_size = original_size_by_color_name.get((
                text(row.get("颜色")),
                text(row.get("原文件名")),
            ))
            if source_size and size != source_size:
                issues.append(
                    f"wpt30 does not keep original dimensions: {output_path} "
                    f"size={size} source={source_size}"
                )
            elif not source_size:
                issues.append(
                    f"wpt30 cannot verify original dimensions without source dimensions: "
                    f"{output_path}"
                )
        elif slot == "tmq":
            if size != (shoe.SHOE_TMQ_CANVAS_SIZE, shoe.SHOE_TMQ_CANVAS_SIZE):
                issues.append(f"tmq is not 800x800: {output_path} size={size}")
            if red_pixel_count(output_path) <= 0:
                issues.append(f"tmq has no visible red style-code box: {output_path}")
            issues.extend(validate_tmq_style_code(output_path, style))
        elif slot == "tms":
            if output_path.suffix.lower() not in {".jpg", ".jpeg"}:
                issues.append(f"tms is not jpg: {output_path}")
            luma = image_background_luma(output_path)
            if luma < shoe.SHOE_WHITE_BACKGROUND_LUMA:
                issues.append(f"tms is not white background: {output_path} luma={luma:.1f}")

    for color_dir in color_dirs:
        issues.extend(validate_wpz5_box_distinct(color_dir))
        if not re.search(r"\d{5}", color_dir.name):
            issues.append(f"{color_dir.name} missing 5-digit color code in folder name")
        required = [
            "wpz (1).jpg",
            "wpz (2).jpg",
            "wpz (3).jpg",
            "wpz (4).jpg",
            "wpz (15).jpg",
            "wpz (16).jpg",
            "tms.jpg",
            "yq (1).jpg",
            "yq (2).jpg",
            "yq (3).jpg",
        ]
        for name in required:
            if not (color_dir / name).is_file():
                issues.append(f"{color_dir.name} missing {name}")
        wpz15 = color_dir / "wpz (15).jpg"
        if wpz15.is_file():
            feature = shoe._binary_pose_feature(wpz15)
            if feature.background_luma >= shoe.SHOE_WHITE_BACKGROUND_LUMA:
                warnings.append(
                    f"{color_dir.name} wpz (15).jpg is not gray background; "
                    f"accepted as fallback when no qualified gray source exists "
                    f"luma={feature.background_luma:.1f}"
                )
    if color_dirs and not (color_dirs[0] / "o.jpg").is_file():
        issues.append(f"{color_dirs[0].name} missing o.jpg")

    tmz_features: list[tuple[int, shoe._BinaryPoseFeature]] = []
    for index in range(1, 6):
        path = style_root / f"tmz ({index}).jpg"
        if path.is_file():
            tmz_features.append((index, shoe._binary_pose_feature(path)))
    if len(tmz_features) == 5:
        category_text = text(category)
        pose3_max_aspect = 0.95 if category_text == "婴童" else 0.82
        pose3_max_coverage = 0.145 if category_text == "婴童" else 0.16
        pose3 = tmz_features[2][1]
        if not (
            0.45 <= pose3.aspect_ratio <= pose3_max_aspect
            and pose3.bounding_coverage <= pose3_max_coverage
        ):
            issues.append(
                "tmz (3).jpg invalid pose3 feature "
                f"aspect={pose3.aspect_ratio:.3f} coverage={pose3.bounding_coverage:.3f}"
            )
        pose5 = tmz_features[4][1]
        if pose5.background_luma < shoe.SHOE_WHITE_BACKGROUND_LUMA:
            issues.append(f"tmz (5).jpg is not white background luma={pose5.background_luma:.1f}")
        for first_index, first_feature in tmz_features:
            for second_index, second_feature in tmz_features:
                if first_index >= second_index:
                    continue
                distance = shoe._binary_pose_distance(first_feature, second_feature)
                if distance <= shoe.SHOE_MAIN_SLOT_DUPLICATE_MAX_DISTANCE:
                    issues.append(
                        f"tmz duplicate pose {first_index}/{second_index} "
                        f"distance={distance:.4f}"
                    )

    for row in report_rows:
        slot = text(row.get("规则槽位"))
        warning = text(row.get("规则告警"))
        action = text(row.get("处理动作"))
        if "已跳过 tmq" in warning or "已跳过 tmq" in action:
            issues.append("report contains skipped tmq")
        if slot in {
            "tmz1",
            "tmz2",
            "tmz3",
            "tmz4",
            "wpz1",
            "wpz2",
            "wpz3",
            "wpz4",
        }:
            color = text(row.get("颜色"))
            source_name = text(row.get("原文件名"))
            if source_exists_without_copy(report_rows, color, source_name):
                issues.append(f"{slot} uses copied white source {source_name}")
    issues.extend(
        validate_semantic_rows(
            report_rows,
            category=category,
            require_evidence=True,
        )
    )
    issues.extend(validate_tms_tmz5_source_families(report_rows))
    return issues, warnings


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("style")
    parser.add_argument("--attempt", required=True)
    parser.add_argument("--source-root", default="/Users/xingyicheng/Downloads/鞋品测试")
    parser.add_argument(
        "--report-xlsx",
        default="/Users/xingyicheng/Downloads/鞋品测试/深绘鞋品上新图包整理结果_20260822-202441.xlsx",
    )
    parser.add_argument(
        "--category-xlsx",
        default="/Users/xingyicheng/Downloads/鞋品品类映射模板测试.xlsx",
    )
    parser.add_argument(
        "--output-root",
        default="/Users/xingyicheng/Downloads/鞋品测试-最终重跑校验-20260824-Sol链",
    )
    parser.add_argument(
        "--pose-models",
        default="gpt-5.6-sol,gpt-5.6-terra,gpt-5.6-luna,gpt-5.5",
        help="Comma-separated model chain for pose recognition; first is primary.",
    )
    parser.add_argument(
        "--label-models",
        default="gpt-5.6-sol,gemini-3.5-flash,qwen3.7-plus,gpt-5.6-terra,kimi-k2.7-code",
        help="Comma-separated model chain for OCR; first is primary.",
    )
    parser.add_argument(
        "--pose-parallelism",
        type=int,
        default=shoe.SHOE_POSE_MAX_CONCURRENT_CALLS,
    )
    parser.add_argument("--pose-timeout", type=float, default=shoe.SHOE_POSE_MODEL_TIMEOUT_SECONDS)
    parser.add_argument(
        "--pose-strategy",
        default=shoe.SHOE_POSE_DEFAULT_STRATEGY,
        choices=[
            shoe.SHOE_POSE_STRATEGY_GLOBAL_PAGES,
            shoe.SHOE_POSE_STRATEGY_BATCH,
            shoe.SHOE_POSE_STRATEGY_BATCH_OVERVIEW,
            shoe.SHOE_POSE_STRATEGY_SINGLE_SHEET,
        ],
    )
    parser.add_argument("--label-timeout", type=float, default=shoe.SHOE_LABEL_OCR_TIMEOUT_SECONDS)
    parser.add_argument("--copy-final", action="store_true")
    parser.add_argument(
        "--finalize-existing",
        action="store_true",
        help="Approve and copy an existing passing attempt without rerunning models.",
    )
    parser.add_argument(
        "--visual-review-note",
        default="",
        help="Optional note recorded with --finalize-existing approval evidence.",
    )
    args = parser.parse_args()

    style = text(args.style)
    output_root = Path(args.output_root)
    attempt_root = _existing_attempt_path(output_root, style, args.attempt)
    if args.finalize_existing:
        _summary, status = finalize_existing_attempt(
            output_root=output_root,
            attempt_root=attempt_root,
            style=style,
            visual_review_note=text(args.visual_review_note),
        )
        return status
    if attempt_root.exists():
        shutil.rmtree(attempt_root)
    attempt_root.mkdir(parents=True, exist_ok=True)

    source_root = Path(args.source_root)
    category_rows = rows_from_xlsx(Path(args.category_xlsx))
    shoe_categories = shoe.parse_shoe_category_rows(category_rows)
    data_rows = prepared_rows(source_root, Path(args.report_xlsx), style)
    logs: list[str] = []
    pose_models = [text(item) for item in args.pose_models.split(",") if text(item)]
    label_models = [text(item) for item in args.label_models.split(",") if text(item)]
    if not pose_models:
        pose_models = ["gpt-5.6-sol"]
    if not label_models:
        label_models = ["gpt-5.6-sol"]
    pose_parallelism = max(1, int(args.pose_parallelism))
    shoe.SHOE_POSE_BATCH_PARALLELISM = pose_parallelism
    shoe.SHOE_POSE_MAX_CONCURRENT_CALLS = pose_parallelism
    shoe.SHOE_POSE_MODEL_TIMEOUT_SECONDS = max(1.0, float(args.pose_timeout))
    shoe.SHOE_LABEL_OCR_TIMEOUT_SECONDS = max(1.0, float(args.label_timeout))

    start = time.time()
    try:
        report_rows, package_roots = shoe.prepare_shoe_packages(
            data_rows=data_rows,
            output_root=attempt_root,
            model_id=pose_models[0],
            pose_strategy=args.pose_strategy,
            fallback_model_ids=pose_models[1:],
            label_model_id=label_models[0],
            label_fallback_model_ids=label_models[1:],
            shoe_categories=shoe_categories,
            log=logs.append,
            preserve_analysis_artifacts=True,
        )
    except Exception as exc:
        elapsed = round(time.time() - start, 2)
        summary = {
            "style": style,
            "attempt": args.attempt,
            "elapsed": elapsed,
            "category": shoe_categories.get(style, ""),
            "pose_models": pose_models,
            "pose_strategy": args.pose_strategy,
            "label_models": label_models,
            "pose_parallelism": shoe.SHOE_POSE_MAX_CONCURRENT_CALLS,
            "pose_timeout": shoe.SHOE_POSE_MODEL_TIMEOUT_SECONDS,
            "label_timeout": shoe.SHOE_LABEL_OCR_TIMEOUT_SECONDS,
            "issues": [f"{type(exc).__name__}: {text(exc)}"],
            "metrics": summarize_logs(logs),
            "logs_tail": logs[-80:],
        }
        (attempt_root / "logs.txt").write_text("\n".join(logs), encoding="utf-8")
        (attempt_root / "validation.json").write_text(
            json.dumps(summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 1
    elapsed = round(time.time() - start, 2)
    style_root = package_roots.get(style, attempt_root / style)
    report_json = attempt_root / "report_rows.json"
    report_json.write_text(json.dumps(report_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    write_report_xlsx(attempt_root / "report.xlsx", report_rows)
    (attempt_root / "logs.txt").write_text("\n".join(logs), encoding="utf-8")
    visual_review = create_visual_review_sheet(
        style_root,
        report_rows,
        attempt_root / "visual-review.jpg",
    )
    contact_sheets = collect_contact_sheet_artifacts(attempt_root)
    analysis_artifacts = collect_analysis_artifacts(attempt_root)

    issues, validation_warnings = validate_style(
        style=style,
        style_root=style_root,
        report_rows=report_rows,
        category=shoe_categories.get(style, ""),
    )
    if visual_review is None or not visual_review.is_file():
        issues.append("visual-review evidence missing")
    if not contact_sheets:
        issues.append("contact sheet evidence missing")
    color_dirs = sorted(
        [
            path.name
            for path in style_root.iterdir()
            if path.is_dir() and re.match(r"^\d+\.", path.name)
        ],
    ) if style_root.is_dir() else []
    summary = {
        "style": style,
        "attempt": args.attempt,
        "elapsed": elapsed,
        "style_root": str(style_root),
        "report_rows": len(report_rows),
        "report_xlsx": str(attempt_root / "report.xlsx"),
        "report_json": str(report_json),
        "visual_review": str(visual_review) if visual_review else "",
        "contact_sheets": [str(path) for path in contact_sheets],
        "analysis_artifacts": [str(path) for path in analysis_artifacts],
        "color_dirs": color_dirs,
        "category": shoe_categories.get(style, ""),
        "pose_models": pose_models,
        "pose_strategy": args.pose_strategy,
        "label_models": label_models,
        "pose_parallelism": shoe.SHOE_POSE_MAX_CONCURRENT_CALLS,
        "pose_timeout": shoe.SHOE_POSE_MODEL_TIMEOUT_SECONDS,
        "label_timeout": shoe.SHOE_LABEL_OCR_TIMEOUT_SECONDS,
        "issues": issues,
        "warnings": validation_warnings,
        "metrics": summarize_logs(logs),
        "logs_tail": logs[-50:],
    }
    if not issues and args.copy_final:
        summary.update(
            copy_final_artifacts(
                output_root=output_root,
                attempt_root=attempt_root,
                style=style,
                style_root=style_root,
                contact_sheets=contact_sheets,
                analysis_artifacts=analysis_artifacts,
            )
        )
    validation_path = attempt_root / "validation.json"
    validation_path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    final_evidence_root = text(summary.get("final_evidence_root"))
    if final_evidence_root:
        shutil.copy2(validation_path, Path(final_evidence_root) / "validation.json")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0 if not issues else 2


if __name__ == "__main__":
    raise SystemExit(main())
