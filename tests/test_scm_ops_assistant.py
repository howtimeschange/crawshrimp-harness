from pathlib import Path
from importlib.util import find_spec
import os
import tempfile
import unittest
from unittest.mock import patch

import yaml

from core import data_sink
from core.api_server import _finalize_scm_ops_assistant_outputs


MANIFEST_PATH = Path("adapters/scm-ops-assistant/manifest.yaml")


class ScmOpsAssistantTests(unittest.TestCase):
    def test_manifest_declares_wash_hangtag_batch_download(self):
        manifest = yaml.safe_load(MANIFEST_PATH.read_text(encoding="utf-8"))
        tasks = {item["id"]: item for item in manifest["tasks"]}
        task = tasks["wash_hangtag_batch_download"]
        params = {item["id"]: item for item in task["params"]}
        output = task["output"][0]

        self.assertEqual(manifest["id"], "scm-ops-assistant")
        self.assertEqual(manifest["name"], "SCM 运营助手")
        self.assertEqual(task["name"], "批量下载洗唛吊牌")
        self.assertEqual(task["script"], "wash-hangtag-batch-download.js")
        self.assertEqual(task["entry_url"], "https://scm.semir.com/scm-quality-mgm/index/scm-qc-wash-appr-index")
        self.assertEqual(params["style_codes"]["type"], "textarea")
        self.assertTrue(params["style_codes"]["required"])
        self.assertEqual(params["export_folder"]["type"], "directory")
        self.assertTrue(params["export_folder"]["required"])
        self.assertTrue(params["only_completed"]["default"])
        self.assertNotIn("max_styles", params)
        self.assertNotIn("request_delay_ms", params)
        self.assertNotIn("download_concurrency", params)
        self.assertEqual(output["sheet_key"], "__sheet_name")
        self.assertIn("成分汇总", [sheet["name"] for sheet in output["sheets"]])
        self.assertIn("下载明细", [sheet["name"] for sheet in output["sheets"]])
        self.assertIn("异常", [sheet["name"] for sheet in output["sheets"]])
        self.assertIn("中文成分", output["columns"])
        self.assertIn("英文成分", output["columns"])

    def test_finalize_copies_excel_into_scm_package_root(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            package_root = base / "SCM洗唛吊牌_20260805"
            wash_file = package_root / "208426103215" / "洗唛文件" / "a.jpg"
            wash_file.parent.mkdir(parents=True)
            wash_file.write_bytes(b"jpg")
            exported = base / "runtime" / "summary.xlsx"
            exported.parent.mkdir()
            exported.write_bytes(b"xlsx")

            logs = []
            refs = _finalize_scm_ops_assistant_outputs(
                data_rows=[{"__scm_package_root": str(package_root), "款号": "208426103215"}],
                runtime_files=[str(wash_file)],
                exported_files=[str(exported)],
                run_params={"export_folder": str(base)},
                log=logs.append,
            )

            copied_excel = package_root / "summary.xlsx"
            self.assertTrue(copied_excel.is_file())
            self.assertIn(str(package_root), refs)
            self.assertIn(str(copied_excel), refs)
            self.assertIn(str(wash_file), refs)
            self.assertTrue(any("SCM 洗唛吊牌已导出" in item for item in logs))

    def test_export_excel_writes_configured_scm_sheets(self):
        if find_spec("openpyxl") is None:
            self.skipTest("openpyxl is required for export_excel in this environment")

        from openpyxl import load_workbook

        manifest = yaml.safe_load(MANIFEST_PATH.read_text(encoding="utf-8"))
        task = next(item for item in manifest["tasks"] if item["id"] == "wash_hangtag_batch_download")
        output = next(item for item in task["output"] if item["type"] == "excel")
        rows = [
            {
                "__sheet_name": "成分汇总",
                "款号": "208426103215",
                "中文成分": "成分主面料:100%聚酯纤维",
                "英文成分": "Composition:Main fabric:100%Polyester",
                "SKC数量": 1,
                "洗唛文件数": 1,
                "吊牌文件数": 1,
                "查询结果": "成功",
            },
            {
                "__sheet_name": "下载明细",
                "款号": "208426103215",
                "SKC": "20842610321500322",
                "文件类型": "洗唛文件",
                "文件名": "wash.jpg",
                "本地文件": "/tmp/pkg/208426103215/洗唛文件/wash.jpg",
                "文件大小": 430115,
                "下载结果": "已下载",
                "中文成分": "成分主面料:100%聚酯纤维",
                "英文成分": "Composition:Main fabric:100%Polyester",
            },
            {
                "__sheet_name": "异常",
                "款号": "208426103216",
                "文件类型": "吊牌文件",
                "下载结果": "下载失败",
                "查询结果": "失败",
                "备注": "HTTP 403",
            },
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            with patch.dict(os.environ, {"CRAWSHRIMP_DATA": tmpdir}, clear=False):
                out_path = data_sink.export_excel(
                    rows,
                    adapter_id="scm-ops-assistant",
                    task_id="wash_hangtag_batch_download",
                    filename_template="test.xlsx",
                    sheet_key=output["sheet_key"],
                    sheet_configs=output["sheets"],
                )

            wb = load_workbook(out_path, read_only=True, data_only=True)
            self.assertEqual(wb.sheetnames, ["成分汇总", "下载明细", "异常"])

            summary_headers = [cell.value for cell in next(wb["成分汇总"].iter_rows(min_row=1, max_row=1))]
            detail_headers = [cell.value for cell in next(wb["下载明细"].iter_rows(min_row=1, max_row=1))]
            exception_headers = [cell.value for cell in next(wb["异常"].iter_rows(min_row=1, max_row=1))]
            self.assertIn("中文成分", summary_headers)
            self.assertIn("英文成分", summary_headers)
            self.assertIn("本地文件", detail_headers)
            self.assertIn("备注", exception_headers)
            self.assertEqual(wb["成分汇总"]["A2"].value, "208426103215")
            self.assertEqual(wb["下载明细"]["T2"].value, "已下载")
            self.assertEqual(wb["异常"]["J2"].value, "HTTP 403")


if __name__ == "__main__":
    unittest.main()
