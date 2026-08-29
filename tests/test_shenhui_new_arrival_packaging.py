import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest.mock import patch

import yaml
from openpyxl import Workbook, load_workbook
from PIL import Image

from core.api_server import (
    _SHENHUI_NEW_ARRIVAL_SINGLE_IMAGE_THRESHOLD_BYTES,
    _cleanup_orphaned_runtime_artifacts,
    _compress_shenhui_label_tile_image_if_beneficial,
    _finalize_shenhui_new_arrival_outputs,
    _prepare_shenhui_shoe_package_rows,
    _serialize_task_param,
    _shenhui_shoe_box_label_candidate_result,
)
from core.models import AdapterManifest
from core.shenhui_apparel_label_processing import ApparelLabelProcessingResult

ROOT = Path(__file__).resolve().parents[1]
MANIFEST_PATH = ROOT / "adapters" / "shenhui-new-arrival" / "manifest.yaml"
SHOE_PACKAGING_PATH = ROOT / "core" / "shenhui_shoe_packaging.py"


class ShenhuiNewArrivalPackagingTests(unittest.TestCase):
    def test_apparel_label_processor_is_called_only_for_prepare_upload_package(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            calls = []

            def fake_processor(**kwargs):
                calls.append(kwargs)
                return ApparelLabelProcessingResult(0, (), (), (), ())

            prepare_runtime = base / "prepare-runtime"
            prepare_runtime.mkdir()
            source = prepare_runtime / "model.jpg"
            Image.new("RGB", (20, 20), "white").save(source)
            with patch(
                "core.api_server.process_prepare_upload_package_labels",
                side_effect=fake_processor,
            ), patch(
                "core.api_server.finalize_pdf_batch_screenshot_outputs",
                return_value=[],
            ):
                _finalize_shenhui_new_arrival_outputs(
                    task_id="prepare_upload_package",
                    data_rows=[{
                        "输入款号": "202426107206",
                        "输入编码": "202426107206",
                        "文件名": "model.jpg",
                        "下载结果": "已下载",
                        "本地文件": str(source),
                        "__shenhui_group_code": "202426107206",
                        "__shenhui_asset_role": "image",
                        "__package_filename": "model.jpg",
                    }],
                    runtime_files=[str(source)],
                    exported_files=[],
                    run_params={"package_name": "prepare-only"},
                    runtime_artifact_dir=str(prepare_runtime),
                    log=lambda _message: None,
                )

                for task_id in (
                    "batch_label_tile_download",
                    "pdf_batch_screenshot",
                    "prepare_shoe_upload_package",
                    "unrelated_task",
                ):
                    runtime = base / f"runtime-{task_id}"
                    runtime.mkdir(exist_ok=True)
                    _finalize_shenhui_new_arrival_outputs(
                        task_id=task_id,
                        data_rows=[],
                        runtime_files=[],
                        exported_files=[],
                        run_params={},
                        runtime_artifact_dir=str(runtime),
                        log=lambda _message: None,
                    )

            self.assertEqual(len(calls), 1)
            self.assertEqual(calls[0]["package_root"].name, "prepare-only")

    def test_new_arrival_compression_single_image_threshold_is_30mb(self):
        self.assertEqual(
            _SHENHUI_NEW_ARRIVAL_SINGLE_IMAGE_THRESHOLD_BYTES,
            30 * 1024 * 1024,
        )

    def test_shoe_packaging_core_is_separate_from_clothing_packaging(self):
        self.assertTrue(SHOE_PACKAGING_PATH.is_file())

    def test_manifest_declares_separate_shoe_upload_package_task(self):
        manifest = yaml.safe_load(MANIFEST_PATH.read_text(encoding="utf-8"))
        task = next(
            item
            for item in manifest["tasks"]
            if item["id"] == "prepare_shoe_upload_package"
        )

        self.assertEqual(task["name"], "【鞋品】整理深绘上新图包")
        self.assertEqual(task["script"], "prepare-shoe-upload-package.js")
        self.assertFalse(task["skip_auth"])
        params = {item["id"]: item for item in task["params"]}
        self.assertNotIn("item_codes", params)
        self.assertNotIn("folder_scan_depth", params)
        self.assertNotIn("download_concurrency", params)
        self.assertEqual(params["shoe_cloud_path"]["type"], "text")
        self.assertEqual(params["shoe_category_file"]["type"], "file_excel")
        self.assertTrue(params["shoe_category_file"].get("required"))
        self.assertEqual(params["mode"]["default"], "new")
        mode_options = {
            option["value"]: option["label"]
            for option in params["mode"]["options"]
        }
        self.assertEqual(mode_options["current"], "当前页面")
        self.assertEqual(mode_options["new"], "全新页面（推荐）")
        self.assertEqual(
            params["shoe_category_file"]["template_file"],
            "assets/鞋品品类映射模板.xlsx",
        )
        self.assertEqual(params["model_chain"]["type"], "model_chain")
        self.assertEqual(params["model_chain"]["label"], "模型策略")
        self.assertEqual(params["model_chain"]["ui_span"], "full")
        default_model = params["model_chain"]["default_model"]
        fallback_models = {
            item["id"]: item
            for item in params["model_chain"]["fallback_models"]
        }
        self.assertEqual(default_model["id"], "model_id")
        self.assertEqual(default_model["label"], "默认模型")
        self.assertEqual(default_model["default"], "gpt-5.6-sol")
        self.assertEqual(
            fallback_models["fallback_model_1"]["default"],
            "gpt-5.6-terra",
        )
        self.assertEqual(fallback_models["fallback_model_1"]["label"], "备选模型 1")
        self.assertEqual(fallback_models["fallback_model_2"]["default"], "gpt-5.6-luna")
        self.assertEqual(fallback_models["fallback_model_2"]["label"], "备选模型 2")
        self.assertEqual(fallback_models["fallback_model_3"]["default"], "gpt-5.5")
        self.assertEqual(fallback_models["fallback_model_3"]["label"], "备选模型 3")
        self.assertEqual(
            fallback_models["fallback_model_4"]["default"],
            "deepseek-official-v4-flash-vision-exp",
        )
        self.assertEqual(fallback_models["fallback_model_4"]["label"], "备选模型 4")
        self.assertEqual(fallback_models["fallback_model_5"]["default"], "kimi-k2.7-code")
        self.assertEqual(fallback_models["fallback_model_5"]["label"], "备选模型 5")
        self.assertNotIn("Fallback", params["model_chain"]["hint"])
        self.assertIn("不使用", [
            option["label"]
            for option in fallback_models["fallback_model_1"]["options"]
        ])
        self.assertIn(
            "multi-model",
            [option["value"] for option in default_model["options"]],
        )
        self.assertIn(
            "gpt-5.5",
            [option["value"] for option in default_model["options"]],
        )
        self.assertIn(
            "qwen3.7-plus",
            [option["value"] for option in default_model["options"]],
        )
        self.assertTrue(
            {
                "deepseek-official-v4-flash-vision-exp",
                "deepseek-official-v4-flash",
                "deepseek-official-v4-pro",
                "glm-official-5.3-flash",
                "deepseek-v4-flash",
                "deepseek-v4-pro",
                "glm-5.2",
                "kimi-k3",
                "kimi-k2.7-code",
            }.issubset(
                {
                    option["value"]
                    for option in default_model["options"]
                }
            )
        )
        self.assertEqual(params["shoe_pose_strategy"]["default"], "single_sheet")
        self.assertEqual(
            [
                option["value"]
                for option in params["shoe_pose_strategy"]["options"]
            ],
            ["single_sheet", "global_pages", "batch_overview", "batch"],
        )
        self.assertEqual(params["export_folder"]["type"], "directory")

        output_columns = task["output"][0]["columns"]
        self.assertIn("品类来源", output_columns)
        self.assertIn("规则槽位", output_columns)
        self.assertIn("规则告警", output_columns)
        self.assertIn("压缩结果", output_columns)

    def test_shoe_model_chain_param_survives_manifest_validation(self):
        manifest = yaml.safe_load(MANIFEST_PATH.read_text(encoding="utf-8"))
        adapter = AdapterManifest(**manifest)
        task = next(
            item
            for item in adapter.tasks
            if item.id == "prepare_shoe_upload_package"
        )
        param = next(item for item in task.params if item.id == "model_chain")

        self.assertEqual(param.type.value, "model_chain")
        self.assertEqual(param.default_model["id"], "model_id")
        self.assertEqual(param.fallback_models[0]["id"], "fallback_model_1")
        self.assertEqual(param.fallback_models[0]["label"], "备选模型 1")

        serialized = _serialize_task_param("shenhui-new-arrival", param)
        self.assertEqual(serialized["default_model"]["default"], "gpt-5.6-sol")
        self.assertEqual(
            serialized["fallback_models"][0]["default"],
            "gpt-5.6-terra",
        )

    def test_manifest_declares_deepdraw_upload_task_with_fail_closed_controls(self):
        manifest = yaml.safe_load(MANIFEST_PATH.read_text(encoding="utf-8"))
        task = next(item for item in manifest["tasks"] if item["id"] == "upload_to_deepdraw")

        self.assertEqual(task["script"], "upload-to-deepdraw.js")
        self.assertEqual(task["entry_url"], "https://www.deepdraw.biz/authorized/merchant/index")
        self.assertTrue(task["skip_auth"])
        self.assertIn(
            "https://www.deepdraw.biz/authorized/merchant/index",
            task["tab_match_prefixes"],
        )

        params = {item["id"]: item for item in task["params"]}
        self.assertEqual(params["package_zip_paths"]["type"], "file_zip")
        self.assertEqual(params["upload_mode"]["default"], "dry_run")
        self.assertNotIn("production_confirm_text", params)
        self.assertNotIn("max_upload_count", params)

    def test_manifest_declares_separate_pdf_batch_screenshot_task(self):
        manifest = yaml.safe_load(MANIFEST_PATH.read_text(encoding="utf-8"))
        self.assertEqual(
            [item["id"] for item in manifest["tasks"]],
            [
                "batch_label_tile_download",
                "prepare_upload_package",
                "prepare_shoe_upload_package",
                "upload_to_deepdraw",
                "pdf_batch_screenshot",
            ],
        )
        task = next(item for item in manifest["tasks"] if item["id"] == "pdf_batch_screenshot")

        self.assertEqual(task["script"], "pdf-batch-screenshot.js")
        self.assertEqual(task["entry_url"], "about:blank")
        self.assertTrue(task["skip_auth"])
        params = {item["id"]: item for item in task["params"]}
        self.assertEqual(params["wash_pdf_files"]["type"], "file_pdf")
        self.assertEqual(params["tag_pdf_files"]["type"], "file_pdf")
        self.assertNotIn("pdf_files", params)
        self.assertNotIn("pdf_type", params)
        self.assertEqual(params["wash_crop_boxes"]["type"], "textarea")
        self.assertEqual(params["wash_crop_boxes"]["default"], '[{"x":0.0892,"y":0.2084,"width":0.4189,"height":0.7546}]')
        self.assertEqual(params["tag_crop_boxes"]["type"], "textarea")
        self.assertEqual(params["tag_crop_boxes"]["default"], '[{"x":0.0113,"y":0.2352,"width":0.1535,"height":0.5058}]')
        self.assertEqual(params["output_mode"]["type"], "select")
        self.assertEqual(params["output_mode"]["default"], "create_package")
        self.assertNotIn("style_color_overrides", params)

        prepare_task = next(item for item in manifest["tasks"] if item["id"] == "prepare_upload_package")
        self.assertEqual(prepare_task["name"], "【服饰】整理深绘上新图包")
        prepare_params = {item["id"]: item for item in prepare_task["params"]}
        self.assertEqual(prepare_params["mode"]["default"], "new")
        prepare_mode_options = {
            option["value"]: option["label"]
            for option in prepare_params["mode"]["options"]
        }
        self.assertEqual(prepare_mode_options["current"], "当前页面")
        self.assertEqual(prepare_mode_options["new"], "全新页面（推荐）")
        self.assertEqual(prepare_params["image_source_type"]["type"], "radio")
        self.assertEqual(prepare_params["image_source_type"]["default"], "all")
        self.assertEqual(
            [item["value"] for item in prepare_params["image_source_type"]["options"]],
            ["all", "still", "model"],
        )
        self.assertEqual(prepare_params["wash_crop_boxes"]["type"], "textarea")
        self.assertEqual(prepare_params["tag_crop_boxes"]["type"], "textarea")
        self.assertEqual(prepare_params["auto_zip_package"]["type"], "checkbox")
        self.assertEqual(prepare_params["auto_zip_package"]["default"], [])
        self.assertEqual(prepare_params["auto_zip_package"]["options"][0]["value"], "yes")
        self.assertIn("压缩结果", prepare_task["output"][0]["columns"])

    def test_manifest_declares_batch_label_tile_download_task(self):
        manifest = yaml.safe_load(MANIFEST_PATH.read_text(encoding="utf-8"))
        task = next(item for item in manifest["tasks"] if item["id"] == "batch_label_tile_download")

        self.assertEqual(task["name"], "批量下载吊牌/洗唛/平铺图")
        self.assertEqual(task["script"], "batch-label-tile-download.js")
        self.assertFalse(task["skip_auth"])
        params = {item["id"]: item for item in task["params"]}
        self.assertEqual(params["mode"]["default"], "new")
        self.assertEqual(params["still_cloud_path"]["type"], "text")
        self.assertTrue(params["still_cloud_path"]["required"])
        self.assertEqual(params["model_cloud_path"]["type"], "text")
        self.assertFalse(params["model_cloud_path"]["required"])
        self.assertEqual(params["item_codes"]["type"], "textarea")
        self.assertTrue(params["item_codes"]["required"])
        self.assertEqual(params["export_folder"]["type"], "directory")

        output_columns = task["output"][0]["columns"]
        self.assertIn("素材类型", output_columns)
        self.assertIn("匹配策略", output_columns)
        self.assertIn("模拍路径命中", output_columns)

    def test_finalize_outputs_groups_images_by_style_code_without_zip_by_default(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            runtime_dir.mkdir()

            model_file = runtime_dir / "runtime-model.jpg"
            yq_file = runtime_dir / "runtime-yq.jpg"
            model_file.write_bytes(b"model")
            yq_file.write_bytes(b"yq")

            exported = base / "summary.xlsx"
            exported.write_bytes(b"excel")

            result = _finalize_shenhui_new_arrival_outputs(
                task_id="prepare_upload_package",
                data_rows=[
                    {
                        "输入款号": "208226103201",
                        "输入编码": "208226103201",
                        "素材来源": "模特图",
                        "文件名": "balaBR05106-72904_P.jpg",
                        "下载结果": "已下载",
                        "本地文件": str(model_file),
                        "__shenhui_group_code": "208226103201",
                        "__shenhui_asset_role": "image",
                        "__package_filename": "balaBR05106-72904_P.jpg",
                    },
                    {
                        "输入款号": "208226103201",
                        "输入编码": "208226103201",
                        "素材来源": "静物图",
                        "文件名": "yq.jpg",
                        "下载结果": "已下载",
                        "本地文件": str(yq_file),
                        "__shenhui_group_code": "208226103201",
                        "__shenhui_asset_role": "image",
                        "__package_filename": "yq.jpg",
                    },
                ],
                runtime_files=[str(model_file), str(yq_file)],
                exported_files=[str(exported)],
                run_params={"package_name": "深绘测试图包"},
                runtime_artifact_dir=str(runtime_dir),
                log=lambda _: None,
            )

            self.assertEqual(len(result), 2)
            package_dir = Path(result[0])
            self.assertTrue(package_dir.is_dir())
            self.assertEqual(package_dir.name, "深绘测试图包")
            self.assertEqual(Path(result[1]), exported)
            self.assertTrue((package_dir / "208226103201" / "balaBR05106-72904_P.jpg").is_file())
            self.assertTrue((package_dir / "208226103201" / "yq.jpg").is_file())

            self.assertFalse(model_file.exists())
            self.assertFalse(yq_file.exists())
            self.assertFalse((runtime_dir / "深绘测试图包").exists())

    def test_finalize_prepare_upload_package_compresses_single_large_images(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            export_dir = base / "exports"
            runtime_dir.mkdir()
            export_dir.mkdir()

            image_file = runtime_dir / "runtime-large.jpg"
            Image.effect_noise((512, 512), 90).convert("RGB").save(
                image_file,
                format="JPEG",
                quality=100,
            )
            before_size = image_file.stat().st_size

            exported = base / "summary.xlsx"
            data_rows = [{
                "输入款号": "208226103201",
                "输入编码": "208226103201",
                "素材来源": "模特图",
                "文件名": "balaBR05106-72904_P.jpg",
                "云盘路径": "模拍原图/208226103201/balaBR05106-72904_P.jpg",
                "处理动作": "保留模特图",
                "下载结果": "已下载",
                "本地文件": str(image_file),
                "压缩结果": "",
                "备注": "",
                "__shenhui_group_code": "208226103201",
                "__shenhui_asset_role": "image",
                "__package_filename": "balaBR05106-72904_P.jpg",
            }]
            columns = [
                "输入款号",
                "输入编码",
                "素材来源",
                "文件名",
                "云盘路径",
                "处理动作",
                "下载结果",
                "本地文件",
                "压缩结果",
                "备注",
            ]
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(columns)
            sheet.append([data_rows[0].get(column, "") for column in columns])
            workbook.save(exported)
            workbook.close()

            with patch(
                "core.api_server._SHENHUI_NEW_ARRIVAL_SINGLE_IMAGE_THRESHOLD_BYTES",
                max(1, before_size - 1),
            ), patch(
                "core.api_server._SHENHUI_NEW_ARRIVAL_STYLE_PACKAGE_THRESHOLD_BYTES",
                before_size * 10,
            ):
                _finalize_shenhui_new_arrival_outputs(
                    task_id="prepare_upload_package",
                    data_rows=data_rows,
                    runtime_files=[str(image_file)],
                    exported_files=[str(exported)],
                    run_params={
                        "package_name": "深绘测试图包",
                        "export_folder": str(export_dir),
                    },
                    runtime_artifact_dir=str(runtime_dir),
                    log=lambda _: None,
                )

            final_image = export_dir / "深绘测试图包" / "208226103201" / "balaBR05106-72904_P.jpg"
            copied_excel = export_dir / "summary.xlsx"
            self.assertTrue(final_image.is_file())
            self.assertLess(final_image.stat().st_size, before_size)
            with Image.open(final_image) as compressed:
                self.assertEqual(compressed.size, (512, 512))
            self.assertIn("单图超过", data_rows[0]["压缩结果"])

            workbook = load_workbook(copied_excel, read_only=True, data_only=True)
            self.assertEqual(workbook.active["H2"].value, str(final_image))
            self.assertTrue(Path(workbook.active["H2"].value).is_file())
            self.assertIn("单图超过", workbook.active["I2"].value)
            workbook.close()

    def test_finalize_batch_label_tile_download_moves_files_under_style_folder(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            export_dir = base / "downloads"
            runtime_dir.mkdir()
            export_dir.mkdir()

            tile_file = runtime_dir / "runtime-tile.jpg"
            tag_file = runtime_dir / "runtime-tag.jpg"
            tile_file.write_bytes(b"tile")
            tag_file.write_bytes(b"tag")

            exported = base / "summary.xlsx"
            data_rows = [
                {
                    "输入款号": "208426108223",
                    "输入编码": "208426108223",
                    "素材类型": "平铺图",
                    "素材来源": "模拍路径",
                    "文件名": "208426108223-00316_有模拍.jpg",
                    "云盘路径": "模拍原图/208426108223/208426108223-00316.jpg",
                    "匹配策略": "优先从模拍路径查找平铺图",
                    "模拍路径命中": "是",
                    "下载结果": "已下载",
                    "本地文件": str(tile_file),
                    "备注": "",
                    "__shenhui_group_code": "208426108223",
                    "__package_filename": "208426108223-00316_有模拍.jpg",
                },
                {
                    "输入款号": "208426108223",
                    "输入编码": "208426108223",
                    "素材类型": "吊牌",
                    "素材来源": "平拍路径",
                    "文件名": "208426108223_吊牌_yq1.jpg",
                    "云盘路径": "平拍原图/208426108223/yq1.jpg",
                    "匹配策略": "优先命中 yq1",
                    "模拍路径命中": "是",
                    "下载结果": "已下载",
                    "本地文件": str(tag_file),
                    "备注": "",
                    "__shenhui_group_code": "208426108223",
                    "__package_filename": "208426108223_吊牌_yq1.jpg",
                },
            ]
            workbook = Workbook()
            sheet = workbook.active
            columns = [
                "输入款号",
                "输入编码",
                "素材类型",
                "素材来源",
                "文件名",
                "云盘路径",
                "匹配策略",
                "模拍路径命中",
                "下载结果",
                "本地文件",
                "备注",
            ]
            sheet.append(columns)
            for row in data_rows:
                sheet.append([row.get(column, "") for column in columns])
            workbook.save(exported)
            workbook.close()

            result = _finalize_shenhui_new_arrival_outputs(
                task_id="batch_label_tile_download",
                data_rows=data_rows,
                runtime_files=[str(tile_file), str(tag_file)],
                exported_files=[str(exported)],
                run_params={
                    "package_name": "测试吊牌洗唛平铺图",
                    "export_folder": str(export_dir),
                },
                runtime_artifact_dir=str(runtime_dir),
                log=lambda _: None,
            )

            package_dir = export_dir / "测试吊牌洗唛平铺图"
            copied_excel = package_dir / "summary.xlsx"
            self.assertEqual({Path(path) for path in result}, {package_dir, copied_excel})
            self.assertTrue((package_dir / "208426108223" / "208426108223-00316_有模拍.jpg").is_file())
            self.assertTrue((package_dir / "208426108223" / "208426108223_吊牌_yq1.jpg").is_file())
            self.assertFalse(tile_file.exists())
            self.assertFalse(tag_file.exists())
            self.assertEqual(
                data_rows[0]["本地文件"],
                str(package_dir / "208426108223" / "208426108223-00316_有模拍.jpg"),
            )

            workbook = load_workbook(copied_excel, read_only=True, data_only=True)
            self.assertEqual(
                workbook.active["J2"].value,
                str(package_dir / "208426108223" / "208426108223-00316_有模拍.jpg"),
            )
            workbook.close()

    def test_finalize_batch_label_tile_download_keeps_only_ocr_selected_shoe_label_candidate(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            export_dir = base / "downloads"
            runtime_dir.mkdir()
            export_dir.mkdir()

            style_file = runtime_dir / "runtime-style.jpg"
            box_file = runtime_dir / "runtime-box.jpg"
            detail_file = runtime_dir / "runtime-detail.jpg"
            for path in (style_file, box_file, detail_file):
                Image.new("RGB", (20, 20), "white").save(path)

            exported = base / "summary.xlsx"
            data_rows = [
                {
                    "输入款号": "204426141122",
                    "输入编码": "204426141122",
                    "素材类型": "款色图",
                    "素材来源": "平拍路径",
                    "文件名": "204426141122-00322.jpg",
                    "云盘路径": "鞋品/204426141122-已写/00322/36/204426141122-00322.jpg",
                    "匹配策略": "鞋品仅保留每个款色的款色命名图",
                    "模拍路径命中": "否",
                    "下载结果": "已下载",
                    "本地文件": str(style_file),
                    "备注": "",
                    "__shenhui_group_code": "204426141122",
                    "__shenhui_asset_role": "shoe_style_color",
                    "__package_filename": "204426141122-00322.jpg",
                    "__shoe_color_code": "00322",
                },
                {
                    "输入款号": "204426141122",
                    "输入编码": "204426141122",
                    "素材类型": "鞋盒标签图/电子吊牌图",
                    "素材来源": "平拍路径",
                    "文件名": "GUDO6815.jpg",
                    "云盘路径": "鞋品/204426141122-已写/00322/36/GUDO6815.jpg",
                    "匹配策略": "鞋品下载少量无语义候选，后端 OCR 识别鞋盒标签",
                    "模拍路径命中": "否",
                    "下载结果": "已下载",
                    "本地文件": str(box_file),
                    "备注": "",
                    "__shenhui_group_code": "204426141122",
                    "__shenhui_asset_role": "shoe_label",
                    "__package_filename": "GUDO6815.jpg",
                    "__shoe_color_code": "00322",
                    "__shoe_label_candidate_kind": "generic_ocr",
                },
                {
                    "输入款号": "204426141122",
                    "输入编码": "204426141122",
                    "素材类型": "鞋盒标签图/电子吊牌图",
                    "素材来源": "平拍路径",
                    "文件名": "GUDO6811.jpg",
                    "云盘路径": "鞋品/204426141122-已写/00322/36/GUDO6811.jpg",
                    "匹配策略": "鞋品下载少量无语义候选，后端 OCR 识别鞋盒标签",
                    "模拍路径命中": "否",
                    "下载结果": "已下载",
                    "本地文件": str(detail_file),
                    "备注": "",
                    "__shenhui_group_code": "204426141122",
                    "__shenhui_asset_role": "shoe_label",
                    "__package_filename": "GUDO6811.jpg",
                    "__shoe_color_code": "00322",
                    "__shoe_label_candidate_kind": "generic_ocr",
                },
            ]
            columns = [
                "输入款号",
                "输入编码",
                "素材类型",
                "素材来源",
                "文件名",
                "云盘路径",
                "匹配策略",
                "模拍路径命中",
                "下载结果",
                "本地文件",
                "备注",
            ]
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(columns)
            for row in data_rows:
                sheet.append([row.get(column, "") for column in columns])
            workbook.save(exported)
            workbook.close()

            def fake_detect(row, _local_path, _log):
                if row["文件名"] == "GUDO6815.jpg":
                    return {
                        "accepted": True,
                        "score": 0.56,
                        "confidence": 64,
                        "note": "OCR确认鞋盒标签：款号 204426141122 色号 00322",
                    }
                return {
                    "accepted": False,
                    "score": 0.08,
                    "confidence": 31,
                    "note": "鞋盒标签候选 OCR 未读到当前款号：204426141122",
                }

            with patch("core.api_server._shenhui_shoe_box_label_candidate_result", side_effect=fake_detect):
                _finalize_shenhui_new_arrival_outputs(
                    task_id="batch_label_tile_download",
                    data_rows=data_rows,
                    runtime_files=[str(style_file), str(box_file), str(detail_file)],
                    exported_files=[str(exported)],
                    run_params={
                        "package_name": "测试鞋品标签下载",
                        "export_folder": str(export_dir),
                    },
                    runtime_artifact_dir=str(runtime_dir),
                    log=lambda _: None,
                )

            package_dir = export_dir / "测试鞋品标签下载"
            final_style = package_dir / "204426141122" / "204426141122-00322.jpg"
            final_box = package_dir / "204426141122" / "GUDO6815.jpg"
            final_detail = package_dir / "204426141122" / "GUDO6811.jpg"
            copied_excel = package_dir / "summary.xlsx"

            self.assertTrue(final_style.is_file())
            self.assertTrue(final_box.is_file())
            self.assertFalse(final_detail.exists())
            self.assertEqual(data_rows[1]["下载结果"], "已下载")
            self.assertIn("OCR确认鞋盒标签", data_rows[1]["备注"])
            self.assertEqual(data_rows[2]["下载结果"], "已跳过")
            self.assertEqual(data_rows[2]["本地文件"], "")
            self.assertIn("未读到当前款号", data_rows[2]["备注"])

            workbook = load_workbook(copied_excel, read_only=True, data_only=True)
            rows = list(workbook.active.iter_rows(values_only=True))
            workbook.close()
            header = list(rows[0])
            download_result_index = header.index("下载结果")
            local_file_index = header.index("本地文件")
            note_index = header.index("备注")
            self.assertEqual(rows[2][download_result_index], "已下载")
            self.assertIn("GUDO6815.jpg", rows[2][local_file_index])
            self.assertEqual(rows[3][download_result_index], "已跳过")
            self.assertEqual(rows[3][local_file_index], None)
            self.assertIn("未读到当前款号", rows[3][note_index])

    def test_shoe_box_label_ocr_accepts_structured_label_without_title_words(self):
        row = {
            "输入款号": "204426141046",
            "__shenhui_group_code": "204426141046",
            "__shoe_color_code": "50301",
        }
        normalized_text = (
            "coobola204426141046GB30585-2024"
            "33-38合成革+织物691467871180620260530"
        )
        with patch(
            "core.api_server._shenhui_shoe_label_visual_metrics",
            return_value={
                "score": 0.612,
                "brown_ratio": 0.234,
                "white_ratio": 0.703,
                "dark_ratio": 0.007,
            },
        ), patch(
            "core.api_server._shenhui_shoe_label_ocr_text",
            return_value=(normalized_text, 46.0, normalized_text),
        ):
            result = _shenhui_shoe_box_label_candidate_result(row, Path("GUDO8190.jpg"), None)

        self.assertTrue(result["accepted"])
        self.assertIn("standard", result["note"])
        self.assertIn("barcode", result["note"])

    def test_shoe_box_label_ocr_uses_crop_when_full_image_misses_small_label(self):
        row = {
            "输入款号": "204426141124",
            "__shenhui_group_code": "204426141124",
            "__shoe_color_code": "20301",
        }
        full_text = "random shoe box background noise"
        crop_text = "balabala204426141124合格证执行标准QB/T4331-2021GB30585-2024"
        with patch(
            "core.api_server._shenhui_shoe_label_visual_metrics",
            return_value={
                "score": 0.679,
                "brown_ratio": 0.271,
                "white_ratio": 0.665,
                "dark_ratio": 0.007,
            },
        ), patch(
            "core.api_server._shenhui_shoe_label_ocr_text",
            return_value=(full_text, 29.0, full_text),
        ), patch(
            "core.api_server._shenhui_shoe_label_crop_ocr_text",
            return_value=(crop_text, 77.0, crop_text),
        ):
            result = _shenhui_shoe_box_label_candidate_result(row, Path("GUDO8513.jpg"), None)

        self.assertTrue(result["accepted"])
        self.assertIn("OCR裁剪确认", result["note"])
        self.assertEqual(result["confidence"], 77.0)

    def test_shoe_box_label_ocr_uses_style_digit_crop_when_crop_misreads_style(self):
        row = {
            "输入款号": "204426141124",
            "__shenhui_group_code": "204426141124",
            "__shoe_color_code": "90001",
        }
        full_text = "random shoe box background noise"
        crop_text = "balabala204424141124合格证执行标准QB/T4331-2021GB30585-2024"
        with patch(
            "core.api_server._shenhui_shoe_label_visual_metrics",
            return_value={
                "score": 0.649,
                "brown_ratio": 0.252,
                "white_ratio": 0.670,
                "dark_ratio": 0.008,
            },
        ), patch(
            "core.api_server._shenhui_shoe_label_ocr_text",
            return_value=(full_text, 29.0, full_text),
        ), patch(
            "core.api_server._shenhui_shoe_label_crop_ocr_text",
            return_value=(crop_text, 35.0, crop_text),
        ), patch(
            "core.api_server._shenhui_shoe_label_crop_style_code_text",
            return_value=("204426141124", 12.0, "204426141124"),
        ):
            result = _shenhui_shoe_box_label_candidate_result(row, Path("GUDO8526.jpg"), None)

        self.assertTrue(result["accepted"])
        self.assertIn("style_digit_crop", result["note"])
        self.assertEqual(result["confidence"], 35.0)

    def test_shoe_box_label_ocr_rejects_style_code_without_label_structure(self):
        row = {
            "输入款号": "204426141046",
            "__shenhui_group_code": "204426141046",
            "__shoe_color_code": "50301",
        }
        with patch(
            "core.api_server._shenhui_shoe_label_visual_metrics",
            return_value={
                "score": 0.612,
                "brown_ratio": 0.234,
                "white_ratio": 0.703,
                "dark_ratio": 0.007,
            },
        ), patch(
            "core.api_server._shenhui_shoe_label_ocr_text",
            return_value=("204426141046", 46.0, "204426141046"),
        ):
            result = _shenhui_shoe_box_label_candidate_result(row, Path("GUDO8190.jpg"), None)

        self.assertFalse(result["accepted"])
        self.assertIn("结构证据", result["note"])

    def test_finalize_batch_label_tile_download_compresses_images_and_updates_summary(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            export_dir = base / "downloads"
            runtime_dir.mkdir()
            export_dir.mkdir()

            image_file = runtime_dir / "runtime-large.jpg"
            Image.effect_noise((256, 256), 90).convert("RGB").save(
                image_file,
                format="JPEG",
                quality=100,
            )
            before_size = image_file.stat().st_size

            exported = base / "summary.xlsx"
            data_rows = [
                {
                    "输入款号": "208426108223",
                    "输入编码": "208426108223",
                    "素材类型": "吊牌",
                    "素材来源": "平拍路径",
                    "文件名": "208426108223_吊牌_yq1.jpg",
                    "云盘路径": "平拍原图/208426108223/yq1.jpg",
                    "匹配策略": "优先命中 yq1",
                    "模拍路径命中": "否",
                    "下载结果": "已下载",
                    "本地文件": str(image_file),
                    "备注": "",
                    "__shenhui_group_code": "208426108223",
                    "__package_filename": "208426108223_吊牌_yq1.jpg",
                },
            ]
            workbook = Workbook()
            sheet = workbook.active
            columns = [
                "输入款号",
                "输入编码",
                "素材类型",
                "素材来源",
                "文件名",
                "云盘路径",
                "匹配策略",
                "模拍路径命中",
                "下载结果",
                "本地文件",
                "备注",
            ]
            sheet.append(columns)
            for row in data_rows:
                sheet.append([row.get(column, "") for column in columns])
            workbook.save(exported)
            workbook.close()

            _finalize_shenhui_new_arrival_outputs(
                task_id="batch_label_tile_download",
                data_rows=data_rows,
                runtime_files=[str(image_file)],
                exported_files=[str(exported)],
                run_params={
                    "package_name": "测试压缩",
                    "export_folder": str(export_dir),
                },
                runtime_artifact_dir=str(runtime_dir),
                log=lambda _: None,
            )

            package_dir = export_dir / "测试压缩"
            final_image = package_dir / "208426108223" / "208426108223_吊牌_yq1.jpg"
            copied_excel = package_dir / "summary.xlsx"
            self.assertTrue(final_image.is_file())
            self.assertLess(final_image.stat().st_size, before_size)
            with Image.open(final_image) as compressed:
                self.assertEqual(compressed.size, (256, 256))
            self.assertIn("已压缩", data_rows[0]["备注"])
            self.assertIn("高保真", data_rows[0]["备注"])

            workbook = load_workbook(copied_excel, read_only=True, data_only=True)
            self.assertIn("已压缩", workbook.active["K2"].value)
            workbook.close()

    def test_label_tile_image_compression_selects_smallest_quality_that_preserves_detail(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            image_file = Path(tmpdir) / "label-tile-source.jpg"
            image = Image.new("RGB", (320, 240), "white")
            pixels = image.load()
            for y in range(image.height):
                for x in range(image.width):
                    pixels[x, y] = (
                        90 + (x * 110 // image.width),
                        110 + (y * 80 // image.height),
                        130 + ((x + y) * 60 // (image.width + image.height)),
                    )
            image.save(image_file, format="JPEG", quality=100)

            with patch(
                "core.api_server._shenhui_label_tile_candidate_preserves_detail",
                return_value=(True, {"pixel_rms": 0.12, "edge_rms": 0.34}),
            ):
                note, before_size, after_size = _compress_shenhui_label_tile_image_if_beneficial(
                    image_file,
                    None,
                )

            self.assertIn("q70", note)
            self.assertLess(after_size, before_size)
            with Image.open(image_file) as compressed:
                self.assertEqual(compressed.size, (320, 240))

    def test_finalize_batch_label_tile_download_compresses_pdfs_and_updates_summary(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            import fitz

            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            export_dir = base / "downloads"
            runtime_dir.mkdir()
            export_dir.mkdir()

            pdf_file = runtime_dir / "runtime-label.pdf"
            document = fitz.open()
            page = document.new_page()
            for index in range(200):
                page.insert_text((72, 72 + (index % 50) * 10), "深绘PDF压缩验证 " * 20)
            document.save(str(pdf_file), garbage=0, deflate=False)
            document.close()
            before_size = pdf_file.stat().st_size

            exported = base / "summary.xlsx"
            data_rows = [
                {
                    "输入款号": "208426108223",
                    "输入编码": "208426108223",
                    "素材类型": "洗唛",
                    "素材来源": "平拍路径",
                    "文件名": "208426108223_洗唛.pdf",
                    "云盘路径": "平拍原图/208426108223/208426108223.pdf",
                    "匹配策略": "按纯款号 PDF 兜底识别洗唛",
                    "模拍路径命中": "否",
                    "下载结果": "已下载",
                    "本地文件": str(pdf_file),
                    "备注": "",
                    "__shenhui_group_code": "208426108223",
                    "__package_filename": "208426108223_洗唛.pdf",
                },
            ]
            workbook = Workbook()
            sheet = workbook.active
            columns = [
                "输入款号",
                "输入编码",
                "素材类型",
                "素材来源",
                "文件名",
                "云盘路径",
                "匹配策略",
                "模拍路径命中",
                "下载结果",
                "本地文件",
                "备注",
            ]
            sheet.append(columns)
            for row in data_rows:
                sheet.append([row.get(column, "") for column in columns])
            workbook.save(exported)
            workbook.close()

            _finalize_shenhui_new_arrival_outputs(
                task_id="batch_label_tile_download",
                data_rows=data_rows,
                runtime_files=[str(pdf_file)],
                exported_files=[str(exported)],
                run_params={
                    "package_name": "测试PDF压缩",
                    "export_folder": str(export_dir),
                },
                runtime_artifact_dir=str(runtime_dir),
                log=lambda _: None,
            )

            package_dir = export_dir / "测试PDF压缩"
            final_pdf = package_dir / "208426108223" / "208426108223_洗唛.pdf"
            copied_excel = package_dir / "summary.xlsx"
            self.assertTrue(final_pdf.is_file())
            self.assertLess(final_pdf.stat().st_size, before_size)
            self.assertIn("已压缩PDF", data_rows[0]["备注"])

            workbook = load_workbook(copied_excel, read_only=True, data_only=True)
            self.assertIn("已压缩PDF", workbook.active["K2"].value)
            workbook.close()

    def test_finalize_shoe_outputs_copies_finished_style_folder_and_result_excel(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            package_dir = runtime_dir / "shoe-packages" / "204326141005"
            color_dir = package_dir / "1.白紫色调00317"
            export_dir = base / "downloads"
            color_dir.mkdir(parents=True)
            (package_dir / "tmz (1).jpg").write_bytes(b"tmz")
            (color_dir / "o.jpg").write_bytes(b"poster")
            (color_dir / "yk1.jpg").write_bytes(b"detail")
            exported = base / "深绘鞋品上新图包整理结果.xlsx"
            data_rows = [{
                "输入款号": "204326141005",
                "下载结果": "已下载",
                "本地文件": str(package_dir / "tmz (1).jpg"),
            }]
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(list(data_rows[0]))
            sheet.append(list(data_rows[0].values()))
            workbook.save(exported)
            workbook.close()
            raw_download = runtime_dir / "raw.jpg"
            raw_download.write_bytes(b"raw")

            result = _finalize_shenhui_new_arrival_outputs(
                task_id="prepare_shoe_upload_package",
                data_rows=data_rows,
                runtime_files=[str(raw_download)],
                exported_files=[str(exported)],
                run_params={
                    "export_folder": str(export_dir),
                    "__shenhui_shoe_package_refs": [str(package_dir)],
                },
                runtime_artifact_dir=str(runtime_dir),
                log=lambda _: None,
            )

            output_package = export_dir / "204326141005"
            self.assertEqual(
                {Path(path) for path in result},
                {
                    output_package,
                    export_dir / "深绘鞋品上新图包整理结果.xlsx",
                },
            )
            self.assertTrue((output_package / "tmz (1).jpg").is_file())
            self.assertTrue((output_package / "1.白紫色调00317" / "o.jpg").is_file())
            self.assertTrue((output_package / "1.白紫色调00317" / "yk1.jpg").is_file())
            self.assertFalse(raw_download.exists())
            expected_path = str(output_package / "tmz (1).jpg")
            self.assertEqual(data_rows[0]["本地文件"], expected_path)
            workbook = load_workbook(
                export_dir / "深绘鞋品上新图包整理结果.xlsx",
                read_only=True,
                data_only=True,
            )
            self.assertEqual(workbook.active["C2"].value, expected_path)
            workbook.close()

    def test_finalize_shoe_outputs_exports_downloaded_materials_when_recognition_failed(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            export_dir = base / "downloads"
            runtime_dir.mkdir()

            raw_file = runtime_dir / "downloaded-shoe.jpg"
            raw_file.write_bytes(b"raw-shoe")
            exported = base / "深绘鞋品上新图包整理结果.xlsx"
            data_rows = [{
                "输入款号": "204426146036",
                "颜色": "60301",
                "原文件名": "GUDO7378.jpg",
                "云盘路径": "鞋品/204426146036/60301/GUDO7378.jpg",
                "规则槽位": "",
                "输出文件名": "",
                "处理动作": "下载后识别鞋品姿势",
                "下载结果": "已下载",
                "本地文件": str(raw_file),
                "压缩结果": "",
                "规则告警": "",
                "品类来源": "",
                "备注": "",
                "__shenhui_group_code": "204426146036",
                "__shoe_color_code": "60301",
                "__shoe_original_filename": "GUDO7378.jpg",
            }]
            columns = [
                "输入款号",
                "颜色",
                "原文件名",
                "云盘路径",
                "规则槽位",
                "输出文件名",
                "处理动作",
                "下载结果",
                "本地文件",
                "压缩结果",
                "规则告警",
                "品类来源",
                "备注",
            ]
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(columns)
            sheet.append([data_rows[0].get(column, "") for column in columns])
            workbook.save(exported)
            workbook.close()

            result = _finalize_shenhui_new_arrival_outputs(
                task_id="prepare_shoe_upload_package",
                data_rows=data_rows,
                runtime_files=[str(raw_file)],
                exported_files=[str(exported)],
                run_params={"export_folder": str(export_dir)},
                runtime_artifact_dir=str(runtime_dir),
                log=lambda _: None,
            )

            fallback_dirs = [Path(path) for path in result if Path(path).is_dir()]
            self.assertEqual(len(fallback_dirs), 1)
            fallback_dir = fallback_dirs[0]
            relocated = fallback_dir / "204426146036" / "60301" / "GUDO7378.jpg"
            copied_excel = export_dir / "深绘鞋品上新图包整理结果.xlsx"
            self.assertTrue(relocated.is_file())
            self.assertTrue(copied_excel.is_file())
            self.assertEqual(data_rows[0]["本地文件"], str(relocated))
            self.assertEqual(data_rows[0]["处理动作"], "原素材兜底导出")
            self.assertIn("正式图包", data_rows[0]["规则告警"])
            self.assertFalse(raw_file.exists())
            self.assertFalse(runtime_dir.exists())

            workbook = load_workbook(copied_excel, read_only=True, data_only=True)
            sheet = workbook.active
            self.assertEqual(sheet.cell(row=2, column=7).value, "原素材兜底导出")
            self.assertEqual(sheet.cell(row=2, column=9).value, str(relocated))
            self.assertIn("正式图包", sheet.cell(row=2, column=11).value)
            workbook.close()

    def test_finalize_shoe_outputs_uses_cached_download_rows_when_all_styles_skipped(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            export_dir = base / "downloads"
            runtime_dir.mkdir()

            raw_file = runtime_dir / "cached-download.jpg"
            raw_file.write_bytes(b"raw-shoe")
            exported = base / "深绘鞋品上新图包整理结果.xlsx"
            data_rows = [{
                "输入款号": "204426146036",
                "颜色": "",
                "原文件名": "",
                "云盘路径": "",
                "规则槽位": "整款",
                "输出文件名": "",
                "处理动作": "失败款跳过",
                "下载结果": "已跳过",
                "本地文件": "",
                "压缩结果": "",
                "规则告警": "204426146036 整理失败，已跳过该款",
                "品类来源": "",
                "备注": "鞋品姿势识别多模型均失败",
            }]
            cached_rows = [{
                "输入款号": "204426146036",
                "颜色": "00317",
                "原文件名": "cached-original.jpg",
                "云盘路径": "鞋品/204426146036/00317/cached-original.jpg",
                "规则槽位": "",
                "输出文件名": "",
                "处理动作": "下载后识别鞋品姿势",
                "下载结果": "已下载",
                "本地文件": str(raw_file),
                "压缩结果": "",
                "规则告警": "",
                "品类来源": "",
                "备注": "",
                "__shenhui_group_code": "204426146036",
                "__shoe_color_code": "00317",
                "__shoe_original_filename": "cached-original.jpg",
            }]
            columns = [
                "输入款号",
                "颜色",
                "原文件名",
                "云盘路径",
                "规则槽位",
                "输出文件名",
                "处理动作",
                "下载结果",
                "本地文件",
                "压缩结果",
                "规则告警",
                "品类来源",
                "备注",
            ]
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(columns)
            sheet.append([data_rows[0].get(column, "") for column in columns])
            workbook.save(exported)
            workbook.close()

            result = _finalize_shenhui_new_arrival_outputs(
                task_id="prepare_shoe_upload_package",
                data_rows=data_rows,
                runtime_files=[str(raw_file)],
                exported_files=[str(exported)],
                run_params={
                    "export_folder": str(export_dir),
                    "__shenhui_shoe_download_rows_for_fallback": cached_rows,
                },
                runtime_artifact_dir=str(runtime_dir),
                log=lambda _: None,
            )

            fallback_dir = next(Path(path) for path in result if Path(path).is_dir())
            relocated = fallback_dir / "204426146036" / "00317" / "cached-original.jpg"
            self.assertTrue(relocated.is_file())
            self.assertEqual(len(data_rows), 2)
            self.assertEqual(data_rows[1]["本地文件"], str(relocated))
            self.assertEqual(data_rows[1]["处理动作"], "原素材兜底导出")

            workbook = load_workbook(
                export_dir / "深绘鞋品上新图包整理结果.xlsx",
                read_only=True,
                data_only=True,
            )
            sheet = workbook.active
            self.assertEqual(sheet.cell(row=2, column=7).value, "失败款跳过")
            self.assertEqual(sheet.cell(row=3, column=1).value, "204426146036")
            self.assertEqual(sheet.cell(row=3, column=2).value, "00317")
            self.assertEqual(sheet.cell(row=3, column=3).value, "cached-original.jpg")
            self.assertEqual(sheet.cell(row=3, column=7).value, "原素材兜底导出")
            self.assertEqual(sheet.cell(row=3, column=9).value, str(relocated))
            workbook.close()

    def test_finalize_shoe_outputs_creates_fallback_excel_for_runtime_files_only(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            export_dir = base / "downloads"
            runtime_dir.mkdir()

            raw_file = runtime_dir / "orphan-downloaded.jpg"
            raw_file.write_bytes(b"raw-shoe")
            data_rows = []

            result = _finalize_shenhui_new_arrival_outputs(
                task_id="prepare_shoe_upload_package",
                data_rows=data_rows,
                runtime_files=[str(raw_file)],
                exported_files=[],
                run_params={"export_folder": str(export_dir)},
                runtime_artifact_dir=str(runtime_dir),
                log=lambda _: None,
            )

            fallback_dirs = [Path(path) for path in result if Path(path).is_dir()]
            fallback_excels = [Path(path) for path in result if Path(path).suffix.lower() == ".xlsx"]
            self.assertEqual(len(fallback_dirs), 1)
            self.assertEqual(len(fallback_excels), 1)
            self.assertEqual(len(list(export_dir.glob("深绘鞋品上新图包整理结果_*.xlsx"))), 1)
            relocated = fallback_dirs[0] / "未分类" / "未识别颜色" / "orphan-downloaded.jpg"
            self.assertTrue(relocated.is_file())
            self.assertTrue(fallback_excels[0].is_file())
            self.assertEqual(len(data_rows), 1)
            self.assertEqual(data_rows[0]["本地文件"], str(relocated))
            self.assertFalse(raw_file.exists())
            self.assertFalse(runtime_dir.exists())

            workbook = load_workbook(fallback_excels[0], read_only=True, data_only=True)
            sheet = workbook.active
            self.assertEqual(sheet.cell(row=2, column=7).value, "原素材兜底导出")
            self.assertEqual(sheet.cell(row=2, column=9).value, str(relocated))
            self.assertIn("任务停止", sheet.cell(row=2, column=11).value)
            workbook.close()

    def test_finalize_shoe_outputs_compresses_when_style_package_exceeds_threshold(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            package_dir = runtime_dir / "shoe-packages" / "204326141005"
            color_dir = package_dir / "1.白紫色调00317"
            export_dir = base / "downloads"
            color_dir.mkdir(parents=True)
            export_dir.mkdir()

            first_image = color_dir / "o.jpg"
            second_image = color_dir / "yk1.jpg"
            for path in (first_image, second_image):
                Image.effect_noise((512, 512), 90).convert("RGB").save(
                    path,
                    format="JPEG",
                    quality=100,
                )
            before_sizes = {
                first_image.name: first_image.stat().st_size,
                second_image.name: second_image.stat().st_size,
            }
            package_total = sum(before_sizes.values())

            data_rows = [
                {
                    "输入款号": "204326141005",
                    "颜色": "白紫色调00317",
                    "原文件名": "source-o.jpg",
                    "云盘路径": "鞋品/204326141005/source-o.jpg",
                    "规则槽位": "o",
                    "输出文件名": "1.白紫色调00317/o.jpg",
                    "处理动作": "已选图并按鞋品规则命名",
                    "下载结果": "已下载",
                    "本地文件": str(first_image),
                    "压缩结果": "",
                    "规则告警": "",
                    "品类来源": "Excel指定",
                    "备注": "",
                },
                {
                    "输入款号": "204326141005",
                    "颜色": "白紫色调00317",
                    "原文件名": "source-yk1.jpg",
                    "云盘路径": "鞋品/204326141005/source-yk1.jpg",
                    "规则槽位": "yk",
                    "输出文件名": "1.白紫色调00317/yk1.jpg",
                    "处理动作": "已选图并按鞋品规则命名",
                    "下载结果": "已下载",
                    "本地文件": str(second_image),
                    "压缩结果": "",
                    "规则告警": "",
                    "品类来源": "Excel指定",
                    "备注": "",
                },
            ]
            columns = [
                "输入款号",
                "颜色",
                "原文件名",
                "云盘路径",
                "规则槽位",
                "输出文件名",
                "处理动作",
                "下载结果",
                "本地文件",
                "压缩结果",
                "规则告警",
                "品类来源",
                "备注",
            ]
            exported = base / "深绘鞋品上新图包整理结果.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(columns)
            for row in data_rows:
                sheet.append([row.get(column, "") for column in columns])
            workbook.save(exported)
            workbook.close()

            with patch(
                "core.api_server._SHENHUI_NEW_ARRIVAL_SINGLE_IMAGE_THRESHOLD_BYTES",
                package_total * 10,
            ), patch(
                "core.api_server._SHENHUI_NEW_ARRIVAL_STYLE_PACKAGE_THRESHOLD_BYTES",
                max(1, package_total - 1),
            ):
                _finalize_shenhui_new_arrival_outputs(
                    task_id="prepare_shoe_upload_package",
                    data_rows=data_rows,
                    runtime_files=[],
                    exported_files=[str(exported)],
                    run_params={
                        "export_folder": str(export_dir),
                        "__shenhui_shoe_package_refs": [str(package_dir)],
                    },
                    runtime_artifact_dir=str(runtime_dir),
                    log=lambda _: None,
                )

            output_package = export_dir / "204326141005"
            compressed_rows = [
                row for row in data_rows if "单款总图包超过" in row.get("压缩结果", "")
            ]
            self.assertTrue(compressed_rows)
            for row in compressed_rows:
                final_image = Path(row["本地文件"])
                self.assertTrue(final_image.is_file())
                self.assertLess(final_image.stat().st_size, before_sizes[final_image.name])
                with Image.open(final_image) as compressed:
                    self.assertEqual(compressed.size, (512, 512))
            self.assertTrue(output_package.is_dir())

            workbook = load_workbook(
                export_dir / "深绘鞋品上新图包整理结果.xlsx",
                read_only=True,
                data_only=True,
            )
            compression_values = [
                workbook.active.cell(row=row_index, column=10).value or ""
                for row_index in range(2, 4)
            ]
            self.assertTrue(
                any("单款总图包超过" in value for value in compression_values)
            )
            workbook.close()

    def test_prepare_shoe_rows_registers_generated_package_before_excel_export(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            runtime_dir = Path(tmpdir) / "runtime"
            package_root = runtime_dir / "shoe-packages" / "204326141005"
            package_root.mkdir(parents=True)
            run_params = {
                "model_id": "qwen3.8-max-preview",
                "fallback_model_1": "gpt-5.6-terra",
                "fallback_model_2": "gpt-5.6-luna",
                "fallback_model_3": "",
                "shoe_pose_strategy": "global_pages",
                "shoe_category_file": {
                    "rows": [
                        {"款号": "208326146209", "品类": "宝宝鞋"},
                        {"款号": "204325141014", "品类": "公主鞋"},
                    ]
                },
            }
            expected_rows = [{
                "输入款号": "204326141005",
                "规则槽位": "o",
                "规则告警": "",
            }]
            progress_events = []
            progress_callback = progress_events.append

            with patch(
                "core.api_server.shenhui_shoe_packaging.prepare_shoe_packages",
                return_value=(expected_rows, {"204326141005": package_root}),
            ) as prepare:
                result = _prepare_shenhui_shoe_package_rows(
                    data_rows=[{"输入款号": "204326141005"}],
                    run_params=run_params,
                    runtime_artifact_dir=str(runtime_dir),
                    log=lambda _: None,
                    progress=progress_callback,
                )

            self.assertEqual(result, expected_rows)
            self.assertEqual(
                run_params["__shenhui_shoe_package_refs"],
                [str(package_root)],
            )
            self.assertEqual(
                prepare.call_args.kwargs["model_id"],
                "qwen3.8-max-preview",
            )
            self.assertEqual(
                prepare.call_args.kwargs["fallback_model_ids"],
                ["gpt-5.6-terra", "gpt-5.6-luna"],
            )
            self.assertEqual(
                prepare.call_args.kwargs["pose_strategy"],
                "global_pages",
            )
            self.assertEqual(
                prepare.call_args.kwargs["label_fallback_model_ids"],
                ["gpt-5.6-terra", "gpt-5.6-luna"],
            )
            self.assertEqual(
                prepare.call_args.kwargs["shoe_categories"],
                {
                    "208326146209": "婴童",
                    "204325141014": "休闲",
                },
            )
            self.assertIs(
                prepare.call_args.kwargs["progress"],
                progress_callback,
            )

            run_params["__shoe_pose_benchmark"] = True
            with patch(
                "core.api_server.shenhui_shoe_packaging.prepare_shoe_packages",
                return_value=(expected_rows, {"204326141005": package_root}),
            ) as benchmark_prepare:
                _prepare_shenhui_shoe_package_rows(
                    data_rows=[{"输入款号": "204326141005"}],
                    run_params=run_params,
                    runtime_artifact_dir=str(runtime_dir),
                    log=lambda _: None,
                )
            self.assertIs(
                benchmark_prepare.call_args.kwargs["analyze_color_label"],
                False,
            )

    def test_prepare_shoe_rows_allows_code_only_rows_to_use_model_category(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            runtime_dir = Path(tmpdir) / "runtime"
            package_root = runtime_dir / "shoe-packages" / "204426146036"
            package_root.mkdir(parents=True)
            run_params = {
                "shoe_category_file": {
                    "rows": [
                        {"款号": "204426146036", "品类": ""},
                        {"款号": "204426146127"},
                    ]
                },
            }
            expected_rows = [{
                "输入款号": "204426146036",
                "规则槽位": "tmz5",
                "规则告警": "",
            }]
            logs = []

            with patch(
                "core.api_server.shenhui_shoe_packaging.prepare_shoe_packages_skip_failed_styles",
                return_value=(expected_rows, {"204426146036": package_root}),
            ) as prepare:
                result = _prepare_shenhui_shoe_package_rows(
                    data_rows=[{"输入款号": "204426146036"}],
                    run_params=run_params,
                    runtime_artifact_dir=str(runtime_dir),
                    log=logs.append,
                )

            self.assertEqual(result, expected_rows)
            self.assertIsNone(prepare.call_args.kwargs["shoe_categories"])
            self.assertTrue(any("模型兜底识别品类" in item for item in logs))

    def test_finalize_outputs_creates_style_zips_when_auto_zip_enabled(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            runtime_dir.mkdir()

            model_file = runtime_dir / "runtime-model.jpg"
            yq_file = runtime_dir / "runtime-yq.jpg"
            model_file.write_bytes(b"model")
            yq_file.write_bytes(b"yq")

            exported = base / "summary.xlsx"
            exported.write_bytes(b"excel")

            result = _finalize_shenhui_new_arrival_outputs(
                task_id="prepare_upload_package",
                data_rows=[
                    {
                        "输入款号": "208226103201",
                        "输入编码": "208226103201",
                        "素材来源": "模特图",
                        "文件名": "balaBR05106-72904_P.jpg",
                        "下载结果": "已下载",
                        "本地文件": str(model_file),
                        "__shenhui_group_code": "208226103201",
                        "__shenhui_asset_role": "image",
                        "__package_filename": "balaBR05106-72904_P.jpg",
                    },
                    {
                        "输入款号": "208226103201",
                        "输入编码": "208226103201",
                        "素材来源": "静物图",
                        "文件名": "yq.jpg",
                        "下载结果": "已下载",
                        "本地文件": str(yq_file),
                        "__shenhui_group_code": "208226103201",
                        "__shenhui_asset_role": "image",
                        "__package_filename": "yq.jpg",
                    },
                ],
                runtime_files=[str(model_file), str(yq_file)],
                exported_files=[str(exported)],
                run_params={"package_name": "深绘测试图包", "auto_zip_package": ["yes"]},
                runtime_artifact_dir=str(runtime_dir),
                log=lambda _: None,
            )

            self.assertEqual(len(result), 2)
            style_zip_path = Path(result[0])
            self.assertTrue(style_zip_path.is_file())
            self.assertEqual(style_zip_path.name, "208226103201.zip")
            self.assertEqual(Path(result[1]), exported)

            with zipfile.ZipFile(style_zip_path) as archive:
                names = archive.namelist()
                self.assertIn("balaBR05106-72904_P.jpg", names)
                self.assertIn("yq.jpg", names)

            self.assertFalse(model_file.exists())
            self.assertFalse(yq_file.exists())
            self.assertFalse((runtime_dir / "深绘测试图包").exists())

    def test_finalize_outputs_removes_runtime_intermediates_after_copying_to_export_folder(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            export_dir = base / "exports"
            runtime_dir.mkdir()

            model_file = runtime_dir / "runtime-model.jpg"
            model_file.write_bytes(b"model")
            exported = base / "summary.xlsx"
            exported.write_bytes(b"excel")

            result = _finalize_shenhui_new_arrival_outputs(
                task_id="prepare_upload_package",
                data_rows=[{
                    "输入款号": "208226103201",
                    "输入编码": "208226103201",
                    "素材来源": "模特图",
                    "文件名": "balaBR05106-72904_P.jpg",
                    "下载结果": "已下载",
                    "本地文件": str(model_file),
                    "__shenhui_group_code": "208226103201",
                    "__shenhui_asset_role": "image",
                    "__package_filename": "balaBR05106-72904_P.jpg",
                }],
                runtime_files=[str(model_file)],
                exported_files=[str(exported)],
                run_params={
                    "package_name": "深绘测试图包",
                    "export_folder": str(export_dir),
                    "auto_zip_package": True,
                },
                runtime_artifact_dir=str(runtime_dir),
                log=lambda _: None,
            )

            self.assertTrue(all(Path(path).is_file() for path in result))
            self.assertEqual({Path(path).parent for path in result}, {export_dir})
            self.assertFalse(runtime_dir.exists())

    def test_finalize_outputs_writes_default_zips_next_to_excel_and_cleans_runtime(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            output_dir = base / "outputs"
            runtime_dir.mkdir()
            output_dir.mkdir()

            model_file = runtime_dir / "runtime-model.jpg"
            model_file.write_bytes(b"model")
            exported = output_dir / "summary.xlsx"
            exported.write_bytes(b"excel")

            result = _finalize_shenhui_new_arrival_outputs(
                task_id="prepare_upload_package",
                data_rows=[{
                    "输入款号": "208226103201",
                    "输入编码": "208226103201",
                    "素材来源": "模特图",
                    "文件名": "balaBR05106-72904_P.jpg",
                    "下载结果": "已下载",
                    "本地文件": str(model_file),
                    "__shenhui_group_code": "208226103201",
                    "__shenhui_asset_role": "image",
                    "__package_filename": "balaBR05106-72904_P.jpg",
                }],
                runtime_files=[str(model_file)],
                exported_files=[str(exported)],
                run_params={"package_name": "深绘测试图包"},
                runtime_artifact_dir=str(runtime_dir),
                log=lambda _: None,
            )

            self.assertEqual(len(result), 2)
            self.assertEqual(Path(result[0]).parent, output_dir)
            self.assertTrue(Path(result[0]).is_dir())
            self.assertEqual(Path(result[1]), exported)
            self.assertFalse(runtime_dir.exists())

    def test_finalize_outputs_cleans_stopped_runtime_download_artifacts(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            runtime_dir.mkdir()

            downloaded = runtime_dir / "downloaded.jpg"
            partial = runtime_dir / "downloaded-2.jpg.part"
            downloaded.write_bytes(b"downloaded")
            partial.write_bytes(b"partial")
            exported = base / "partial.xlsx"
            exported.write_bytes(b"excel")

            result = _finalize_shenhui_new_arrival_outputs(
                task_id="prepare_upload_package",
                data_rows=[],
                runtime_files=[str(downloaded)],
                exported_files=[str(exported)],
                run_params={"package_name": "深绘测试图包"},
                runtime_artifact_dir=str(runtime_dir),
                log=lambda _: None,
            )

            self.assertEqual(result, [str(exported)])
            self.assertFalse(runtime_dir.exists())

    def test_orphaned_active_run_cleanup_removes_runtime_download_artifacts(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "data" / "shenhui-new-arrival" / "prepare_upload_package" / "runtime" / "123"
            runtime_dir.mkdir(parents=True)

            downloaded = runtime_dir / "downloaded.jpg"
            partial = runtime_dir / "downloaded-2.jpg.part"
            downloaded.write_bytes(b"downloaded")
            partial.write_bytes(b"partial")

            with patch("core.data_sink.artifact_dir_path", return_value=runtime_dir):
                _cleanup_orphaned_runtime_artifacts([{
                    "id": 123,
                    "adapter_id": "shenhui-new-arrival",
                    "task_id": "prepare_upload_package",
                }])

            self.assertFalse(runtime_dir.exists())

    def test_finalize_pdf_batch_screenshot_outputs_cropped_zip(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            runtime_dir.mkdir()

            pdf_file = base / "208226103201-label.pdf"
            rendered = base / "rendered-yq.png"
            exported = base / "summary.xlsx"
            pdf_file.write_bytes(b"%PDF-fake")
            rendered.write_bytes(b"png")
            exported.write_bytes(b"excel")

            with patch(
                "core.shenhui_pdf_screenshot.convert_pdf_to_yq_images",
                return_value=[rendered],
            ), patch(
                "core.shenhui_pdf_screenshot.extract_pdf_text",
                return_value="",
            ):
                result = _finalize_shenhui_new_arrival_outputs(
                    task_id="pdf_batch_screenshot",
                    data_rows=[{
                        "PDF文件": "208226103201-label.pdf",
                        "原始路径": str(pdf_file),
                        "__pdf_path": str(pdf_file),
                    }],
                    runtime_files=[],
                    exported_files=[str(exported)],
                    run_params={"package_name": "PDF截图测试"},
                    runtime_artifact_dir=str(runtime_dir),
                    log=lambda _: None,
                )

            self.assertEqual(len(result), 2)
            zip_path = Path(result[0])
            self.assertTrue(zip_path.is_file())
            self.assertEqual(Path(result[1]), exported)

            with zipfile.ZipFile(zip_path) as archive:
                names = archive.namelist()
                self.assertTrue(any(name.endswith("PDF截图测试/208226103201/yq(1).png") for name in names))

    def test_finalize_prepare_upload_package_crops_downloaded_pdf_into_style_zip(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            runtime_dir.mkdir()

            model_file = runtime_dir / "runtime-model.jpg"
            pdf_file = runtime_dir / "runtime-tag.pdf"
            exported = base / "summary.xlsx"
            model_file.write_bytes(b"model")
            pdf_file.write_bytes(b"%PDF-fake")
            exported.write_bytes(b"excel")

            def fake_processor(**kwargs):
                output = kwargs["package_root"] / "208226103201" / "yq1.jpg"
                output.parent.mkdir(parents=True, exist_ok=True)
                Image.new("RGB", (800, 800), "white").save(output)
                kwargs["pdf_rows"][0][0]["处理动作"] = "AI 识别裁图完成"
                return ApparelLabelProcessingResult(1, (), (), (), ())

            rows = [
                {
                    "输入款号": "208226103201",
                    "输入编码": "208226103201",
                    "素材来源": "模特图",
                    "文件名": "balaBR05106-72904_P.jpg",
                    "下载结果": "已下载",
                    "本地文件": str(model_file),
                    "__shenhui_group_code": "208226103201",
                    "__shenhui_asset_role": "image",
                    "__package_filename": "balaBR05106-72904_P.jpg",
                },
                {
                    "输入款号": "208226103201",
                    "输入编码": "208226103201",
                    "素材来源": "静物图",
                    "文件名": "208226103201吊牌.pdf",
                    "下载结果": "已下载",
                    "本地文件": str(pdf_file),
                    "__shenhui_group_code": "208226103201",
                    "__shenhui_asset_role": "pdf_yq",
                    "__package_filename": "208226103201吊牌.pdf",
                    "__pdf_path": str(pdf_file),
                    "__pdf_type": "hang_tag",
                    "__style_code": "208226103201",
                },
            ]

            with patch(
                "core.api_server.process_prepare_upload_package_labels",
                side_effect=fake_processor,
            ):
                result = _finalize_shenhui_new_arrival_outputs(
                    task_id="prepare_upload_package",
                    data_rows=rows,
                    runtime_files=[str(model_file), str(pdf_file)],
                    exported_files=[str(exported)],
                    run_params={
                        "package_name": "深绘测试图包",
                        "tag_crop_boxes": '[{"x":0.01,"y":0.2,"width":0.2,"height":0.5}]',
                        "auto_zip_package": True,
                    },
                    runtime_artifact_dir=str(runtime_dir),
                    log=lambda _: None,
                )

            style_zip_path = Path(result[0])
            self.assertTrue(style_zip_path.is_file())
            with zipfile.ZipFile(style_zip_path) as archive:
                names = archive.namelist()
                self.assertIn("balaBR05106-72904_P.jpg", names)
                self.assertIn("yq1.jpg", names)
                self.assertNotIn("_PDF待裁图/runtime-tag.pdf", names)
            self.assertEqual(rows[1]["本地文件"], "")
            self.assertEqual(rows[1]["最终裁图"], "")

    def test_finalize_prepare_upload_package_rewrites_generated_crop_paths_in_exported_workbook(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            export_dir = base / "downloads"
            runtime_dir.mkdir()
            export_dir.mkdir()

            pdf_file = runtime_dir / "runtime-tag.pdf"
            pdf_file.write_bytes(b"%PDF-fake")
            external_crop = base / "outside" / "external-yq1.jpg"
            external_crop.parent.mkdir()
            Image.new("RGB", (800, 800), "white").save(external_crop)
            exported = base / "summary.xlsx"
            columns = [
                "输入款号",
                "文件名",
                "下载结果",
                "本地文件",
                "处理动作",
                "最终裁图",
            ]
            data_rows = [
                {
                    "输入款号": "208226103201",
                    "输入编码": "208226103201",
                    "文件名": "208226103201吊牌.pdf",
                    "下载结果": "已下载",
                    "本地文件": str(pdf_file),
                    "处理动作": "",
                    "最终裁图": "",
                    "__shenhui_group_code": "208226103201",
                    "__shenhui_asset_role": "pdf_yq",
                    "__package_filename": "208226103201吊牌.pdf",
                    "__pdf_type": "hang_tag",
                    "__style_code": "208226103201",
                },
                {
                    "输入款号": "999999999999",
                    "文件名": "external-yq1.jpg",
                    "下载结果": "已跳过",
                    "本地文件": str(external_crop),
                    "处理动作": "外部素材保留",
                    "最终裁图": str(external_crop),
                },
            ]
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(columns)
            for row in data_rows:
                sheet.append([row.get(column, "") for column in columns])
            workbook.save(exported)
            workbook.close()

            def fake_processor(**kwargs):
                output = kwargs["package_root"] / "208226103201" / "yq1.jpg"
                output.parent.mkdir(parents=True, exist_ok=True)
                Image.new("RGB", (800, 800), "white").save(output)
                generated_row = kwargs["data_rows"][0]
                generated_row["文件名"] = "yq1.jpg"
                generated_row["处理动作"] = "AI 识别裁图完成"
                generated_row["本地文件"] = str(output)
                generated_row["最终裁图"] = str(output)
                return ApparelLabelProcessingResult(1, (), (), (), ())

            with patch(
                "core.api_server.process_prepare_upload_package_labels",
                side_effect=fake_processor,
            ):
                result = _finalize_shenhui_new_arrival_outputs(
                    task_id="prepare_upload_package",
                    data_rows=data_rows,
                    runtime_files=[str(pdf_file)],
                    exported_files=[str(exported)],
                    run_params={
                        "package_name": "深绘测试图包",
                        "export_folder": str(export_dir),
                    },
                    runtime_artifact_dir=str(runtime_dir),
                    log=lambda _: None,
                )

            package_dir = export_dir / "深绘测试图包"
            final_crop = package_dir / "208226103201" / "yq1.jpg"
            copied_workbook = export_dir / "summary.xlsx"
            self.assertIn(str(package_dir), result)
            self.assertTrue(final_crop.is_file())
            self.assertEqual(data_rows[0]["本地文件"], str(final_crop))
            self.assertEqual(data_rows[0]["最终裁图"], str(final_crop))
            self.assertEqual(data_rows[1]["本地文件"], str(external_crop))
            self.assertEqual(data_rows[1]["最终裁图"], str(external_crop))
            self.assertFalse(runtime_dir.exists())

            workbook = load_workbook(copied_workbook, read_only=True, data_only=True)
            sheet = workbook.active
            self.assertEqual(sheet.cell(row=2, column=2).value, "yq1.jpg")
            self.assertEqual(sheet.cell(row=2, column=4).value, str(final_crop))
            self.assertEqual(sheet.cell(row=2, column=6).value, str(final_crop))
            self.assertEqual(sheet.cell(row=3, column=4).value, str(external_crop))
            self.assertEqual(sheet.cell(row=3, column=6).value, str(external_crop))
            workbook.close()

    def test_finalize_prepare_upload_package_clears_zip_packaged_crop_paths_in_exported_workbook(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            export_dir = base / "downloads"
            runtime_dir.mkdir()
            export_dir.mkdir()

            pdf_file = runtime_dir / "runtime-tag.pdf"
            pdf_file.write_bytes(b"%PDF-fake")
            external_crop = base / "outside" / "external-yq1.jpg"
            external_crop.parent.mkdir()
            Image.new("RGB", (800, 800), "white").save(external_crop)
            exported = base / "summary.xlsx"
            columns = [
                "输入款号",
                "文件名",
                "下载结果",
                "本地文件",
                "处理动作",
                "最终裁图",
                "备注",
            ]
            data_rows = [
                {
                    "输入款号": "208226103201",
                    "输入编码": "208226103201",
                    "文件名": "208226103201吊牌.pdf",
                    "下载结果": "已下载",
                    "本地文件": str(pdf_file),
                    "处理动作": "",
                    "最终裁图": "",
                    "备注": "原备注",
                    "__shenhui_group_code": "208226103201",
                    "__shenhui_asset_role": "pdf_yq",
                    "__package_filename": "208226103201吊牌.pdf",
                    "__pdf_type": "hang_tag",
                    "__style_code": "208226103201",
                },
                {
                    "输入款号": "999999999999",
                    "文件名": "external-yq1.jpg",
                    "下载结果": "已跳过",
                    "本地文件": str(external_crop),
                    "处理动作": "外部素材保留",
                    "最终裁图": str(external_crop),
                    "备注": "外部备注",
                },
            ]
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(columns)
            for row in data_rows:
                sheet.append([row.get(column, "") for column in columns])
            workbook.save(exported)
            workbook.close()
            runtime_crop_paths = []

            def fake_processor(**kwargs):
                output = kwargs["package_root"] / "208226103201" / "yq1.jpg"
                output.parent.mkdir(parents=True, exist_ok=True)
                Image.new("RGB", (800, 800), "white").save(output)
                runtime_crop_paths.append(str(output))
                generated_row = kwargs["data_rows"][0]
                generated_row["处理动作"] = "AI 识别裁图完成"
                generated_row["本地文件"] = str(output)
                generated_row["最终裁图"] = str(output)
                return ApparelLabelProcessingResult(1, (), (), (), ())

            with patch(
                "core.api_server.process_prepare_upload_package_labels",
                side_effect=fake_processor,
            ):
                result = _finalize_shenhui_new_arrival_outputs(
                    task_id="prepare_upload_package",
                    data_rows=data_rows,
                    runtime_files=[str(pdf_file)],
                    exported_files=[str(exported)],
                    run_params={
                        "package_name": "深绘测试图包",
                        "export_folder": str(export_dir),
                        "auto_zip_package": True,
                    },
                    runtime_artifact_dir=str(runtime_dir),
                    log=lambda _: None,
                )

            zip_path = export_dir / "208226103201.zip"
            copied_workbook = export_dir / "summary.xlsx"
            self.assertIn(str(zip_path), result)
            self.assertTrue(zip_path.is_file())
            self.assertFalse(Path(runtime_crop_paths[0]).exists())
            self.assertEqual(data_rows[0]["最终裁图"], "")
            self.assertIn("原备注", data_rows[0]["备注"])
            self.assertIn("最终裁图已打包至款号 ZIP", data_rows[0]["备注"])
            self.assertIn("208226103201.zip", data_rows[0]["备注"])
            self.assertEqual(data_rows[1]["最终裁图"], str(external_crop))
            self.assertEqual(data_rows[1]["备注"], "外部备注")

            workbook = load_workbook(copied_workbook, read_only=True, data_only=True)
            sheet = workbook.active
            self.assertNotEqual(sheet.cell(row=2, column=6).value, runtime_crop_paths[0])
            self.assertIsNone(sheet.cell(row=2, column=6).value)
            self.assertIn("最终裁图已打包至款号 ZIP", sheet.cell(row=2, column=7).value)
            self.assertIn("208226103201.zip", sheet.cell(row=2, column=7).value)
            self.assertEqual(sheet.cell(row=3, column=6).value, str(external_crop))
            self.assertEqual(sheet.cell(row=3, column=7).value, "外部备注")
            workbook.close()

    def test_finalize_prepare_upload_package_uses_short_pdf_work_dir(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            runtime_dir.mkdir()

            pdf_file = runtime_dir / "runtime-tag.pdf"
            exported = base / "summary.xlsx"
            pdf_file.write_bytes(b"%PDF-fake")
            exported.write_bytes(b"excel")
            work_dirs = []

            def fake_processor(**kwargs):
                work_dirs.append(Path(kwargs["work_dir"]))
                return ApparelLabelProcessingResult(0, (), (), (), ())

            with patch(
                "core.api_server.process_prepare_upload_package_labels",
                side_effect=fake_processor,
            ):
                _finalize_shenhui_new_arrival_outputs(
                    task_id="prepare_upload_package",
                    data_rows=[{
                        "输入款号": "208226133212",
                        "输入编码": "208226133212",
                        "素材来源": "静物图",
                        "文件名": "208226133212__still__f2e2b3449a6c550fa3e6a9fab8c3bfcb37d1466b__208226133212吊牌.pdf",
                        "下载结果": "已下载",
                        "本地文件": str(pdf_file),
                        "__shenhui_group_code": "208226133212",
                        "__shenhui_asset_role": "pdf_yq",
                        "__package_filename": "208226133212吊牌.pdf",
                        "__pdf_path": str(pdf_file),
                        "__pdf_type": "hang_tag",
                        "__style_code": "208226133212",
                    }],
                    runtime_files=[str(pdf_file)],
                    exported_files=[str(exported)],
                    run_params={
                        "package_name": "深绘上新图包_20260609_174655",
                        "auto_zip_package": True,
                    },
                    runtime_artifact_dir=str(runtime_dir),
                    log=lambda _: None,
                )

            self.assertEqual(work_dirs, [runtime_dir / "_pdf_work"])

    def test_finalize_prepare_upload_package_preserves_pdf_when_screenshot_unavailable(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            runtime_dir = base / "runtime"
            runtime_dir.mkdir()

            pdf_file = runtime_dir / "runtime-tag.pdf"
            exported = base / "summary.xlsx"
            pdf_file.write_bytes(b"%PDF-fake")
            exported.write_bytes(b"excel")

            with patch(
                "core.shenhui_pdf_screenshot.convert_pdf_to_yq_images",
                return_value=[],
            ), patch(
                "core.shenhui_pdf_screenshot.extract_pdf_text",
                return_value="",
            ):
                result = _finalize_shenhui_new_arrival_outputs(
                    task_id="prepare_upload_package",
                    data_rows=[{
                        "输入款号": "208226103201",
                        "输入编码": "208226103201",
                        "素材来源": "静物图",
                        "文件名": "208226103201吊牌.pdf",
                        "下载结果": "已下载",
                        "本地文件": str(pdf_file),
                        "__shenhui_group_code": "208226103201",
                        "__shenhui_asset_role": "pdf_yq",
                        "__package_filename": "208226103201吊牌.pdf",
                        "__pdf_path": str(pdf_file),
                        "__pdf_type": "hang_tag",
                        "__style_code": "208226103201",
                    }],
                    runtime_files=[str(pdf_file)],
                    exported_files=[str(exported)],
                    run_params={"package_name": "深绘测试图包"},
                    runtime_artifact_dir=str(runtime_dir),
                    log=lambda _: None,
                )

            package_dir = Path(result[0])
            preserved_pdf = package_dir / "208226103201" / "_PDF待裁图" / "208226103201吊牌.pdf"
            self.assertTrue(preserved_pdf.is_file())
            self.assertEqual(Path(result[1]), exported)
            self.assertFalse(pdf_file.exists())


if __name__ == "__main__":
    unittest.main()
